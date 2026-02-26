import json
import re
import traceback
import unicodedata
from datetime import datetime
from decimal import Decimal, InvalidOperation
from uuid import uuid4

import openpyxl
import pdfplumber
import xlwt
from dateutil import parser
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (Case, CharField, Count, IntegerField, Prefetch,
                              Q, Sum, Value, When)
from django.db.models.functions import Cast
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.timezone import is_aware
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from facturacion.forms import OrdenCompraFacturacionForm
from facturacion.models import (CartolaMovimiento, FacturaOC,
                                OrdenCompraFacturacion)
from operaciones.forms import MovimientoUsuarioForm
from operaciones.models import MonthlyPayment, ServicioCotizado
from usuarios.decoradores import rol_requerido

from .forms import (CartolaAbonoForm, CartolaGastoForm,
                    CartolaMovimientoCompletoForm, FacturaOCForm,
                    ImportarFacturasForm, ProyectoForm, TipoGastoForm)
from .models import Proyecto, TipoGasto

User = get_user_model()

@login_required
@rol_requerido('facturacion', 'admin')
def listar_ordenes_compra(request):
    q = request.GET.copy()

    # --- Filtros de texto (raw para mantener DU con prefijo en inputs) ---
    du_raw = (q.get('du') or '')
    du = du_raw.strip().upper().replace('DU', '') if du_raw else ''
    id_claro = (q.get('id_claro') or '').strip()
    id_new = (q.get('id_new') or '').strip()
    mes_produccion = (q.get('mes_produccion') or '').strip()
    estado = (q.get('estado') or '').strip()

    # IDs de OC ya facturadas
    oc_facturadas_ids = list(
        FacturaOC.objects.values_list('orden_compra_id', flat=True)
    )

    # Prefetch SOLO OC no facturadas (se respeta en .ordenes_compra.all())
    oc_no_facturadas_qs = OrdenCompraFacturacion.objects.exclude(
        pk__in=oc_facturadas_ids
    )

    servicios = (
        ServicioCotizado.objects
        .select_related(
            'pm_aprueba', 'tecnico_aceptado', 'tecnico_finalizo',
            'supervisor_aprobo', 'supervisor_rechazo', 'supervisor_asigna',
            'usuario_informe'
        )
        .prefetch_related(
            Prefetch('ordenes_compra', queryset=oc_no_facturadas_qs),
            'trabajadores_asignados'
        )
        .order_by('-fecha_creacion')
    )

    # 🚫 NO mostrar bonos / adelantos / descuentos en este listado
    servicios = servicios.exclude(
        estado__in=['ajuste_bono', 'ajuste_adelanto', 'ajuste_descuento']
    )

    # ✅ contadores para decidir correctamente cuándo mostrar “Sin orden de compra”
    servicios = servicios.annotate(
        oc_total=Count('ordenes_compra', distinct=True),
        oc_sin_fact=Count(
            'ordenes_compra',
            filter=~Q(ordenes_compra__id__in=oc_facturadas_ids),
            distinct=True
        )
    )

    # 🔎 Filtrado por estado (si el usuario lo elige)
    if estado:
        servicios = servicios.filter(estado=estado)

    # 🔎 Filtros de texto
    if du:
        servicios = servicios.filter(du__iexact=du)
    if id_claro:
        servicios = servicios.filter(id_claro__icontains=id_claro)
    if id_new:
        servicios = servicios.filter(id_new__icontains=id_new)
    if mes_produccion:
        servicios = servicios.filter(mes_produccion__icontains=mes_produccion)

    # 🔎 **Filtro clave ANTES de paginar**:
    # Solo paginamos lo que el tbody realmente muestra:
    # - servicios con OC no facturada (oc_sin_fact > 0)
    # - servicios que nunca tuvieron OC (oc_total == 0)
    servicios = servicios.filter(Q(oc_sin_fact__gt=0) | Q(oc_total=0))

    # --- Paginación (máx. 100) ---
    cantidad = (q.get("cantidad") or "10").strip().lower()

    if cantidad == "todos":
        # "todos" = mostrar hasta 100
        per_page = 100
    else:
        try:
            # mínimo 5, máximo 100
            per_page = max(5, min(int(cantidad), 100))
        except ValueError:
            per_page = 10
            cantidad = "10"

    pagina = Paginator(servicios, per_page).get_page(q.get("page"))

    # === MonthlyPayment (como ya lo tenías)
    visibles = list(pagina.object_list)

    def _pick_tech_id(s):
        return s.tecnico_finalizo_id or s.tecnico_aceptado_id

    tech_ids = {tid for tid in (_pick_tech_id(s) for s in visibles) if tid}
    months = {(s.mes_produccion or "")[:7]
              for s in visibles if (s.mes_produccion or "")[:7]}

    mp_map = {}
    if tech_ids and months:
        for mp in MonthlyPayment.objects.filter(
            technician_id__in=tech_ids, month__in=months
        ):
            mp_map[(mp.technician_id, mp.month)] = mp.status

    status_label_map = dict(MonthlyPayment.STATUS)
    for s in visibles:
        tid = _pick_tech_id(s)
        mm = (s.mes_produccion or "")[:7]
        st = mp_map.get((tid, mm))
        s.payment_status = st or ""
        s.payment_status_label = status_label_map.get(st, "")

    # --- Querystrings persistentes ---
    qs_pag = q.copy()
    qs_pag.pop("page", None)
    qs_cant = q.copy()
    qs_cant.pop("cantidad", None)

    context = {
        'pagina': pagina,
        'cantidad': cantidad,
        'filtros': {
            'du': du_raw,
            'id_claro': id_claro,
            'id_new': id_new,
            'mes_produccion': mes_produccion,
            'estado': estado,
        },
        'estado_choices': ServicioCotizado.ESTADOS,
        'monthly_status_choices': status_label_map,
        'qs_pag': qs_pag.urlencode(),
        'qs_cant': qs_cant.urlencode(),
    }
    return render(request, 'facturacion/listar_ordenes_compra.html', context)


# =========================
#  IMPORTAR / PREVIEW OC
# =========================

@login_required
@rol_requerido('facturacion', 'admin')
def importar_orden_compra(request):
    if request.method == 'POST':
        archivos = request.FILES.getlist('archivo_pdf')
        if not archivos:
            messages.error(request, "Selecciona al menos un archivo PDF.")
            return redirect('facturacion:importar_orden_compra')

        datos_extraidos = []
        nombres_archivos = []

        def _norm_line(s: str) -> str:
            s = unicodedata.normalize("NFKC", s or "")
            s = s.replace("–", "-").replace("—", "-")
            return s

        def _extraer_de_pdf(ruta_absoluta: str) -> list[dict]:
            filas = []
            with pdfplumber.open(ruta_absoluta) as pdf:
                for pagina in pdf.pages:
                    texto = pagina.extract_text() or ""
                    lineas = [_norm_line(l).strip()
                              for l in texto.split("\n") if l.strip()]

                    # 1) Número de OC
                    numero_oc = None
                    for idx, linea in enumerate(lineas[:40]):
                        if 'ORDEN DE COMPRA' in linea.upper():
                            for k in range(1, 5):
                                if idx + k < len(lineas):
                                    m = re.search(
                                        r'\b(\d{10})\b', lineas[idx + k])
                                    if m:
                                        numero_oc = m.group(1)
                                        break
                            if numero_oc:
                                break
                    if not numero_oc:
                        for cab in lineas[:60]:
                            m = re.search(r'\b(\d{10})\b', cab)
                            if m:
                                numero_oc = m.group(1)
                                break
                    if not numero_oc:
                        numero_oc = 'NO_ENCONTRADO'

                    # 2) Filas detalle
                    i = 0
                    while i < len(lineas):
                        linea = lineas[i]
                        if re.match(r'^\d+\s+\d+\s+SER', linea):
                            partes = re.split(r'\s{2,}', linea.strip())
                            if len(partes) < 7:
                                partes = linea.split()
                            if len(partes) >= 8:
                                pos = partes[0]
                                cantidad = partes[1]
                                unidad = partes[2]
                                material = partes[3]
                                descripcion = ' '.join(partes[4:-3])
                                fecha_entrega = partes[-3]
                                precio_unitario = partes[-2].replace(',', '.')
                                monto = partes[-1].replace(',', '.')
                                id_new = None
                                if i + 1 < len(lineas):
                                    siguiente = lineas[i + 1]
                                    pat = r'(CL-\d{2}-[A-ZÑ]{2}-\d{5}-\d{2})'
                                    m2 = re.search(
                                        pat, siguiente, flags=re.IGNORECASE)
                                    if m2:
                                        id_new = m2.group(1).upper()

                                filas.append({
                                    'orden_compra': numero_oc,
                                    'pos': pos,
                                    'cantidad': cantidad,
                                    'unidad_medida': unidad,
                                    'material_servicio': material,
                                    'descripcion_sitio': descripcion,
                                    'fecha_entrega': fecha_entrega,
                                    'precio_unitario': precio_unitario,
                                    'monto': monto,
                                    'id_new': id_new,
                                })
                                i += 1
                        i += 1
            return filas

        # Procesar PDFs
        for archivo in archivos:
            nombre_archivo = archivo.name
            nombres_archivos.append(nombre_archivo)
            ruta_temporal = default_storage.save(
                f"temp_oc/{uuid4().hex}_{nombre_archivo}",
                ContentFile(archivo.read())
            )
            ruta_absoluta = default_storage.path(ruta_temporal)
            try:
                datos_extraidos.extend(_extraer_de_pdf(ruta_absoluta))
            finally:
                default_storage.delete(ruta_temporal)

        # ====== Construir filas de PREVIEW con candidatos ======
        filas_preview = []
        ids_no_encontrados = set()

        for idx, fila in enumerate(datos_extraidos):
            idn = (fila.get('id_new') or "").strip().upper()
            candidatos = []
            if idn:
                # 👉 SOLO servicios que NO tienen OC (ya sea facturada o no)
                qs = (
                    ServicioCotizado.objects
                    .filter(id_new=idn, ordenes_compra__isnull=True)
                    .order_by('-du', '-fecha_creacion')
                    .distinct()
                )
                for s in qs:
                    label = f"DU{s.du or ''} — {s.id_claro or '—'} — {s.mes_produccion or '—'}"
                    candidatos.append({'id': s.id, 'label': label})
            else:
                ids_no_encontrados.add("SIN_ID")

            if not candidatos:
                ids_no_encontrados.add(idn or "SIN_ID")

            fila_dict = dict(fila)
            fila_dict['idx'] = idx
            fila_dict['candidatos'] = candidatos
            fila_dict['selected_id'] = candidatos[0]['id'] if len(
                candidatos) == 1 else ""
            filas_preview.append(fila_dict)

        # Guardar datos crudos en sesión para usarlos al guardar
        request.session['ordenes_previsualizadas'] = datos_extraidos

        return render(request, 'facturacion/preview_oc.html', {
            'datos': filas_preview,
            'nombre_archivo': ", ".join(nombres_archivos),
            'ids_no_encontrados': ids_no_encontrados,
        })

    # GET
    return render(request, 'facturacion/importar_orden_compra.html')


@login_required
@rol_requerido('facturacion', 'admin')
def guardar_ordenes_compra(request):
    if request.method != 'POST':
        return redirect('facturacion:listar_oc_facturacion')

    datos_previsualizados = request.session.get('ordenes_previsualizadas')
    if not datos_previsualizados:
        messages.error(request, "No hay datos para guardar.")
        return redirect('facturacion:importar_orden_compra')

    ordenes_guardadas = 0
    ordenes_sin_oc_libre = []
    ordenes_sin_servicio = []

    for idx, item in enumerate(datos_previsualizados):
        id_new = item.get('id_new')
        if not id_new:
            continue

        # 1) Si el usuario eligió explícitamente un DU
        target_du_id = request.POST.get(f"target_du_{idx}", "").strip()
        servicio_sin_oc = None

        if target_du_id.isdigit():
            try:
                s = ServicioCotizado.objects.get(pk=int(target_du_id))
                # Validamos que aún no tenga OC
                if not s.ordenes_compra.exists():
                    servicio_sin_oc = s
            except ServicioCotizado.DoesNotExist:
                servicio_sin_oc = None

        # 2) Fallback: tomar el más reciente SIN OC
        if servicio_sin_oc is None:
            servicios = (
                ServicioCotizado.objects
                .filter(id_new=id_new, ordenes_compra__isnull=True)
                .order_by('-du', '-fecha_creacion')
                .distinct()
            )
            if not servicios.exists():
                ordenes_sin_servicio.append(f"ID NEW: {id_new}")
                continue
            servicio_sin_oc = servicios.first()

        # 3) Crear OC
        try:
            cantidad = Decimal(
                str(item.get('cantidad') or '0').replace(',', '.'))
            precio_unitario = Decimal(
                str(item.get('precio_unitario') or '0').replace(',', '.'))
            monto = Decimal(str(item.get('monto') or '0').replace(',', '.'))

            fecha_entrega = None
            fecha_texto = item.get('fecha_entrega')
            if fecha_texto:
                try:
                    fecha_entrega = datetime.strptime(
                        fecha_texto, '%d.%m.%Y').date()
                except ValueError:
                    pass

            OrdenCompraFacturacion.objects.create(
                du=servicio_sin_oc,
                orden_compra=item.get('orden_compra'),
                pos=item.get('pos'),
                cantidad=cantidad,
                unidad_medida=item.get('unidad_medida'),
                material_servicio=item.get('material_servicio'),
                descripcion_sitio=item.get('descripcion_sitio'),
                fecha_entrega=fecha_entrega,
                precio_unitario=precio_unitario,
                monto=monto,
            )
            ordenes_guardadas += 1

        except Exception as e:
            print(f"❌ Error al guardar datos de OC: {e}")
            continue

    # Limpiar sesión
    request.session.pop('ordenes_previsualizadas', None)

    if ordenes_guardadas > 0:
        messages.success(
            request, f"{ordenes_guardadas} líneas de la orden de compra fueron guardadas correctamente.")
    if ordenes_sin_oc_libre:
        messages.warning(request,
                         "Se omitieron líneas porque ya no hay servicios disponibles sin OC para asociar:<br>" +
                         "<br>".join(ordenes_sin_oc_libre)
                         )
    if ordenes_sin_servicio:
        messages.error(request,
                       "No existe servicio creado (o todos ya tienen OC) para estos ID NEW:<br>" +
                       "<br>".join(set(ordenes_sin_servicio)) +
                       "<br><br>Comunícate con el PM para que cree el servicio y vuelve a importar la OC."
                       )

    return redirect('facturacion:listar_oc_facturacion')


# =========================
#  CRUD OC
# =========================

@login_required
@rol_requerido('facturacion', 'admin')
def editar_orden_compra(request, pk):
    """
    pk puede ser:
    - id de OrdenCompraFacturacion  -> edita esa OC
    - id de ServicioCotizado        -> muestra form para crear OC de ese DU (sin guardar aún)
    """
    oc = OrdenCompraFacturacion.objects.filter(
        pk=pk).select_related('du').first()
    servicio = None

    if oc is None:
        # pk no es OC -> interpretarlo como id de Servicio
        servicio = get_object_or_404(ServicioCotizado, pk=pk)
        # Instancia SIN GUARDAR para renderizar el form sin crear registros
        oc = OrdenCompraFacturacion(du=servicio)

    if request.method == 'POST':
        form = OrdenCompraFacturacionForm(request.POST, instance=oc)
        if form.is_valid():
            oc = form.save(commit=False)
            # Si veníamos por DU (instancia no guardada), fijar su DU antes de guardar
            if oc.du_id is None and servicio is not None:
                oc.du = servicio
            oc.save()
            messages.success(
                request, "Orden de compra guardada correctamente.")
            return redirect('facturacion:listar_oc_facturacion')
    else:
        form = OrdenCompraFacturacionForm(instance=oc)

    return render(request, 'facturacion/editar_orden_compra.html', {
        'form': form,
        'oc': oc,
    })


@login_required
@rol_requerido('facturacion', 'admin')
def eliminar_orden_compra(request, pk):
    orden = get_object_or_404(OrdenCompraFacturacion, pk=pk)

    if request.method == 'POST':
        orden.delete()
        messages.success(request, "Orden de compra eliminada correctamente.")
        return redirect('facturacion:listar_oc_facturacion')

    return render(request, 'facturacion/eliminar_orden_compra.html', {'orden': orden})


# =========================
#  EXPORTAR EXCEL
# =========================

@login_required
@rol_requerido('facturacion', 'admin')
def exportar_ordenes_compra_excel(request):
    # Filtros (mismos que en listar)
    du = request.GET.get('du', '')
    id_claro = request.GET.get('id_claro', '')
    id_new = request.GET.get('id_new', '')
    mes_produccion = request.GET.get('mes_produccion', '')
    estado = request.GET.get('estado', '')

    estados_validos = [
        'cotizado',
        'aprobado_pendiente',
        'asignado',
        'en_progreso',
        'finalizado_trabajador',
        'rechazado_supervisor',
        'aprobado_supervisor',
        'informe_subido',
        'finalizado'
    ]

    servicios = ServicioCotizado.objects.select_related(
        'pm_aprueba', 'tecnico_aceptado', 'tecnico_finalizo', 'supervisor_aprobo',
        'supervisor_rechazo', 'supervisor_asigna', 'usuario_informe'
    ).prefetch_related(
        'ordenes_compra', 'trabajadores_asignados'
    ).filter(
        estado__in=estados_validos
    ).order_by('-fecha_creacion')

    if du:
        du = du.strip().upper().replace('DU', '')
        servicios = servicios.filter(du__iexact=du)
    if id_claro:
        servicios = servicios.filter(id_claro__icontains=id_claro)
    if id_new:
        servicios = servicios.filter(id_new__icontains=id_new)
    if mes_produccion:
        servicios = servicios.filter(mes_produccion__icontains=mes_produccion)
    if estado:
        servicios = servicios.filter(estado=estado)

    # Crear libro Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Órdenes de Compra"

    # Encabezados
    columnas = [
        "DU", "ID CLARO", "ID NEW", "DETALLE TAREA", "ASIGNADOS",
        "M. COTIZADO (UF)", "M. MMOO (CLP)", "FECHA FIN", "STATUS",
        "OC", "POS", "CANT", "UM", "MATERIAL", "DESCRIPCIÓN SITIO",
        "FECHA ENTREGA", "P. UNITARIO", "MONTO"
    ]
    ws.append(columnas)

    header_fill = PatternFill(start_color="D9D9D9",
                              end_color="D9D9D9", fill_type="solid")
    for col_num, col_name in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for servicio in servicios:
        asignados = ", ".join([u.get_full_name()
                              for u in servicio.trabajadores_asignados.all()]) or ''
        if servicio.ordenes_compra.exists():
            for oc in servicio.ordenes_compra.all():
                ws.append([
                    f"DU{servicio.du or ''}",
                    servicio.id_claro or '',
                    servicio.id_new or '',
                    servicio.detalle_tarea or '',
                    asignados,
                    servicio.monto_cotizado or 0,
                    servicio.monto_mmoo or 0,
                    servicio.fecha_aprobacion_supervisor.strftime(
                        "%d-%m-%Y") if servicio.fecha_aprobacion_supervisor else '',
                    servicio.get_estado_display(),
                    oc.orden_compra or '',
                    oc.pos or '',
                    oc.cantidad or '',
                    oc.unidad_medida or '',
                    oc.material_servicio or '',
                    oc.descripcion_sitio or '',
                    oc.fecha_entrega.strftime(
                        "%d-%m-%Y") if oc.fecha_entrega else '',
                    oc.precio_unitario or 0,
                    oc.monto or 0,
                ])
        else:
            ws.append([
                f"DU{servicio.du or ''}",
                servicio.id_claro or '',
                servicio.id_new or '',
                servicio.detalle_tarea or '',
                asignados,
                servicio.monto_cotizado or 0,
                servicio.monto_mmoo or 0,
                servicio.fecha_aprobacion_supervisor.strftime(
                    "%d-%m-%Y") if servicio.fecha_aprobacion_supervisor else '',
                servicio.get_estado_display(),
                '', '', '', '', '', '', '', '', ''
            ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="ordenes_compra.xlsx"'
    wb.save(response)
    return response


@login_required
@rol_requerido('facturacion', 'admin')
def listar_facturas(request):
    # Traer solo facturas existentes, no órdenes vacías
    facturas = FacturaOC.objects.select_related("orden_compra__du")

    # Filtros dinámicos (raw para mantener lo que escribe el usuario en el input)
    du_raw = (request.GET.get("du", "") or "").strip()
    id_claro = (request.GET.get("id_claro", "") or "").strip()
    id_new = (request.GET.get("id_new", "") or "").strip()
    mes_produccion = (request.GET.get("mes_produccion", "") or "").strip()
    estado = (request.GET.get("estado", "") or "").strip()

    # --- DU: normalizar para aceptar "DU00000020", "DU20", "20", etc.
    if du_raw:
        du_digits = re.sub(r"\D+", "", du_raw.upper())  # solo dígitos
        if du_digits:
            facturas = facturas.filter(
                Q(orden_compra__du__du__iexact=du_digits) |
                Q(orden_compra__du__du__endswith=du_digits)
            )

    if id_claro:
        facturas = facturas.filter(
            orden_compra__du__id_claro__icontains=id_claro)
    if id_new:
        facturas = facturas.filter(orden_compra__du__id_new__icontains=id_new)
    if mes_produccion:
        facturas = facturas.filter(
            orden_compra__du__mes_produccion__icontains=mes_produccion)
    if estado:
        facturas = facturas.filter(orden_compra__du__estado=estado)

    # Paginación (usa 'cantidad' del GET) → máx. 100
    cantidad = (request.GET.get("cantidad") or "10").strip().lower()

    if cantidad == "todos":
        # "todos" se interpreta como máximo 100
        per_page = 100
    else:
        try:
            # mínimo 5, máximo 100
            per_page = max(5, min(int(cantidad), 100))
        except ValueError:
            per_page = 10
            cantidad = "10"

    paginator = Paginator(facturas, per_page)
    page_number = request.GET.get('page')
    pagina = paginator.get_page(page_number)

    return render(request, "facturacion/listar_facturas.html", {
        "pagina": pagina,
        "filtros": {
            "du": du_raw,  # mantenemos lo que escribió el usuario
            "id_claro": id_claro,
            "id_new": id_new,
            "mes_produccion": mes_produccion,
            "estado": estado,
        },
        "estado_choices": ServicioCotizado.ESTADOS,
        "cantidad": cantidad,
    })


@login_required
@rol_requerido('facturacion', 'admin')
def enviar_a_facturacion(request):
    if request.method != "POST":
        return redirect('facturacion:listar_oc_facturacion')

    ids = request.POST.getlist('seleccionados') or []
    if not ids:
        messages.error(request, "Debes seleccionar al menos una orden.")
        return redirect('facturacion:listar_oc_facturacion')

    enviados, omitidos = [], []

    # Hacemos la operación atómica e idempotente
    with transaction.atomic():
        ocs = (OrdenCompraFacturacion.objects
               .select_related('du')
               .filter(id__in=ids))

        for oc in ocs:
            # Validar campos requeridos (si falta algo, no la mandamos)
            if not all([
                oc.orden_compra, oc.pos, oc.cantidad, oc.unidad_medida,
                oc.material_servicio, oc.descripcion_sitio,
                oc.fecha_entrega, oc.precio_unitario, oc.monto
            ]):
                omitidos.append(f"DU {oc.du.du} - POS {oc.pos}")
                continue

            # Evitar duplicados en facturación (idempotente)
            _, created = FacturaOC.objects.get_or_create(
                orden_compra=oc,
                defaults={'mes_produccion': oc.du.mes_produccion}
            )
            if created:
                enviados.append(str(oc.id))
            else:
                omitidos.append(
                    f"DU {oc.du.du} - POS {oc.pos} (ya en facturación)")

    if enviados:
        messages.success(
            request, f"{len(enviados)} órdenes fueron movidas a facturación correctamente."
        )
    if omitidos:
        messages.warning(
            request, "Las siguientes órdenes no fueron movidas:<br>" +
            "<br>".join(omitidos)
        )

    # Vamos a la lista de facturas (lo que esperas al “mover” a facturación)
    return redirect('facturacion:listar_facturas')


def limpiar_fecha(valor):
    """
    Intenta convertir múltiples formatos de fecha a YYYY-MM-DD.
    Acepta: 01-08-2025, 2025-08-01, '8 de Julio del 2025', etc.
    """
    if not valor:
        return None
    try:
        if isinstance(valor, datetime):
            return valor.date()
        fecha = parser.parse(str(valor), dayfirst=True, fuzzy=True)
        return fecha.date()
    except Exception:
        return None


@login_required
@rol_requerido('facturacion', 'admin')
def importar_facturas(request):
    datos = []
    if request.method == "POST":
        form = ImportarFacturasForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                id_claro, oc, hes, valor_en_clp, conformidad, num_factura, fecha_facturacion = row[
                    :7]

                # Limpiar y normalizar la fecha
                fecha_limpia = limpiar_fecha(fecha_facturacion)

                datos.append({
                    "fila": i,
                    "id_claro": str(id_claro).strip() if id_claro else None,
                    "oc": str(oc).strip() if oc else None,
                    "hes": hes,
                    "valor_en_clp": valor_en_clp,
                    "conformidad": conformidad,
                    "num_factura": num_factura,
                    "fecha_facturacion": fecha_limpia.strftime("%Y-%m-%d") if fecha_limpia else None,
                })

            request.session["facturas_previsualizadas"] = datos
            messages.info(
                request, "Previsualización cargada. Revisa los datos antes de guardar."
            )
            return render(request, "facturacion/importar_facturas.html", {"form": form, "datos": datos})
    else:
        form = ImportarFacturasForm()
    return render(request, "facturacion/importar_facturas.html", {"form": form})


def limpiar_monto(valor):
    """
    Convierte un valor con formato chileno (con $ y puntos) a Decimal.
    Ej: "1.041.063" -> 1041063
    """
    from decimal import Decimal, InvalidOperation
    if valor is None or valor == "":
        return None
    if isinstance(valor, (int, float, Decimal)):
        return Decimal(str(valor))
    try:
        limpio = str(valor).strip()
        limpio = re.sub(r"[^\d,.-]", "", limpio)  # Quitar símbolos
        limpio = limpio.replace(".", "")          # Eliminar puntos (miles)
        limpio = limpio.replace(",", ".")         # Reemplazar coma por punto
        return Decimal(limpio)
    except (InvalidOperation, ValueError):
        return None




@login_required
@rol_requerido('facturacion', 'admin')
def guardar_facturas(request):
    datos = request.session.get("facturas_previsualizadas")
    if not datos:
        messages.error(request, "No hay datos para guardar.")
        return redirect("facturacion:importar_facturas")

    actualizados, omitidos = 0, []
    facturas_actualizadas_en_sesion = set()

    for fila in datos:
        id_claro = fila.get("id_claro")
        oc = fila.get("oc")

        # Validar que existan id_claro y oc
        if not id_claro or not oc:
            faltantes = []
            if not id_claro:
                faltantes.append("Sin ID CLARO")
            if not oc:
                faltantes.append("Sin OC")
            omitidos.append(
                f"Fila {fila.get('fila')}: {', '.join(faltantes)}.")
            continue

        # Limpiar y convertir el valor a Decimal
        valor = limpiar_monto(fila.get("valor_en_clp"))
        if valor is None:
            omitidos.append(f"Fila {fila.get('fila')}: Valor en CLP inválido.")
            continue

        # Convertir fecha
        fecha = None
        if fila.get("fecha_facturacion"):
            try:
                if isinstance(fila["fecha_facturacion"], str):
                    fecha = datetime.strptime(
                        fila["fecha_facturacion"], "%Y-%m-%d").date()
                elif isinstance(fila["fecha_facturacion"], datetime):
                    fecha = fila["fecha_facturacion"].date()
            except ValueError:
                omitidos.append(f"Fila {fila.get('fila')}: Fecha inválida.")
                continue

        # Validar obligatorios
        if not all([fila.get("hes"), valor, fila.get("conformidad")]):
            omitidos.append(
                f"Fila {fila.get('fila')}: Faltan datos obligatorios.")
            continue

        # Buscar todas las facturas que coincidan con ID_CLARO + OC
        facturas = FacturaOC.objects.filter(
            orden_compra__orden_compra=oc,
            orden_compra__du__id_claro=id_claro
        ).order_by('id')  # más antiguas primero

        if not facturas.exists():
            omitidos.append(
                f"Fila {fila.get('fila')}: No existe Factura para ID_CLARO {id_claro} y OC {oc}."
            )
            continue

        # Buscar la primera factura sin conformidad que no haya sido usada en esta sesión
        factura = None
        for f in facturas:
            if not f.conformidad and f.id not in facturas_actualizadas_en_sesion:
                factura = f
                break

        if not factura:
            # Todas tienen conformidad → no se puede actualizar
            omitidos.append(
                f"Fila {fila.get('fila')}: Todas las facturas para ID_CLARO {id_claro} y OC {oc} ya tienen conformidad."
            )
            continue

        # Actualizar la factura seleccionada
        factura.hes = fila.get("hes")
        factura.valor_en_clp = valor
        factura.conformidad = fila.get("conformidad")
        factura.num_factura = fila.get("num_factura")
        factura.fecha_facturacion = fecha
        factura.save()

        # Marcar como usada en esta sesión
        facturas_actualizadas_en_sesion.add(factura.id)
        actualizados += 1

    # Limpiar sesión
    request.session.pop("facturas_previsualizadas", None)

    # Mensajes
    if actualizados:
        messages.success(
            request, f"{actualizados} facturas actualizadas correctamente.")
    if omitidos:
        messages.warning(request, "Omitidas:<br>" + "<br>".join(omitidos))
    return redirect("facturacion:listar_facturas")


@login_required
@rol_requerido('facturacion', 'admin')
def editar_factura(request, pk):
    factura = get_object_or_404(FacturaOC, pk=pk)
    if request.method == "POST":
        form = FacturaOCForm(request.POST, instance=factura)
        if form.is_valid():
            form.save()
            messages.success(request, "Factura actualizada correctamente.")
            return redirect('facturacion:listar_facturas')
    else:
        form = FacturaOCForm(instance=factura)
    return render(request, "facturacion/editar_factura.html", {"form": form})


@login_required
@rol_requerido('admin')
def eliminar_factura(request, pk):
    factura = get_object_or_404(FacturaOC, pk=pk)
    if request.method == "POST":
        factura.delete()
        messages.success(request, "Factura eliminada correctamente.")
        return redirect('facturacion:listar_facturas')
    return render(request, "facturacion/eliminar_factura.html", {"factura": factura})


@csrf_exempt
def actualizar_factura_ajax(request, pk):
    if request.method == "POST":
        factura = get_object_or_404(FacturaOC, pk=pk)
        campo = request.POST.get("campo")
        valor = request.POST.get("valor")

        # Conversión según tipo de campo
        if campo in ["valor_en_clp"]:
            try:
                valor = float(valor.replace(",", "").replace("$", "").strip())
            except:
                return JsonResponse({"success": False, "error": "Valor inválido"})
        if campo in ["factorizado"]:
            valor = valor.lower() in ["1", "true", "sí", "si"]

        # Guardar valor
        setattr(factura, campo, valor if valor != "" else None)
        factura.save()

        # Recalcular el estado dinámicamente
        nuevo_status = factura.get_status_factura()

        return JsonResponse({
            "success": True,
            "valor": valor,
            "nuevo_status": nuevo_status  # <-- Devolvemos el nuevo estado
        })
    return JsonResponse({"success": False, "error": "Método no permitido"})


@login_required
@rol_requerido('facturacion', 'admin')
def exportar_facturacion_excel(request):
    import openpyxl
    from django.http import HttpResponse
    from openpyxl.styles import Alignment, Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Facturación"

    columnas = [
        "DU", "ID CLARO", "ID NEW", "DETALLE TAREA", "ASIGNADOS",
        "M. COTIZADO (UF)", "M. MMOO (CLP)", "FECHA FIN", "STATUS SERVICIO",
        "OC", "POS", "CANT", "UM", "MATERIAL", "DESCRIPCIÓN SITIO", "FECHA ENTREGA",
        "P. UNITARIO", "MONTO", "HES", "VALOR EN CLP", "CONFORMIDAD",
        "N° FACTURA", "FECHA FACTURACIÓN", "MES DE PRODUCCIÓN",
        "FACTORIZADO", "FECHA FACTORING", "STATUS FACTURA"
    ]
    for col_num, column_title in enumerate(columnas, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    facturas = (
        FacturaOC.objects
        .select_related(
            'orden_compra', 'orden_compra__du',
            'orden_compra__du__pm_aprueba',
            'orden_compra__du__tecnico_aceptado',
            'orden_compra__du__tecnico_finalizo',
            'orden_compra__du__supervisor_aprobo',
            'orden_compra__du__supervisor_rechazo',
            'orden_compra__du__supervisor_asigna',
            'orden_compra__du__usuario_informe',
        )
        .prefetch_related('orden_compra__du__trabajadores_asignados')
    )

    for row_num, factura in enumerate(facturas, start=2):
        oc = factura.orden_compra
        du = oc.du if oc else None

        # Servicio
        ws.cell(row=row_num, column=1,
                value=f"DU{du.du}" if du and du.du else "")
        ws.cell(row=row_num, column=2, value=du.id_claro if du else "")
        ws.cell(row=row_num, column=3, value=du.id_new if du else "")
        ws.cell(row=row_num, column=4, value=du.detalle_tarea if du else "")
        ws.cell(row=row_num, column=5, value=", ".join(
            [u.get_full_name() for u in du.trabajadores_asignados.all()]) if du else "")
        ws.cell(row=row_num, column=6, value=float(
            du.monto_cotizado) if du and du.monto_cotizado else 0)
        ws.cell(row=row_num, column=7, value=float(
            du.monto_mmoo) if du and du.monto_mmoo else 0)
        ws.cell(row=row_num, column=8, value=du.fecha_aprobacion_supervisor.strftime(
            "%d-%m-%Y") if du and du.fecha_aprobacion_supervisor else "")
        ws.cell(row=row_num, column=9,
                value=du.get_estado_display() if du else "")

        # Orden de compra
        ws.cell(row=row_num, column=10, value=oc.orden_compra if oc else "")
        ws.cell(row=row_num, column=11, value=oc.pos if oc else "")
        ws.cell(row=row_num, column=12, value=float(
            oc.cantidad) if oc and oc.cantidad else 0)
        ws.cell(row=row_num, column=13, value=oc.unidad_medida if oc else "")
        ws.cell(row=row_num, column=14,
                value=oc.material_servicio if oc else "")
        ws.cell(row=row_num, column=15,
                value=oc.descripcion_sitio if oc else "")
        ws.cell(row=row_num, column=16, value=oc.fecha_entrega.strftime(
            "%d-%m-%Y") if oc and oc.fecha_entrega else "")
        ws.cell(row=row_num, column=17, value=float(
            oc.precio_unitario) if oc and oc.precio_unitario else 0)
        ws.cell(row=row_num, column=18, value=float(
            oc.monto) if oc and oc.monto else 0)

        # Factura
        ws.cell(row=row_num, column=19, value=factura.hes or "")
        ws.cell(row=row_num, column=20, value=float(
            factura.valor_en_clp) if factura.valor_en_clp else 0)
        ws.cell(row=row_num, column=21, value=factura.conformidad or "")
        ws.cell(row=row_num, column=22, value=factura.num_factura or "")
        ws.cell(row=row_num, column=23, value=factura.fecha_facturacion.strftime(
            "%d-%m-%Y") if factura.fecha_facturacion else "")
        ws.cell(row=row_num, column=24, value=factura.mes_produccion or "")
        ws.cell(row=row_num, column=25,
                value="Sí" if factura.factorizado else "No")
        ws.cell(row=row_num, column=26, value=factura.fecha_factoring.strftime(
            "%d-%m-%Y") if factura.fecha_factoring else "")
        ws.cell(row=row_num, column=27, value=factura.get_status_factura())

    # Ajustar ancho
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename=Lista_Facturacion.xlsx'
    wb.save(response)
    return response


import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Case, CharField, IntegerField, Q, Value, When
from django.db.models.functions import Cast
from django.shortcuts import render

# Ajusta imports según tu estructura real
# from .models import CartolaMovimiento
# from .decorators import rol_requerido


@login_required
@rol_requerido('facturacion', 'admin')
def listar_cartola(request):
    # --- Cantidad/paginación (mantén string para la UI)
    cantidad_param = request.GET.get('cantidad', '10')

    if cantidad_param == 'todos':
        page_size = 100
    else:
        try:
            page_size = max(5, min(int(cantidad_param), 100))
        except ValueError:
            page_size = 10
            cantidad_param = '10'

    # ============================================================
    # ✅ Excel filters (JSON) recibido por GET
    # ============================================================
    params = request.GET.copy()
    excel_filters_raw = (params.get('excel_filters') or '').strip()
    try:
        excel_filters = json.loads(excel_filters_raw) if excel_filters_raw else {}
    except json.JSONDecodeError:
        excel_filters = {}

    # --- Filtros clásicos
    usuario = (request.GET.get('usuario') or '').strip()
    fecha_txt = (request.GET.get('fecha') or '').strip()
    fecha_real_txt = (request.GET.get('fecha_real') or '').strip()
    proyecto = (request.GET.get('proyecto') or '').strip()
    categoria = (request.GET.get('categoria') or '').strip()
    tipo = (request.GET.get('tipo') or '').strip()
    rut_factura = (request.GET.get('rut_factura') or '').strip()
    estado = (request.GET.get('estado') or '').strip()
    hora_servicio = (request.GET.get('hora_servicio') or '').strip()
    kilometraje_servicio = (request.GET.get('kilometraje_servicio') or '').strip()

    # --- Base queryset
    movimientos = CartolaMovimiento.objects.select_related(
        'usuario',
        'proyecto',
        'tipo',
        'vehiculo_flota',
        'tipo_servicio_flota',
        'servicio_flota',
        'aprobado_por_supervisor',
        'aprobado_por_pm',
        'aprobado_por_finanzas',
    ).all()

    # Anotaciones auxiliares para filtros parciales
    movimientos = movimientos.annotate(
        fecha_iso=Cast('fecha', CharField()),
        fecha_transaccion_iso=Cast('fecha_transaccion', CharField()),
        km_servicio_iso=Cast('kilometraje_servicio_flota', CharField()),
    )

    # --- Filtro usuario
    if usuario:
        movimientos = movimientos.filter(
            Q(usuario__username__icontains=usuario) |
            Q(usuario__first_name__icontains=usuario) |
            Q(usuario__last_name__icontains=usuario)
        )

    # --- Helper fecha parcial (DD, DD-MM, DD-MM-YYYY)
    def aplicar_filtro_fecha_parcial(qs, field_name, valor, warning_label):
        if not valor:
            return qs

        fecha_normalizada = valor.replace('/', '-').strip()
        m = re.match(r'^(\d{1,2})(?:-(\d{1,2}))?(?:-(\d{1,4}))?$', fecha_normalizada)

        if not m:
            messages.warning(request, f"Formato de {warning_label} inválido. Use DD, DD-MM o DD-MM-YYYY.")
            return qs

        dia_str = m.group(1)
        mes_str = m.group(2)
        anio_str = m.group(3)

        try:
            q_fecha = Q(**{f"{field_name}__day": int(dia_str)})

            if mes_str:
                q_fecha &= Q(**{f"{field_name}__month": int(mes_str)})

            if anio_str and len(anio_str) == 4:
                q_fecha &= Q(**{f"{field_name}__year": int(anio_str)})

            return qs.filter(q_fecha)

        except ValueError:
            messages.warning(request, f"Formato de {warning_label} inválido. Use DD, DD-MM o DD-MM-YYYY.")
            return qs

    # --- Filtros fecha
    movimientos = aplicar_filtro_fecha_parcial(movimientos, 'fecha', fecha_txt, 'fecha')

    if fecha_real_txt:
        movimientos = movimientos.filter(fecha_transaccion__isnull=False)
        movimientos = aplicar_filtro_fecha_parcial(
            movimientos, 'fecha_transaccion', fecha_real_txt, 'fecha real del gasto'
        )

    # --- Otros filtros clásicos
    if proyecto:
        movimientos = movimientos.filter(proyecto__nombre__icontains=proyecto)

    if categoria:
        movimientos = movimientos.filter(tipo__categoria__icontains=categoria)

    if tipo:
        movimientos = movimientos.filter(tipo__nombre__icontains=tipo)

    if rut_factura:
        movimientos = movimientos.filter(rut_factura__icontains=rut_factura)

    if estado:
        movimientos = movimientos.filter(status=estado)

    if hora_servicio:
        movimientos = movimientos.annotate(
            hora_servicio_iso=Cast('hora_servicio_flota', CharField())
        ).filter(hora_servicio_iso__icontains=hora_servicio)

    if kilometraje_servicio:
        movimientos = movimientos.filter(km_servicio_iso__icontains=kilometraje_servicio)

    # ==========================================================
    # ✅ MOSTRAR TODO EL PROCESO + ABONOS (solo fuera de historial)
    # ==========================================================
    estados_visibles = [codigo for codigo, _label in CartolaMovimiento.ESTADOS]

    movimientos = movimientos.filter(
        en_historial=False,
        status__in=estados_visibles
    )

    # --- Orden personalizado
    movimientos = movimientos.annotate(
        prioridad=Case(
            When(status='pendiente_supervisor', then=Value(0)),
            When(status='aprobado_supervisor', then=Value(1)),
            When(status='aprobado_pm', then=Value(2)),
            When(status='pendiente_abono_usuario', then=Value(3)),
            When(status__startswith='rechazado', then=Value(4)),
            When(status='aprobado_finanzas', then=Value(5)),
            When(status='aprobado_abono_usuario', then=Value(6)),
            When(status__startswith='aprobado', then=Value(7)),
            When(status__startswith='pendiente', then=Value(8)),
            default=Value(9),
            output_field=IntegerField(),
        )
    ).order_by(
        'prioridad',
        '-fecha_transaccion',
        '-fecha',
        '-id'
    )

    # ============================================================
    # ✅ Excel filters (python) + excel_global_json (TODO el set filtrado)
    #    IMPORTANTE: Los valores deben ser EXACTAMENTE los mismos que en data-excel-value del template
    # ============================================================
    movimientos_list = list(movimientos)

    def _norm(s):
        s = "" if s is None else str(s)
        s = s.replace("\u00A0", " ")
        s = re.sub(r"\s+", " ", s).strip().lower()
        return s

    def format_clp(n):
        try:
            n = 0 if n is None else n
            n_int = int(n)
            return f"${n_int:,}".replace(",", ".")
        except Exception:
            return "$0"

    def format_km_cell(n):
        """
        Debe calzar con el template:
        data-excel-value="{% if mov.kilometraje_servicio_flota %}{{ mov.kilometraje_servicio_flota|miles }} KM{% else %}—{% endif %}"
        """
        try:
            if n is None or n == "":
                return "—"
            n_int = int(float(n))
            return f"{n_int:,}".replace(",", ".") + " KM"
        except Exception:
            return "—"

    def _fmt_fecha_cell(dt_or_d):
        """
        Debe calzar con el template:
        data-excel-value="{{ mov.fecha|date:'d-m-Y' }}"
        (soporta DateField o DateTimeField)
        """
        if not dt_or_d:
            return ""
        try:
            # si es datetime aware/naive, intentamos localtime si aplica
            if hasattr(dt_or_d, "hour"):
                try:
                    if timezone.is_aware(dt_or_d):
                        dt_or_d = timezone.localtime(dt_or_d)
                except Exception:
                    pass
                return dt_or_d.strftime("%d-%m-%Y")
            return dt_or_d.strftime("%d-%m-%Y")
        except Exception:
            try:
                return str(dt_or_d)
            except Exception:
                return ""

    def _fmt_hora_cell(t):
        """
        Debe calzar con template:
        data-excel-value="{% if mov.hora_servicio_flota %}{{ mov.hora_servicio_flota|time:'H:i' }}{% else %}—{% endif %}"
        """
        if not t:
            return "—"
        try:
            return t.strftime("%H:%M")
        except Exception:
            return str(t)

    def _fmt_obs_cell(obs):
        """
        Debe calzar con template:
        data-excel-value="{{ mov.observaciones|default:'(Vacías)' }}"
        """
        s = (obs or "").strip()
        return s if s else "(Vacías)"

    def _status_cell(mov):
        """
        Debe calzar con template:
        data-excel-value="{{ mov.get_status_display }}"
        """
        try:
            return mov.get_status_display() if getattr(mov, "status", None) else ""
        except Exception:
            return ""

    if excel_filters:
        # normalizamos valores seleccionados (lo mismo que hace tu JS)
        excel_filters_norm = {}
        for col, values in (excel_filters or {}).items():
            if not values:
                continue
            try:
                excel_filters_norm[str(col)] = set(_norm(v) for v in values)
            except Exception:
                continue

        def matches_excel_filters(mov):
            for col, allowed_norm in excel_filters_norm.items():
                if not allowed_norm:
                    continue

                # Índices tabla cartola (CON checkbox):
                # 0  Checkbox
                # 1  Usuario
                # 2  Fecha
                # 3  Fecha real del gasto
                # 4  Vehículo
                # 5  Hora servicio
                # 6  Kilometraje servicio
                # 7  Proyecto
                # 8  Categoría
                # 9  Tipo
                # 10 RUT factura
                # 11 Tipo de documento
                # 12 Número de documento
                # 13 Observaciones
                # 14 N° Transferencia
                # 15 Comprobante
                # 16 Cargos
                # 17 Abonos
                # 18 Status
                # 19 Acciones (no se filtra)

                if col == "1":
                    label = str(mov.usuario) if mov.usuario else ""

                elif col == "2":
                    label = _fmt_fecha_cell(getattr(mov, "fecha", None))

                elif col == "3":
                    d = getattr(mov, "fecha_transaccion", None) or getattr(mov, "fecha", None)
                    label = _fmt_fecha_cell(d)

                elif col == "4":
                    label = str(getattr(mov, "vehiculo_flota", None) or "—").strip() or "—"

                elif col == "5":
                    label = _fmt_hora_cell(getattr(mov, "hora_servicio_flota", None))

                elif col == "6":
                    label = format_km_cell(getattr(mov, "kilometraje_servicio_flota", None))

                elif col == "7":
                    label = str(mov.proyecto) if mov.proyecto else ""

                elif col == "8":
                    if mov.tipo and getattr(mov.tipo, "categoria", None):
                        label = (mov.tipo.categoria or "").title()
                    else:
                        label = ""

                elif col == "9":
                    label = str(mov.tipo) if mov.tipo else ""

                elif col == "10":
                    label = (getattr(mov, "rut_factura", None) or "—").strip() or "—"

                elif col == "11":
                    label = (getattr(mov, "tipo_doc", None) or "—").strip() or "—"

                elif col == "12":
                    label = (getattr(mov, "numero_doc", None) or "—").strip() or "—"

                elif col == "13":
                    label = _fmt_obs_cell(getattr(mov, "observaciones", None))

                elif col == "14":
                    label = (getattr(mov, "numero_transferencia", None) or "—").strip() or "—"

                elif col == "15":
                    label = "Ver" if getattr(mov, "comprobante", None) else "—"

                elif col == "16":
                    label = format_clp(getattr(mov, "cargos", 0) or 0)

                elif col == "17":
                    label = format_clp(getattr(mov, "abonos", 0) or 0)

                elif col == "18":
                    label = _status_cell(mov)

                else:
                    continue

                if _norm(label) not in allowed_norm:
                    return False

            return True

        movimientos_list = [m for m in movimientos_list if matches_excel_filters(m)]

    # --- excel_global (para panel Excel, basado en TODO el dataset filtrado)
    #     IMPORTANTE: producir los mismos textos que van en data-excel-value
    excel_global = {}

    excel_global[1] = sorted({str(m.usuario) for m in movimientos_list if m.usuario})

    excel_global[2] = sorted({
        _fmt_fecha_cell(getattr(m, "fecha", None))
        for m in movimientos_list
        if getattr(m, "fecha", None)
    })

    excel_global[3] = sorted({
        _fmt_fecha_cell(getattr(m, "fecha_transaccion", None) or getattr(m, "fecha", None))
        for m in movimientos_list
        if (getattr(m, "fecha_transaccion", None) or getattr(m, "fecha", None))
    })

    excel_global[4] = sorted({
        (str(getattr(m, "vehiculo_flota", None) or "—").strip() or "—")
        for m in movimientos_list
    })

    excel_global[5] = sorted({
        _fmt_hora_cell(getattr(m, "hora_servicio_flota", None))
        for m in movimientos_list
    })

    excel_global[6] = sorted({
        format_km_cell(getattr(m, "kilometraje_servicio_flota", None))
        for m in movimientos_list
    })

    excel_global[7] = sorted({str(m.proyecto) for m in movimientos_list if m.proyecto})

    excel_global[8] = sorted({
        (m.tipo.categoria or "").title()
        for m in movimientos_list
        if m.tipo and getattr(m.tipo, "categoria", None)
    })

    excel_global[9] = sorted({str(m.tipo) for m in movimientos_list if m.tipo})

    excel_global[10] = sorted({
        (getattr(m, "rut_factura", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    excel_global[11] = sorted({
        (getattr(m, "tipo_doc", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    excel_global[12] = sorted({
        (getattr(m, "numero_doc", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    excel_global[13] = sorted({
        _fmt_obs_cell(getattr(m, "observaciones", None))
        for m in movimientos_list
    })

    excel_global[14] = sorted({
        (getattr(m, "numero_transferencia", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    excel_global[15] = sorted({
        "Ver" if getattr(m, "comprobante", None) else "—"
        for m in movimientos_list
    })

    excel_global[16] = sorted({
        format_clp(getattr(m, "cargos", 0) or 0)
        for m in movimientos_list
    })

    excel_global[17] = sorted({
        format_clp(getattr(m, "abonos", 0) or 0)
        for m in movimientos_list
    })

    # ✅ CLAVE: status debe salir EXACTAMENTE como en la celda (get_status_display),
    # NO como mapping manual (que puede diferir y romper el filtro)
    excel_global[18] = sorted({
        _status_cell(m)
        for m in movimientos_list
        if getattr(m, "status", None)
    })

    excel_global_json = json.dumps(excel_global)

    # --- Paginación (después de aplicar excel_filters)
    paginator = Paginator(movimientos_list, page_size)
    page_number = request.GET.get('page')
    pagina = paginator.get_page(page_number)

    # --- Choices / filtros para template
    estado_choices = CartolaMovimiento.ESTADOS
    filtros = {
        'usuario': usuario,
        'fecha': fecha_txt,
        'fecha_real': fecha_real_txt,
        'proyecto': proyecto,
        'categoria': categoria,
        'tipo': tipo,
        'rut_factura': rut_factura,
        'estado': estado,
        'hora_servicio': hora_servicio,
        'kilometraje_servicio': kilometraje_servicio,
    }

    # ✅ para construir links preservando querystring
    params_no_page = params.copy()
    params_no_page.pop('page', None)
    base_qs = params_no_page.urlencode()
    full_qs = params.urlencode()

    return render(
        request,
        'facturacion/listar_cartola.html',
        {
            'pagina': pagina,
            'cantidad': cantidad_param,
            'estado_choices': estado_choices,
            'filtros': filtros,
            'excel_global_json': excel_global_json,
            'base_qs': base_qs,
            'full_qs': full_qs,
        }
    )


@login_required
@rol_requerido('facturacion', 'admin')
def registrar_abono(request):
    if request.method == 'POST':
        form = CartolaAbonoForm(request.POST, request.FILES)
        if form.is_valid():
            movimiento = form.save(commit=False)
            # Si es abono, setear automáticamente valores
            from .models import TipoGasto
            tipo_abono = TipoGasto.objects.filter(categoria='abono').first()
            movimiento.tipo = tipo_abono
            movimiento.cargos = 0
            movimiento.save()
            messages.success(request, "Movimiento registrado correctamente.")
            # <-- Redirección después de guardar
            return redirect('facturacion:listar_cartola')
        else:
            messages.error(
                request, "Por favor corrige los errores antes de continuar.")
    else:
        form = CartolaAbonoForm()
    return render(request, 'facturacion/registrar_abono.html', {'form': form})


@login_required
@rol_requerido('facturacion', 'admin')
def crear_tipo(request):
    if request.method == 'POST':
        form = TipoGastoForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                tipos = TipoGasto.objects.all().order_by('-id')
                html = render_to_string(
                    'facturacion/partials/tipo_gasto_table.html',
                    {'tipos': tipos},
                    request=request
                )
                return JsonResponse({'success': True, 'html': html})
            messages.success(request, "Tipo de gasto creado correctamente.")
            return redirect('facturacion:crear_tipo')
    else:
        form = TipoGastoForm()

    tipos = TipoGasto.objects.all().order_by('-id')
    return render(request, 'facturacion/crear_tipo.html', {'form': form, 'tipos': tipos})


@login_required
@rol_requerido('admin')
def editar_tipo(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)

    if request.method == 'POST':
        form = TipoGastoForm(request.POST, instance=tipo)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                tipos = TipoGasto.objects.all().order_by('-id')
                html = render_to_string(
                    'facturacion/partials/tipo_gasto_table.html',
                    {'tipos': tipos},
                    request=request
                )
                return JsonResponse({'success': True, 'html': html})
            messages.success(request, "Tipo de gasto actualizado correctamente.")
            return redirect('facturacion:crear_tipo')
    else:
        form = TipoGastoForm(instance=tipo)

    tipos = TipoGasto.objects.all().order_by('-id')
    return render(request, 'facturacion/crear_tipo.html', {'form': form, 'tipos': tipos, 'editando': True})


@login_required
@rol_requerido('admin')
def eliminar_tipo(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)
    tipo.delete()

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        tipos = TipoGasto.objects.all().order_by('-id')
        html = render_to_string(
            'facturacion/partials/tipo_gasto_table.html',
            {'tipos': tipos},
            request=request
        )
        return JsonResponse({'success': True, 'html': html})

    messages.success(request, "Tipo de gasto eliminado correctamente.")
    return redirect('facturacion:crear_tipo')


# ✅ NUEVA VISTA: activar/desactivar tipo de gasto
@login_required
@rol_requerido('admin')
def toggle_disponible_tipo(request, pk):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método no permitido'}, status=405)

    tipo = get_object_or_404(TipoGasto, pk=pk)
    tipo.disponible = not tipo.disponible
    tipo.save(update_fields=['disponible'])

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        tipos = TipoGasto.objects.all().order_by('-id')
        html = render_to_string(
            'facturacion/partials/tipo_gasto_table.html',
            {'tipos': tipos},
            request=request
        )
        return JsonResponse({
            'success': True,
            'html': html,
            'disponible': tipo.disponible,
            'msg': f'Tipo "{tipo.nombre}" {"activado" if tipo.disponible else "desactivado"} correctamente.'
        })

    messages.success(
        request,
        f'Tipo "{tipo.nombre}" {"activado" if tipo.disponible else "desactivado"} correctamente.'
    )
    return redirect('facturacion:crear_tipo')


# Listar y crear
@login_required
@rol_requerido('facturacion', 'admin')
def crear_proyecto(request):
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Proyecto creado correctamente.")
            return redirect('facturacion:crear_proyecto')
    else:
        form = ProyectoForm()
    proyectos = Proyecto.objects.all().order_by('-id')
    return render(request, 'facturacion/crear_proyecto.html', {'form': form, 'proyectos': proyectos})

# Editar


@login_required
@rol_requerido('admin')
def editar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proyecto)
        if form.is_valid():
            form.save()
            messages.success(request, "Proyecto actualizado correctamente.")
            return redirect('facturacion:crear_proyecto')
    else:
        form = ProyectoForm(instance=proyecto)
    proyectos = Proyecto.objects.all().order_by('-id')
    return render(request, 'facturacion/crear_proyecto.html', {'form': form, 'proyectos': proyectos})

# Eliminar


@login_required
@rol_requerido('admin')
def eliminar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    if request.method == 'POST':
        proyecto.delete()
        messages.success(request, "Proyecto eliminado correctamente.")
        return redirect('facturacion:crear_proyecto')
    return redirect('facturacion:crear_proyecto')


from django.http import JsonResponse
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST


def _is_ajax(request):
    hdr = (request.headers.get('x-requested-with') or '').lower()
    return hdr in ('xmlhttprequest', 'fetch') or request.headers.get('Hx-Request') == 'true'

@login_required
@rol_requerido('facturacion', 'supervisor', 'pm', 'admin')
@require_POST
@csrf_protect
def aprobar_movimiento(request, pk):
    mov = get_object_or_404(CartolaMovimiento, pk=pk)
    changed = False

    # 👇 da poder a admin (staff) además del superusuario
    es_admin = getattr(request.user, 'es_admin', False) or request.user.is_staff

    # Helper interno: confirma lectura de km en flota cuando FINANZAS aprueba
    def _confirmar_lectura_flota_si_corresponde(movimiento):
        """
        Marca como aprobada la lectura pendiente creada al momento de la declaración.
        Ajusta nombres según tu app/modelo de flota.
        """
        try:
            # ✅ Cambia este import por el real de tu módulo de flota
            from flota.models import LecturaKilometraje

            # Buscamos la lectura creada desde cartola para este movimiento
            lectura = (
                LecturaKilometraje.objects
                .filter(cartola_movimiento=movimiento)
                .order_by('-id')
                .first()
            )

            if lectura:
                # Confirmar aprobación financiera
                lectura.aprobado = True
                lectura.aprobado_en = timezone.now()
                lectura.aprobado_por = request.user
                lectura.estado = 'aprobado'  # si manejas estado string
                lectura.save(update_fields=[
                    'aprobado', 'aprobado_en', 'aprobado_por', 'estado'
                ])
        except Exception:
            # No rompemos la aprobación contable por un error de flota
            # (pero idealmente loguear)
            traceback.print_exc()

    if mov.tipo and mov.tipo.categoria != "abono":
        # superuser o admin pueden avanzar en cualquier etapa
        if request.user.is_superuser or es_admin:
            if mov.status == 'pendiente_supervisor':
                mov.status = 'aprobado_supervisor'
                mov.aprobado_por_supervisor = request.user
                changed = True

            elif mov.status == 'aprobado_supervisor':
                mov.status = 'aprobado_pm'
                mov.aprobado_por_pm = request.user
                changed = True

            elif mov.status == 'aprobado_pm':
                mov.status = 'aprobado_finanzas'
                mov.aprobado_por_finanzas = request.user

                # ✅ Al aprobar finanzas: pasa a historial
                now = timezone.now()
                mov.aprobado_finanzas_en = now
                mov.en_historial = True
                mov.historial_enviado_el = now
                mov.historial_enviado_por = request.user

                changed = True

        # flujo por roles específicos
        elif getattr(request.user, 'es_supervisor', False) and mov.status == 'pendiente_supervisor':
            mov.status = 'aprobado_supervisor'
            mov.aprobado_por_supervisor = request.user
            changed = True

        elif getattr(request.user, 'es_pm', False) and mov.status == 'aprobado_supervisor':
            mov.status = 'aprobado_pm'
            mov.aprobado_por_pm = request.user
            changed = True

        elif getattr(request.user, 'es_facturacion', False) and mov.status == 'aprobado_pm':
            mov.status = 'aprobado_finanzas'
            mov.aprobado_por_finanzas = request.user

            # ✅ Al aprobar finanzas: pasa a historial
            now = timezone.now()
            mov.aprobado_finanzas_en = now
            mov.en_historial = True
            mov.historial_enviado_el = now
            mov.historial_enviado_por = request.user

            changed = True

        if changed:
            mov.motivo_rechazo = ''
            update_fields = ['status', 'motivo_rechazo']

            if mov.status == 'aprobado_supervisor':
                update_fields.append('aprobado_por_supervisor')

            elif mov.status == 'aprobado_pm':
                update_fields.append('aprobado_por_pm')

            elif mov.status == 'aprobado_finanzas':
                update_fields.extend([
                    'aprobado_por_finanzas',
                    'aprobado_finanzas_en',
                    'en_historial',
                    'historial_enviado_el',
                    'historial_enviado_por',
                ])

            mov.save(update_fields=update_fields)

            # ✅ Punto 3: confirmar lectura en flota SOLO cuando finanzas aprueba
            if mov.status == 'aprobado_finanzas':
                _confirmar_lectura_flota_si_corresponde(mov)

            if _is_ajax(request):
                return JsonResponse({"ok": True, "id": mov.pk, "new_status": mov.status})

            messages.success(request, "Gasto aprobado correctamente.")

        else:
            if _is_ajax(request):
                return JsonResponse(
                    {"ok": False, "error": "No autorizado o estado inválido."},
                    status=403
                )
            messages.warning(request, "No puedes aprobar este movimiento en su estado actual.")
    else:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "error": "Movimiento inválido."}, status=400)
        messages.error(request, "Movimiento inválido.")

    next_url = (
        request.POST.get('next')
        or request.GET.get('next')
        or request.META.get('HTTP_REFERER')
        or reverse('facturacion:listar_cartola')
    )
    return redirect(next_url)


@login_required
@rol_requerido('facturacion', 'supervisor', 'pm', 'admin')
def rechazar_movimiento(request, pk):
    mov = get_object_or_404(CartolaMovimiento, pk=pk)

    if request.method == 'POST':
        motivo = request.POST.get('motivo_rechazo', '').strip()
        changed = False

        if mov.tipo and mov.tipo.categoria != "abono":
            # Nueva rama: superusuario puede rechazar según etapa actual
            if request.user.is_superuser:
                if mov.status == 'pendiente_supervisor':
                    mov.status = 'rechazado_supervisor'
                    mov.aprobado_por_supervisor = request.user
                    changed = True
                elif mov.status == 'aprobado_supervisor':
                    mov.status = 'rechazado_pm'
                    mov.aprobado_por_pm = request.user
                    changed = True
                elif mov.status == 'aprobado_pm':
                    mov.status = 'rechazado_finanzas'
                    mov.aprobado_por_finanzas = request.user
                    changed = True

            # Ramas existentes (se mantienen tal cual)
            elif request.user.es_supervisor and mov.status == 'pendiente_supervisor':
                mov.status = 'rechazado_supervisor'
                mov.aprobado_por_supervisor = request.user
                changed = True
            elif request.user.es_pm and mov.status == 'aprobado_supervisor':
                mov.status = 'rechazado_pm'
                mov.aprobado_por_pm = request.user
                changed = True
            elif request.user.es_facturacion and mov.status == 'aprobado_pm':
                mov.status = 'rechazado_finanzas'
                mov.aprobado_por_finanzas = request.user
                changed = True

            if changed:
                mov.motivo_rechazo = motivo
                update_fields = ['status', 'motivo_rechazo']
                if mov.status == 'rechazado_supervisor':
                    update_fields.append('aprobado_por_supervisor')
                elif mov.status == 'rechazado_pm':
                    update_fields.append('aprobado_por_pm')
                elif mov.status == 'rechazado_finanzas':
                    update_fields.append('aprobado_por_finanzas')

                mov.save(update_fields=update_fields)
                messages.success(request, "Gasto rechazado correctamente.")

                if _is_ajax(request):
                    return JsonResponse({
                        "ok": True,
                        "id": mov.pk,
                        "new_status": mov.status,
                        "motivo": mov.motivo_rechazo,
                    })
            else:
                if _is_ajax(request):
                    return JsonResponse({"ok": False, "error": "No autorizado o estado inválido."}, status=403)
                messages.warning(
                    request, "No puedes rechazar este movimiento en su estado actual.")

    next_url = (
        request.POST.get('next')
        or request.GET.get('next')
        or request.META.get('HTTP_REFERER')
        or reverse('facturacion:listar_cartola')
    )
    return redirect(next_url)




from flota.models import Vehicle  # ✅ importar Vehicle


@login_required
@rol_requerido('facturacion', 'admin')
def editar_movimiento(request, pk):
    movimiento = get_object_or_404(CartolaMovimiento, pk=pk)

    if movimiento.tipo and movimiento.tipo.categoria == "abono":
        FormClass = CartolaAbonoForm
        estado_restaurado = 'pendiente_abono_usuario'
    else:
        FormClass = MovimientoUsuarioForm
        estado_restaurado = 'pendiente_supervisor'

    # ✅ helper interno para sincronizar con Flota
    def _sync_movimiento_flota(mov):
        """
        Sincroniza KM/último movimiento del vehículo cuando la rendición es de Servicios.
        No baja el odómetro si el KM editado es menor al actual (strict=False ya lo maneja).
        """
        try:
            # Validar tipo = Servicios
            tipo = getattr(mov, "tipo", None)
            tipo_text = ""
            if tipo:
                tipo_text = (
                    getattr(tipo, "nombre", None)
                    or getattr(tipo, "name", None)
                    or str(tipo)
                    or ""
                ).strip().lower()

            es_servicio = "servicio" in tipo_text
            if not es_servicio:
                return

            vehiculo = getattr(mov, "vehiculo_flota", None)
            km = getattr(mov, "kilometraje_servicio_flota", None)

            if not vehiculo or km in (None, ""):
                return

            km = int(km)

            # Fecha/hora del movimiento para last_movement_at
            fecha = getattr(mov, "fecha_servicio_flota", None) or getattr(mov, "fecha_transaccion", None)
            hora = getattr(mov, "hora_servicio_flota", None)

            if fecha and hora:
                dt_naive = timezone.datetime.combine(fecha, hora)
                dt_mov = timezone.make_aware(dt_naive, timezone.get_current_timezone())
            elif fecha:
                dt_naive = timezone.datetime.combine(fecha, timezone.datetime.min.time())
                dt_mov = timezone.make_aware(dt_naive, timezone.get_current_timezone())
            else:
                dt_mov = timezone.now()

            # ✅ Actualiza odómetro (no baja si km es menor al actual)
            vehiculo.update_kilometraje(
                nuevo_km=km,
                source="rendicion",
                ref=f"Rendición #{mov.pk}",
                strict=False,
            )

            # ✅ Actualiza último movimiento (siempre)
            Vehicle.objects.filter(pk=vehiculo.pk).update(
                last_movement_at=dt_mov,
                updated_at=timezone.now(),
            )

        except Exception:
            # Silencioso para no romper guardado de la rendición
            # Si quieres, aquí puedes loguear con logger.exception(...)
            pass

    if request.method == 'POST':
        # ✅ PASAR USUARIO AL FORM
        form = FormClass(request.POST, request.FILES, instance=movimiento, user=request.user)

        if form.is_valid():
            campos_editados = form.changed_data

            try:
                with transaction.atomic():
                    if campos_editados:
                        movimiento.status = estado_restaurado
                        movimiento.motivo_rechazo = ""

                    movimiento = form.save()

                    # ✅ NUEVO: sincronizar con flota si aplica
                    _sync_movimiento_flota(movimiento)

                messages.success(request, "Movimiento actualizado correctamente.")

            except Exception:
                messages.error(request, "No se pudo actualizar el movimiento.")
                return render(
                    request,
                    'facturacion/editar_movimiento.html',
                    {
                        'form': form,
                        'movimiento': movimiento,
                        'next': (
                            request.POST.get('next')
                            or request.GET.get('next')
                            or request.META.get('HTTP_REFERER')
                            or reverse('facturacion:listar_cartola')
                        ),
                    }
                )

            next_url = (
                request.POST.get('next')
                or request.GET.get('next')
                or request.META.get('HTTP_REFERER')
                or reverse('facturacion:listar_cartola')
            )
            return redirect(next_url)

        messages.error(request, "No se pudo guardar. Revisa los campos marcados en rojo.")

    else:
        # ✅ PASAR USUARIO AL FORM
        form = FormClass(instance=movimiento, user=request.user)

    next_url = (
        request.GET.get('next')
        or request.META.get('HTTP_REFERER')
        or reverse('facturacion:listar_cartola')
    )

    return render(
        request,
        'facturacion/editar_movimiento.html',
        {
            'form': form,
            'movimiento': movimiento,
            'next': next_url,
        }
    )

@login_required
@rol_requerido('admin')
def eliminar_movimiento(request, pk):
    from django.db import transaction
    from django.db.models import Max
    from django.utils import timezone

    from flota.models import VehicleOdometerEvent, VehicleService

    movimiento = get_object_or_404(CartolaMovimiento, pk=pk)

    def _recalcular_km_y_ultimo_movimiento(vehicle_id: int):
        from flota.models import Vehicle

        max_km_service = (
            VehicleService.objects
            .filter(vehicle_id=vehicle_id, kilometraje_declarado__isnull=False)
            .aggregate(m=Max('kilometraje_declarado'))
            .get('m')
        ) or 0

        max_km_event = (
            VehicleOdometerEvent.objects
            .filter(vehicle_id=vehicle_id)
            .aggregate(m=Max('new_km'))
            .get('m')
        ) or 0

        nuevo_km = max(int(max_km_service), int(max_km_event))

        last_service = (
            VehicleService.objects
            .filter(vehicle_id=vehicle_id)
            .order_by('-service_date', '-created_at', '-pk')
            .first()
        )

        last_movement_at = None
        if last_service:
            if last_service.service_time:
                dt_naive = timezone.datetime.combine(last_service.service_date, last_service.service_time)
            else:
                dt_naive = timezone.datetime.combine(last_service.service_date, timezone.datetime.min.time())
            last_movement_at = timezone.make_aware(dt_naive, timezone.get_current_timezone())

        Vehicle.objects.filter(pk=vehicle_id).update(
            kilometraje_actual=nuevo_km,
            last_movement_at=last_movement_at,
            updated_at=timezone.now(),
        )

    if request.method == 'POST':
        try:
            with transaction.atomic():
                servicio = getattr(movimiento, "servicio_flota", None)
                vehiculo_id = getattr(movimiento, "vehiculo_flota_id", None) or (servicio.vehicle_id if servicio else None)

                if servicio:
                    # borrar odometer events del servicio
                    try:
                        if getattr(servicio, "service_code", None) is not None:
                            VehicleOdometerEvent.objects.filter(
                                vehicle_id=servicio.vehicle_id,
                                source="servicio",
                                reference=f"Servicio #{servicio.service_code}"
                            ).delete()
                    except Exception:
                        pass

                    try:
                        servicio.delete()
                    except Exception:
                        pass

                # borrar odometer event de la rendición (si existe)
                if vehiculo_id:
                    try:
                        VehicleOdometerEvent.objects.filter(
                            vehicle_id=vehiculo_id,
                            source="rendicion",
                            reference=f"Rendición #{movimiento.pk}"
                        ).delete()
                    except Exception:
                        pass

                movimiento.delete()

                if vehiculo_id:
                    _recalcular_km_y_ultimo_movimiento(vehiculo_id)

            messages.success(request, "Movimiento eliminado correctamente.")
            return redirect('facturacion:listar_cartola')

        except Exception as e:
            messages.error(request, f"No se pudo eliminar el movimiento correctamente: {e}")
            return redirect('facturacion:listar_cartola')

    return render(request, 'facturacion/eliminar_movimiento.html', {'movimiento': movimiento})


@login_required
@rol_requerido('facturacion', 'admin')
def listar_saldos_usuarios(request):
    # Cantidad por página (máx. 100)
    cantidad_param = (request.GET.get('cantidad') or '5').strip().lower()

    # Agrupar por usuario y calcular rendido y disponible
    saldos = (
        CartolaMovimiento.objects
        .values('usuario__id', 'usuario__first_name', 'usuario__last_name', 'usuario__email')
        .annotate(
            # 👉 cargos SOLO de movimientos que NO están rechazados
            monto_rendido=Sum(
                'cargos',
                filter=~Q(status__startswith='rechazado')
            ),
            monto_asignado=Sum('abonos'),
        )
        .order_by('usuario__first_name')
    )

    # Calcular monto disponible
    for s in saldos:
        s['monto_disponible'] = (s['monto_asignado'] or 0) - (s['monto_rendido'] or 0)

    # Paginación (máx. 100)
    if cantidad_param == 'todos':
        per_page = 100
    else:
        try:
            per_page = max(5, min(int(cantidad_param), 100))
        except ValueError:
            per_page = 5
            cantidad_param = '5'

    paginator = Paginator(saldos, per_page)
    page_number = request.GET.get('page')
    pagina = paginator.get_page(page_number)

    return render(request, 'facturacion/listar_saldos_usuarios.html', {
        'saldos': pagina,
        'pagina': pagina,
        'cantidad': cantidad_param,
    })

@login_required
@rol_requerido('facturacion', 'admin')
def exportar_cartola_finanzas(request):
    # (Opcional) aquí podrías aplicar los mismos filtros que la vista listar_cartola.
    movimientos = CartolaMovimiento.objects.all().order_by('-fecha')

    response = HttpResponse(content_type='application/octet-stream')
    response['Content-Disposition'] = 'attachment; filename="cartola_finanzas.xls"'
    response['X-Content-Type-Options'] = 'nosniff'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Cartola Finanzas')

    header_style = xlwt.easyxf('font: bold on; align: horiz center')
    date_style = xlwt.easyxf(num_format_str='DD-MM-YYYY')

    # ✅ Agregamos campos flota + "Fecha real del gasto"
    columns = [
        "Usuario",
        "Fecha",                 # fecha registro (mov.fecha)
        "Fecha real del gasto",  # mov.fecha_transaccion
        "Vehículo",              # ✅ nuevo
        "Hora servicio",         # ✅ nuevo
        "Kilometraje servicio",  # ✅ nuevo
        "Proyecto",
        "Categoría",
        "Tipo",
        "Tipo servicio (Flota)", # ✅ nuevo
        "RUT Factura",
        "Tipo de Documento",
        "Número de Documento",
        "Observaciones",
        "N° Transferencia",
        "Cargos",
        "Abonos",
        "Status",
    ]
    for col_num, column_title in enumerate(columns):
        ws.write(0, col_num, column_title, header_style)

    for row_num, mov in enumerate(movimientos, start=1):
        # ---- Fecha registro (mov.fecha) ----
        fecha_excel = mov.fecha
        if isinstance(fecha_excel, datetime):
            if is_aware(fecha_excel):
                fecha_excel = fecha_excel.astimezone().replace(tzinfo=None)
            fecha_excel = fecha_excel.date()

        # ---- Fecha real del gasto (mov.fecha_transaccion) ----
        fecha_real_excel = getattr(mov, "fecha_transaccion", None)
        if isinstance(fecha_real_excel, datetime):
            if is_aware(fecha_real_excel):
                fecha_real_excel = fecha_real_excel.astimezone().replace(tzinfo=None)
            fecha_real_excel = fecha_real_excel.date()

        # ---- Hora servicio flota ----
        hora_servicio = getattr(mov, "hora_servicio_flota", None)
        hora_servicio_txt = hora_servicio.strftime("%H:%M") if hora_servicio else ""

        ws.write(row_num, 0, mov.usuario.get_full_name() if mov.usuario else "")
        ws.write(row_num, 1, fecha_excel, date_style)

        if fecha_real_excel:
            ws.write(row_num, 2, fecha_real_excel, date_style)
        else:
            ws.write(row_num, 2, "")

        ws.write(row_num, 3, str(getattr(mov, "vehiculo_flota", "") or ""))
        ws.write(row_num, 4, hora_servicio_txt)
        ws.write(row_num, 5, float(getattr(mov, "kilometraje_servicio_flota", 0) or 0))

        ws.write(row_num, 6, str(getattr(mov, "proyecto", "") or ""))
        ws.write(row_num, 7, getattr(getattr(mov, "tipo", None), "categoria", "") or "")
        ws.write(row_num, 8, str(getattr(mov, "tipo", "") or ""))
        ws.write(row_num, 9, str(getattr(mov, "tipo_servicio_flota", "") or ""))

        ws.write(row_num, 10, getattr(mov, "rut_factura", "") or "")
        ws.write(row_num, 11, getattr(mov, "tipo_doc", "") or "")
        ws.write(row_num, 12, getattr(mov, "numero_doc", "") or "")
        ws.write(row_num, 13, getattr(mov, "observaciones", "") or "")
        ws.write(row_num, 14, getattr(mov, "numero_transferencia", "") or "")
        ws.write(row_num, 15, float(getattr(mov, "cargos", 0) or 0))
        ws.write(row_num, 16, float(getattr(mov, "abonos", 0) or 0))
        ws.write(row_num, 17, mov.get_status_display() if hasattr(mov, "get_status_display") else str(getattr(mov, "status", "") or ""))

    wb.save(response)
    return response



@login_required
@rol_requerido('admin', 'facturacion')  # Solo Admin y Finanzas
def exportar_saldos_disponibles(request):
    # Obtener los mismos datos que usa la vista principal
    from django.db.models import F, Sum, Value

    from facturacion.models import CartolaMovimiento

    saldos = (CartolaMovimiento.objects
              .values('usuario__first_name', 'usuario__last_name')
              .annotate(
                  monto_rendido=Sum('cargos', default=0),
                  monto_disponible=Sum(F('abonos') - F('cargos'), default=0)
              ).order_by('usuario__first_name'))

    # Crear respuesta como Excel
    response = HttpResponse(content_type='application/octet-stream')
    response['Content-Disposition'] = 'attachment; filename="saldos_disponibles.xls"'
    response['X-Content-Type-Options'] = 'nosniff'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Saldos Disponibles')

    header_style = xlwt.easyxf('font: bold on; align: horiz center')
    currency_style = xlwt.easyxf(num_format_str='#,##0')

    # Cabeceras
    columns = ["Usuario", "Monto Rendido", "Monto Disponible"]
    for col_num, column_title in enumerate(columns):
        ws.write(0, col_num, column_title, header_style)

    # Filas
    for row_num, s in enumerate(saldos, start=1):
        usuario = f"{s['usuario__first_name']} {s['usuario__last_name']}"
        ws.write(row_num, 0, usuario)
        ws.write(row_num, 1, float(s['monto_rendido'] or 0), currency_style)
        ws.write(row_num, 2, float(s['monto_disponible'] or 0), currency_style)

    wb.save(response)
    return response


from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.utils import timezone


@login_required
@rol_requerido('facturacion', 'admin')
@require_POST
@csrf_protect
def enviar_a_historial(request):
    """
    Recibe ids[] por POST y marca esos movimientos como en_historial=True.

    Permite mover al historial:
    - Rendiciones normales: status='aprobado_finanzas'
    - Abonos: status='aprobado_abono_usuario'

    Solo si aún no están en historial.
    """
    ids = request.POST.getlist("ids[]") or []
    ids = [int(i) for i in ids if str(i).isdigit()]

    if not ids:
        return JsonResponse({"ok": False, "error": "No se recibieron IDs."}, status=400)

    estados_permitidos_historial = [
        'aprobado_finanzas',       # rendición normal
        'aprobado_abono_usuario',  # abono
    ]

    # Traemos todos los seleccionados (para diagnosticar)
    seleccionados = CartolaMovimiento.objects.filter(id__in=ids)

    if not seleccionados.exists():
        return JsonResponse({
            "ok": False,
            "error": "No se encontraron movimientos con los IDs enviados."
        }, status=404)

    # Filtramos los que sí se pueden mover
    qs_validos = seleccionados.filter(
        status__in=estados_permitidos_historial,
        en_historial=False,
    )

    moved_count = qs_validos.count()

    # Si no hay ninguno válido, devolvemos detalle
    if moved_count == 0:
        resumen = list(
            seleccionados.values("id", "status", "en_historial")
        )

        return JsonResponse({
            "ok": False,
            "code": "NO_FINALIZADO_PARA_HISTORIAL",
            "error": (
                "No se puede enviar al historial porque los movimientos seleccionados "
                "no están finalizados (Aprobado por Finanzas o Abono aprobado por Usuario), "
                "o ya fueron enviados al historial."
            ),
            "detalle": resumen,  # útil para depurar en consola
        }, status=400)

    # Movemos solo los válidos
    qs_validos.update(
        en_historial=True,
        historial_enviado_el=timezone.now(),
        historial_enviado_por=request.user
    )

    # Cantidad no movida (si hubo mezcla)
    not_moved_count = len(ids) - moved_count

    return JsonResponse({
        "ok": True,
        "moved": moved_count,
        "not_moved": not_moved_count,
        "message": (
            f"Se enviaron {moved_count} movimiento(s) al historial."
            + (f" {not_moved_count} no calificaban para historial." if not_moved_count > 0 else "")
        )
    }, status=200)



@login_required
@rol_requerido('facturacion', 'admin')
def listar_cartola_historial(request):
    """
    Vista de historial (idéntica en look), pero muestra SOLO lo que ya está en_historial=True.
    """
    import json
    import re
    from datetime import datetime

    from django.contrib import messages
    from django.core.paginator import Paginator
    from django.db.models import CharField, Q
    from django.db.models.functions import Cast
    from django.utils import timezone

    # --- Cantidad/paginación (mantén string para la UI)
    cantidad_param = request.GET.get('cantidad', '10')

    if cantidad_param == 'todos':
        page_size = 100
    else:
        try:
            page_size = max(5, min(int(cantidad_param), 100))
        except ValueError:
            page_size = 10
            cantidad_param = '10'

    # ---------- Filtros tipo Excel recibidos por GET ----------
    params = request.GET.copy()
    excel_filters_raw = (params.get('excel_filters') or '').strip()
    try:
        excel_filters = json.loads(excel_filters_raw) if excel_filters_raw else {}
    except json.JSONDecodeError:
        excel_filters = {}

    # --- Filtros clásicos
    usuario = (request.GET.get('usuario') or '').strip()
    fecha_txt = (request.GET.get('fecha') or '').strip()
    proyecto = (request.GET.get('proyecto') or '').strip()
    categoria = (request.GET.get('categoria') or '').strip()
    tipo = (request.GET.get('tipo') or '').strip()
    rut_factura = (request.GET.get('rut_factura') or '').strip()
    estado = (request.GET.get('estado') or '').strip()

    # --- Base queryset (SOLO HISTORIAL)
    movimientos = (
        CartolaMovimiento.objects
        .select_related(
            'usuario',
            'proyecto',
            'tipo',
            'vehiculo_flota',
            'tipo_servicio_flota',
        )
        .filter(en_historial=True)
    )

    movimientos = movimientos.annotate(
        fecha_iso=Cast('fecha', CharField())
    )

    if usuario:
        movimientos = movimientos.filter(
            Q(usuario__username__icontains=usuario) |
            Q(usuario__first_name__icontains=usuario) |
            Q(usuario__last_name__icontains=usuario)
        )

    if fecha_txt:
        fecha_normalizada = fecha_txt.replace('/', '-').strip()
        m = re.match(r'^(\d{1,2})(?:-(\d{1,2}))?(?:-(\d{1,4}))?$', fecha_normalizada)
        if m:
            dia_str = m.group(1)
            mes_str = m.group(2)
            anio_str = m.group(3)

            try:
                q_fecha = Q()
                dia = int(dia_str)
                q_fecha &= Q(fecha__day=dia)

                if mes_str:
                    mes = int(mes_str)
                    q_fecha &= Q(fecha__month=mes)

                if anio_str and len(anio_str) == 4:
                    anio = int(anio_str)
                    q_fecha &= Q(fecha__year=anio)

                movimientos = movimientos.filter(q_fecha)

            except ValueError:
                messages.warning(request, "Formato de fecha inválido. Use DD, DD-MM o DD-MM-YYYY.")
        else:
            messages.warning(request, "Formato de fecha inválido. Use DD, DD-MM o DD-MM-YYYY.")

    if proyecto:
        movimientos = movimientos.filter(proyecto__nombre__icontains=proyecto)
    if categoria:
        movimientos = movimientos.filter(tipo__categoria__icontains=categoria)
    if tipo:
        movimientos = movimientos.filter(tipo__nombre__icontains=tipo)
    if rut_factura:
        movimientos = movimientos.filter(rut_factura__icontains=rut_factura)
    if estado:
        movimientos = movimientos.filter(status=estado)

    # Orden: más recientes arriba
    movimientos = movimientos.order_by('-fecha', '-id')

    # ============================================================
    # Excel filters (python) + excel_global_json
    # ============================================================
    movimientos_list = list(movimientos)

    def _norm(s):
        s = "" if s is None else str(s)
        s = s.replace("\u00A0", " ")
        s = re.sub(r"\s+", " ", s).strip().lower()
        return s

    def format_clp(n):
        try:
            n = 0 if n is None else n
            n_int = int(n)
            return f"${n_int:,}".replace(",", ".")
        except Exception:
            return "$0"

    def format_km_cell(n):
        # debe calzar con template: "{{ mov.kilometraje_servicio_flota|miles }} KM" o "—"
        try:
            if n is None or n == "":
                return "—"
            n_int = int(float(n))
            return f"{n_int:,}".replace(",", ".") + " KM"
        except Exception:
            return "—"

    def _fmt_fecha_cell(dt_or_d):
        # debe calzar con date:'d-m-Y' del template (soporta DateTime/Date)
        if not dt_or_d:
            return ""
        try:
            if isinstance(dt_or_d, datetime):
                if timezone.is_aware(dt_or_d):
                    dt_or_d = timezone.localtime(dt_or_d)
                return dt_or_d.strftime("%d-%m-%Y")
            return dt_or_d.strftime("%d-%m-%Y")
        except Exception:
            return ""

    def _fmt_hora_cell(t):
        # debe calzar con time:'H:i' del template o "—"
        if not t:
            return "—"
        try:
            return t.strftime("%H:%M")
        except Exception:
            return str(t)

    def _fmt_obs_cell(obs):
        # debe calzar con default:'(Vacías)'
        s = (obs or "").strip()
        return s if s else "(Vacías)"

    def _status_cell(mov):
        # debe calzar con data-excel-value="{{ mov.get_status_display }}"
        try:
            return mov.get_status_display() if getattr(mov, "status", None) else ""
        except Exception:
            return ""

    if excel_filters:
        excel_filters_norm = {}
        for col, values in (excel_filters or {}).items():
            if not values:
                continue
            try:
                excel_filters_norm[str(col)] = set(_norm(v) for v in values)
            except Exception:
                continue

        def matches_excel_filters(mov):
            for col, allowed_norm in excel_filters_norm.items():
                if not allowed_norm:
                    continue

                # Índices tabla historial (CON checkbox):
                # 0  Checkbox
                # 1  Usuario
                # 2  Fecha
                # 3  Fecha real del gasto
                # 4  Vehículo
                # 5  Hora servicio
                # 6  Kilometraje servicio
                # 7  Proyecto
                # 8  Categoría
                # 9  Tipo
                # 10 Tipo servicio (Flota)
                # 11 RUT factura
                # 12 Tipo de documento
                # 13 N° Documento
                # 14 Observaciones
                # 15 N° Transferencia
                # 16 Comprobante
                # 17 Cargos
                # 18 Abonos
                # 19 Status

                if col == "1":
                    label = str(mov.usuario) if mov.usuario else ""

                elif col == "2":
                    label = _fmt_fecha_cell(getattr(mov, "fecha", None))

                elif col == "3":
                    d = getattr(mov, "fecha_transaccion", None) or getattr(mov, "fecha", None)
                    label = _fmt_fecha_cell(d)

                elif col == "4":
                    label = str(getattr(mov, "vehiculo_flota", None) or "—").strip() or "—"

                elif col == "5":
                    label = _fmt_hora_cell(getattr(mov, "hora_servicio_flota", None))

                elif col == "6":
                    label = format_km_cell(getattr(mov, "kilometraje_servicio_flota", None))

                elif col == "7":
                    label = str(mov.proyecto) if mov.proyecto else ""

                elif col == "8":
                    if mov.tipo and getattr(mov.tipo, "categoria", None):
                        label = (mov.tipo.categoria or "").title()
                    else:
                        label = ""

                elif col == "9":
                    label = str(mov.tipo) if mov.tipo else ""

                elif col == "10":
                    label = str(getattr(mov, "tipo_servicio_flota", None) or "—").strip() or "—"

                elif col == "11":
                    label = (getattr(mov, "rut_factura", None) or "—").strip() or "—"

                elif col == "12":
                    label = (getattr(mov, "tipo_doc", None) or "—").strip() or "—"

                elif col == "13":
                    label = (getattr(mov, "numero_doc", None) or "—").strip() or "—"

                elif col == "14":
                    label = _fmt_obs_cell(getattr(mov, "observaciones", None))

                elif col == "15":
                    label = (getattr(mov, "numero_transferencia", None) or "—").strip() or "—"

                elif col == "16":
                    label = "Ver" if getattr(mov, "comprobante", None) else "—"

                elif col == "17":
                    label = format_clp(getattr(mov, "cargos", 0) or 0)

                elif col == "18":
                    label = format_clp(getattr(mov, "abonos", 0) or 0)

                elif col == "19":
                    label = _status_cell(mov)

                else:
                    continue

                if _norm(label) not in allowed_norm:
                    return False

            return True

        movimientos_list = [m for m in movimientos_list if matches_excel_filters(m)]

    # ==========================
    # excel_global_json (solo historial)
    # ==========================
    excel_global = {}

    excel_global[1] = sorted({str(m.usuario) for m in movimientos_list if m.usuario})

    excel_global[2] = sorted({
        _fmt_fecha_cell(getattr(m, "fecha", None))
        for m in movimientos_list
        if getattr(m, "fecha", None)
    })

    excel_global[3] = sorted({
        _fmt_fecha_cell(getattr(m, "fecha_transaccion", None) or getattr(m, "fecha", None))
        for m in movimientos_list
        if (getattr(m, "fecha_transaccion", None) or getattr(m, "fecha", None))
    })

    excel_global[4] = sorted({
        (str(getattr(m, "vehiculo_flota", None) or "—").strip() or "—")
        for m in movimientos_list
    })

    excel_global[5] = sorted({
        _fmt_hora_cell(getattr(m, "hora_servicio_flota", None))
        for m in movimientos_list
    })

    excel_global[6] = sorted({
        format_km_cell(getattr(m, "kilometraje_servicio_flota", None))
        for m in movimientos_list
    })

    excel_global[7] = sorted({str(m.proyecto) for m in movimientos_list if m.proyecto})

    excel_global[8] = sorted({
        (m.tipo.categoria or "").title()
        for m in movimientos_list
        if m.tipo and getattr(m.tipo, "categoria", None)
    })

    excel_global[9] = sorted({str(m.tipo) for m in movimientos_list if m.tipo})

    excel_global[10] = sorted({
        (str(getattr(m, "tipo_servicio_flota", None) or "—").strip() or "—")
        for m in movimientos_list
    })

    excel_global[11] = sorted({
        (getattr(m, "rut_factura", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    excel_global[12] = sorted({
        (getattr(m, "tipo_doc", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    excel_global[13] = sorted({
        (getattr(m, "numero_doc", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    excel_global[14] = sorted({
        _fmt_obs_cell(getattr(m, "observaciones", None))
        for m in movimientos_list
    })

    excel_global[15] = sorted({
        (getattr(m, "numero_transferencia", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    excel_global[16] = sorted({
        "Ver" if getattr(m, "comprobante", None) else "—"
        for m in movimientos_list
    })

    excel_global[17] = sorted({
        format_clp(getattr(m, "cargos", 0) or 0)
        for m in movimientos_list
    })

    excel_global[18] = sorted({
        format_clp(getattr(m, "abonos", 0) or 0)
        for m in movimientos_list
    })

    excel_global[19] = sorted({
        _status_cell(m)
        for m in movimientos_list
        if getattr(m, "status", None)
    })

    excel_global_json = json.dumps(excel_global)

    # ============================================================
    # ✅ PAGINACIÓN
    # ✅ FIX: si hay excel_filters activos, forzar page=1
    # ============================================================
    paginator = Paginator(movimientos_list, page_size)

    page_number = request.GET.get('page') or '1'
    if excel_filters and str(page_number) != '1':
        page_number = '1'

    pagina = paginator.get_page(page_number)

    estado_choices = CartolaMovimiento.ESTADOS
    filtros = {
        'usuario': usuario,
        'fecha': fecha_txt,
        'proyecto': proyecto,
        'categoria': categoria,
        'tipo': tipo,
        'rut_factura': rut_factura,
        'estado': estado,
    }

    params_no_page = params.copy()
    params_no_page.pop('page', None)
    base_qs = params_no_page.urlencode()
    full_qs = params.urlencode()

    return render(
        request,
        'facturacion/listar_cartola_historial.html',
        {
            'pagina': pagina,
            'cantidad': cantidad_param,
            'estado_choices': estado_choices,
            'filtros': filtros,
            'excel_global_json': excel_global_json,
            'base_qs': base_qs,
            'full_qs': full_qs,
        }
    )



@login_required
@rol_requerido('facturacion', 'admin')
@require_POST
@csrf_protect
def devolver_a_cartola(request):
    """
    Recibe ids[] por POST y devuelve esos movimientos desde Historial a Cartola.
    - en_historial=False
    - limpia historial_enviado_el / historial_enviado_por
    """
    ids = request.POST.getlist("ids[]") or []
    ids = [i for i in ids if str(i).isdigit()]

    if not ids:
        return JsonResponse({"ok": False, "error": "No se recibieron IDs."}, status=400)

    qs = CartolaMovimiento.objects.filter(id__in=ids, en_historial=True)

    count = qs.count()
    if count == 0:
        return JsonResponse({"ok": False, "error": "No hay registros válidos en historial para devolver."}, status=400)

    qs.update(
        en_historial=False,
        historial_enviado_el=None,
        historial_enviado_por=None,
    )

    return JsonResponse({"ok": True, "restored": count})

@login_required
@rol_requerido('facturacion', 'admin')
def exportar_cartola_historial(request):
    import json
    import re
    from datetime import datetime

    import xlwt
    from django.db.models import CharField, Q
    from django.db.models.functions import Cast
    from django.http import HttpResponse
    from django.utils.timezone import is_aware

    from .models import CartolaMovimiento

    # ---------- Params (copiamos GET, pero sin page) ----------
    params = request.GET.copy()
    params.pop('page', None)

    # Excel filters (JSON)
    excel_filters_raw = (params.get('excel_filters') or '').strip()
    try:
        excel_filters = json.loads(excel_filters_raw) if excel_filters_raw else {}
    except json.JSONDecodeError:
        excel_filters = {}

    # --- Filtros (string trimming)
    usuario = (params.get('usuario') or '').strip()
    fecha_txt = (params.get('fecha') or '').strip()
    proyecto = (params.get('proyecto') or '').strip()
    categoria = (params.get('categoria') or '').strip()
    tipo = (params.get('tipo') or '').strip()
    rut_factura = (params.get('rut_factura') or '').strip()
    estado = (params.get('estado') or '').strip()

    # ============================================================
    # Base queryset: SOLO HISTORIAL
    # ============================================================
    movimientos = CartolaMovimiento.objects.filter(en_historial=True)

    movimientos = movimientos.annotate(
        fecha_iso=Cast('fecha', CharField())
    )

    if usuario:
        movimientos = movimientos.filter(
            Q(usuario__username__icontains=usuario) |
            Q(usuario__first_name__icontains=usuario) |
            Q(usuario__last_name__icontains=usuario)
        )

    if fecha_txt:
        fecha_normalizada = fecha_txt.replace('/', '-').strip()
        m = re.match(r'^(\d{1,2})(?:-(\d{1,2}))?(?:-(\d{1,4}))?$', fecha_normalizada)
        if m:
            dia_str = m.group(1)
            mes_str = m.group(2)
            anio_str = m.group(3)

            try:
                q_fecha = Q()
                dia = int(dia_str)
                q_fecha &= Q(fecha__day=dia)

                if mes_str:
                    mes = int(mes_str)
                    q_fecha &= Q(fecha__month=mes)

                if anio_str and len(anio_str) == 4:
                    anio = int(anio_str)
                    q_fecha &= Q(fecha__year=anio)

                movimientos = movimientos.filter(q_fecha)

            except ValueError:
                pass
        else:
            pass

    if proyecto:
        movimientos = movimientos.filter(proyecto__nombre__icontains=proyecto)
    if categoria:
        movimientos = movimientos.filter(tipo__categoria__icontains=categoria)
    if tipo:
        movimientos = movimientos.filter(tipo__nombre__icontains=tipo)
    if rut_factura:
        movimientos = movimientos.filter(rut_factura__icontains=rut_factura)
    if estado:
        movimientos = movimientos.filter(status=estado)

    # Orden (igual al historial)
    movimientos = movimientos.order_by('-fecha')

    movimientos_list = list(movimientos)

    # ============================================================
    # Excel filters (python) — actualizado con nuevos campos
    # ============================================================
    def format_clp(n):
        try:
            n = 0 if n is None else n
            n_int = int(n)
            return f"${n_int:,}".replace(",", ".")
        except Exception:
            return "$0"

    def format_km(n):
        try:
            if n is None or n == "":
                return "—"
            n_int = int(float(n))
            return f"{n_int:,}".replace(",", ".") + " KM"
        except Exception:
            return "—"

    if excel_filters:
        def matches_excel_filters(mov):
            for col, values in excel_filters.items():
                if not values:
                    continue
                values_set = set(values)

                # Índices tabla historial (CON checkbox):
                # 0  Checkbox (no)
                # 1  Usuario
                # 2  Fecha
                # 3  Fecha real del gasto
                # 4  Vehículo
                # 5  Hora servicio
                # 6  Kilometraje servicio
                # 7  Proyecto
                # 8  Categoría
                # 9  Tipo
                # 10 Tipo servicio (Flota)
                # 11 RUT factura
                # 12 Tipo de documento
                # 13 N° Documento
                # 14 Observaciones
                # 15 N° Transferencia
                # 16 Comprobante
                # 17 Cargos
                # 18 Abonos
                # 19 Status

                if col == "1":
                    label = str(mov.usuario) if mov.usuario else ""
                elif col == "2":
                    d = getattr(mov, "fecha", None)
                    label = d.strftime("%d-%m-%Y") if d else ""
                elif col == "3":
                    d = getattr(mov, "fecha_transaccion", None) or getattr(mov, "fecha", None)
                    label = d.strftime("%d-%m-%Y") if d else ""
                elif col == "4":
                    label = str(getattr(mov, "vehiculo_flota", None) or "—").strip() or "—"
                elif col == "5":
                    h = getattr(mov, "hora_servicio_flota", None)
                    label = h.strftime("%H:%M") if h else "—"
                elif col == "6":
                    label = format_km(getattr(mov, "kilometraje_servicio_flota", None))
                elif col == "7":
                    label = str(mov.proyecto) if mov.proyecto else ""
                elif col == "8":
                    if mov.tipo and getattr(mov.tipo, "categoria", None):
                        label = (mov.tipo.categoria or "").title()
                    else:
                        label = ""
                elif col == "9":
                    label = str(mov.tipo) if mov.tipo else ""
                elif col == "10":
                    label = str(getattr(mov, "tipo_servicio_flota", None) or "—").strip() or "—"
                elif col == "11":
                    label = (getattr(mov, "rut_factura", None) or "—").strip() or "—"
                elif col == "12":
                    label = (getattr(mov, "tipo_doc", None) or "—").strip() or "—"
                elif col == "13":
                    label = (getattr(mov, "numero_doc", None) or "—").strip() or "—"
                elif col == "14":
                    label = (getattr(mov, "observaciones", None) or "").strip() or "(Vacías)"
                elif col == "15":
                    label = (getattr(mov, "numero_transferencia", None) or "—").strip() or "—"
                elif col == "16":
                    label = "Ver" if getattr(mov, "comprobante", None) else "—"
                elif col == "17":
                    label = format_clp(getattr(mov, "cargos", 0) or 0)
                elif col == "18":
                    label = format_clp(getattr(mov, "abonos", 0) or 0)
                elif col == "19":
                    label = mov.get_status_display() if getattr(mov, "status", None) else ""
                else:
                    continue

                if label not in values_set:
                    return False

            return True

        movimientos_list = [m for m in movimientos_list if matches_excel_filters(m)]

    # ============================================================
    # Excel (xlwt)
    # ============================================================
    response = HttpResponse(content_type='application/octet-stream')
    response['Content-Disposition'] = 'attachment; filename="historial_cartola.xls"'
    response['X-Content-Type-Options'] = 'nosniff'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Historial Cartola')

    header_style = xlwt.easyxf('font: bold on; align: horiz center')
    date_style = xlwt.easyxf(num_format_str='DD-MM-YYYY')
    currency_style = xlwt.easyxf(num_format_str='#,##0')

    columns = [
        "Usuario",
        "Fecha",
        "Fecha real del gasto",
        "Vehículo",               # ✅ nuevo
        "Hora servicio",          # ✅ nuevo
        "Kilometraje servicio",   # ✅ nuevo
        "Proyecto",
        "Categoría",
        "Tipo",
        "Tipo servicio (Flota)",  # ✅ nuevo
        "RUT Factura",
        "Tipo de Documento",
        "Número de Documento",
        "Observaciones",
        "N° Transferencia",
        "Comprobante",
        "Cargos",
        "Abonos",
        "Status",
    ]

    for col_num, title in enumerate(columns):
        ws.write(0, col_num, title, header_style)

    for row_num, mov in enumerate(movimientos_list, start=1):
        # Fecha registro (mov.fecha) -> date
        fecha_excel = getattr(mov, "fecha", None)
        if isinstance(fecha_excel, datetime):
            if is_aware(fecha_excel):
                fecha_excel = fecha_excel.astimezone().replace(tzinfo=None)
            fecha_excel = fecha_excel.date()

        # Fecha real del gasto (fecha_transaccion) -> date
        fecha_real = getattr(mov, "fecha_transaccion", None)
        if isinstance(fecha_real, datetime):
            if is_aware(fecha_real):
                fecha_real = fecha_real.astimezone().replace(tzinfo=None)
            fecha_real = fecha_real.date()

        # Comprobante
        comp = getattr(mov, "comprobante", None)
        comp_val = comp.url if comp else "—"

        # Hora servicio
        hora_servicio = getattr(mov, "hora_servicio_flota", None)
        hora_servicio_txt = hora_servicio.strftime("%H:%M") if hora_servicio else "—"

        ws.write(row_num, 0, mov.usuario.get_full_name() if mov.usuario else "")
        ws.write(row_num, 1, fecha_excel if fecha_excel else "", date_style)
        ws.write(row_num, 2, fecha_real if fecha_real else "", date_style)

        ws.write(row_num, 3, str(getattr(mov, "vehiculo_flota", "") or "—"))
        ws.write(row_num, 4, hora_servicio_txt)
        ws.write(row_num, 5, float(getattr(mov, "kilometraje_servicio_flota", 0) or 0))

        ws.write(row_num, 6, str(getattr(mov, "proyecto", "") or ""))
        ws.write(row_num, 7, getattr(getattr(mov, "tipo", None), "categoria", "") or "")
        ws.write(row_num, 8, str(getattr(mov, "tipo", "") or ""))
        ws.write(row_num, 9, str(getattr(mov, "tipo_servicio_flota", "") or "—"))

        ws.write(row_num, 10, (getattr(mov, "rut_factura", "") or "—"))
        ws.write(row_num, 11, (getattr(mov, "tipo_doc", "") or "—"))
        ws.write(row_num, 12, (getattr(mov, "numero_doc", "") or "—"))
        ws.write(row_num, 13, (getattr(mov, "observaciones", "") or ""))
        ws.write(row_num, 14, (getattr(mov, "numero_transferencia", "") or "—"))
        ws.write(row_num, 15, comp_val)

        ws.write(row_num, 16, float(getattr(mov, "cargos", 0) or 0), currency_style)
        ws.write(row_num, 17, float(getattr(mov, "abonos", 0) or 0), currency_style)

        ws.write(
            row_num,
            18,
            mov.get_status_display() if hasattr(mov, "get_status_display") else str(getattr(mov, "status", "") or "")
        )

    wb.save(response)
    return response


