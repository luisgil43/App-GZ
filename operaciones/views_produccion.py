# operaciones/views_produccion.py
# si no lo tienes, añade este import
import datetime as dt
import io
import logging
import mimetypes
import os
import re
import uuid
from datetime import date, datetime
from decimal import Decimal
from urllib.parse import urlencode
from uuid import uuid4

import boto3
from botocore.client import Config
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model  # 👈 agregado
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import F, IntegerField, Max, Q, Sum, Value
from django.db.models.functions import Cast, Replace
from django.http import (Http404, HttpResponse, HttpResponseBadRequest,
                         HttpResponseForbidden, JsonResponse)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_http_methods, require_POST
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate, Spacer,
                                Table, TableStyle)

from usuarios.decoradores import rol_requerido

from .models import MonthlyPayment  # <-- CAMBIA esto por tu modelo real
from .models import ServicioCotizado
from .utils.http import login_required_json

# === Estados producción + ajustes (usar en TODO el módulo) ===
EST_PROD = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}
AJ_POS = {"ajuste_bono", "ajuste_adelanto"}   # suman
AJ_NEG = {"ajuste_descuento"}                 # restan
ESTADOS_PROD_Y_AJUSTES = EST_PROD | AJ_POS | AJ_NEG


def _monto_firmado_por_estado(monto, estado: str) -> Decimal:
    """
    Normaliza el signo según el estado:
      - descuento => negativo
      - bono/adelanto => positivo
      - producción normal => tal cual
    """
    base = Decimal(str(monto or 0))
    if estado in AJ_NEG:
        return -abs(base)
    if estado in AJ_POS:
        return abs(base)
    return base


@login_required
def produccion_admin(request):
    """
    Producción aprobada (por técnico) + ajustes (bono/adelanto/descuento).
    - Filas por técnico asignado; monto prorrateado: monto_mmoo / N_tecnicos
    - Filtros: du, id_claro, id_new, mes_produccion, estado (opcional),
               f_tecnico (nombre/usuario) y f_region (opcional).
    """
    def _month_token(dt) -> str:
        d = dt.date() if hasattr(dt, "date") else dt
        return f"{d.year:04d}-{d.month:02d}"

    current_month = _yyyy_mm(timezone.localdate())

    # ---------- Filtros GET ----------
    filtros = {
        "du": (request.GET.get("du") or "").strip(),
        "id_claro": (request.GET.get("id_claro") or "").strip(),
        "id_new": (request.GET.get("id_new") or "").strip(),
        "mes_produccion": (request.GET.get("mes_produccion") or "").strip(),
        "estado": (request.GET.get("estado") or "").strip(),
    }
    f_tecnico = (request.GET.get("f_tecnico") or "").strip()
    f_region = (request.GET.get("f_region") or "").strip()
    cantidad = request.GET.get("cantidad", "10")

    # ✅ incluir también los ajustes
    estados_ok = {
        "aprobado_supervisor",
        "aprobado_pm",
        "aprobado_finanzas",
        "ajuste_bono",
        "ajuste_adelanto",
        "ajuste_descuento",
    }

    qs = (
        ServicioCotizado.objects
        .filter(estado__in=estados_ok)
        .prefetch_related('trabajadores_asignados')
        .order_by('-fecha_aprobacion_supervisor', '-id')
    )

    # ---------- Aplicar filtros de cabecera ----------
    if filtros["du"]:
        # Normaliza: "DU000062", "du 62", "62" -> "62"
        raw_du = filtros["du"]
        m = re.search(r"du\s*0*([0-9]+)", raw_du, re.IGNORECASE)
        term = m.group(1) if m else re.sub(r"\D+", "", raw_du)
        if term:
            qs = qs.filter(du__icontains=term)

    if filtros["id_claro"]:
        qs = qs.filter(id_claro__icontains=filtros["id_claro"])
    if filtros["id_new"]:
        qs = qs.filter(id_new__icontains=filtros["id_new"])
    if filtros["mes_produccion"]:
        qs = qs.filter(mes_produccion__icontains=filtros["mes_produccion"])
    if filtros["estado"]:
        if filtros["estado"] in estados_ok:
            qs = qs.filter(estado=filtros["estado"])
        else:
            qs = qs.none()
    if f_region:
        qs = qs.filter(region__icontains=f_region)

    # ---------- Construir filas por técnico (prorrateo) ----------
    filas = []
    for s in qs:
        asignados = list(s.trabajadores_asignados.all())
        n = len(asignados) or 1
        total = Decimal(str(s.monto_mmoo or 0))
        parte = (total / n).quantize(Decimal('0.01'))

        if asignados:
            for tec in asignados:
                if f_tecnico:
                    target = f_tecnico.lower()
                    full_name = ((tec.first_name or "") + " " +
                                 (tec.last_name or "")).strip().lower()
                    username = (tec.username or "").lower()
                    if target not in full_name and target not in username:
                        continue

                filas.append({
                    "id": s.id,  # 👈 necesario para acciones
                    "du": s.du,
                    "id_claro": s.id_claro or "",
                    "region": s.region or "",
                    "mes_produccion": s.mes_produccion or "",
                    "id_new": s.id_new or "",
                    "detalle_tarea": s.detalle_tarea or "",
                    "monto_tecnico": parte,
                    "tecnico": tec,
                    "fecha_fin": s.fecha_aprobacion_supervisor,
                    "status": s.estado,
                })
        else:
            filas.append({
                "id": s.id,  # 👈 necesario también aquí
                "du": s.du,
                "id_claro": s.id_claro or "",
                "region": s.region or "",
                "mes_produccion": s.mes_produccion or "",
                "id_new": s.id_new or "",
                "detalle_tarea": s.detalle_tarea or "",
                "monto_tecnico": parte,
                "tecnico": None,
                "fecha_fin": s.fecha_aprobacion_supervisor,
                "status": s.estado,
            })

    # ---------- Ordenar: mes actual hacia abajo ----------
    MONTHS_ES = {
        "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
        "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9, "octubre": 10,
        "noviembre": 11, "diciembre": 12,
    }

    def parse_mes_token(txt: str) -> int | None:
        """
        Devuelve YYYYMM (p.ej. 202509) a partir de mes_produccion.
        Acepta: '2025-09', '2025/9', '09-2025', 'septiembre 2025', 'Sep 2025', etc.
        """
        if not txt:
            return None
        s = str(txt).strip().lower()

        m = re.search(r"(\d{4})[/-](\d{1,2})", s)  # YYYY-MM o YYYY/MM
        if m:
            y, mm = int(m.group(1)), int(m.group(2))
            return y * 100 + mm

        m = re.search(r"(\d{1,2})[/-](\d{4})", s)  # MM-YYYY o MM/YYYY
        if m:
            mm, y = int(m.group(1)), int(m.group(2))
            return y * 100 + mm

        m = re.search(r"([a-záéíóúü]{3,})\s+(\d{4})", s)  # nombre mes + año
        if m:
            name = m.group(1)
            y = int(m.group(2))
            for full, num in MONTHS_ES.items():
                if name.startswith(full[:3]):
                    return y * 100 + num
        return None

    for r in filas:
        ym = parse_mes_token(r["mes_produccion"])
        r["_ym_key"] = ym if ym is not None else -1

        # Normalizar fecha_fin a date para evitar comparar datetime vs date
        fin = r["fecha_fin"]
        if isinstance(fin, datetime):
            dkey = fin.date()
        elif isinstance(fin, date):
            dkey = fin
        else:
            dkey = date(1900, 1, 1)
        r["_date_key"] = dkey

    # Orden final: mes (YYYYMM) desc y, a igualdad, fecha_fin desc
    filas.sort(key=lambda x: (x["_ym_key"], x["_date_key"]), reverse=True)

    # ---------- Paginación ----------
    if cantidad != "todos":
        try:
            per_page = max(5, min(int(cantidad), 100))
        except ValueError:
            per_page = 10
        paginator = Paginator(filas, per_page)
        page_number = request.GET.get("page") or 1
        pagina = paginator.get_page(page_number)
    else:
        pagina = filas

    estado_choices = [
        ('aprobado_supervisor', 'Aprobado por Supervisor'),
        ('aprobado_pm', 'Aprobado por PM'),
        ('aprobado_finanzas', 'Aprobado por Finanzas'),
        ('ajuste_bono', '🎁 Bono'),
        ('ajuste_adelanto', '⏩ Adelanto'),
        ('ajuste_descuento', '➖ Descuento'),
    ]

    filters_dict = {
        **filtros,
        "f_tecnico": f_tecnico,
        "f_region": f_region,
        "cantidad": cantidad,
    }
    filters_qs = urlencode({k: v for k, v in filters_dict.items() if v})

    # Usuarios activos para el modal de Ajuste
    User = get_user_model()
    usuarios = User.objects.filter(is_active=True).order_by(
        'first_name', 'last_name', 'username')

    return render(request, "operaciones/produccion_admin.html", {
        "current_month": current_month,
        "pagina": pagina,
        "cantidad": cantidad,
        "filtros": filtros,
        "estado_choices": estado_choices,
        "f_tecnico": f_tecnico,
        "f_region": f_region,
        "filters_qs": filters_qs,
        "usuarios": usuarios,
    })


@login_required
def exportar_produccion_admin(request):
    """
    Export a Excel (respeta filtros). Una fila por técnico con monto prorrateado,
    incluyendo ajustes (bono/adelanto/descuento).
    """
    filtros = {
        "du": (request.GET.get("du") or "").strip(),
        "id_claro": (request.GET.get("id_claro") or "").strip(),
        "id_new": (request.GET.get("id_new") or "").strip(),
        "mes_produccion": (request.GET.get("mes_produccion") or "").strip(),
        "estado": (request.GET.get("estado") or "").strip(),
    }
    f_tecnico = (request.GET.get("f_tecnico") or "").strip()
    f_region = (request.GET.get("f_region") or "").strip()

    # ⬅️ incluir producción + ajustes
    estados_ok = ESTADOS_PROD_Y_AJUSTES

    qs = (
        ServicioCotizado.objects
        .filter(estado__in=estados_ok)
        .prefetch_related('trabajadores_asignados')
        .order_by('-fecha_aprobacion_supervisor', '-id')
    )

    # Filtros
    if filtros["du"]:
        raw_du = filtros["du"]
        m = re.search(r"du\s*0*([0-9]+)", raw_du, re.IGNORECASE)
        term = m.group(1) if m else re.sub(r"\D+", "", raw_du)
        if term:
            qs = qs.filter(du__icontains=term)

    if filtros["id_claro"]:
        qs = qs.filter(id_claro__icontains=filtros["id_claro"])
    if filtros["id_new"]:
        qs = qs.filter(id_new__icontains=filtros["id_new"])
    if filtros["mes_produccion"]:
        qs = qs.filter(mes_produccion__icontains=filtros["mes_produccion"])
    if filtros["estado"]:
        if filtros["estado"] in estados_ok:
            qs = qs.filter(estado=filtros["estado"])
        else:
            qs = qs.none()
    if f_region:
        qs = qs.filter(region__icontains=f_region)

    # Armar filas export
    filas = []
    for s in qs:
        asignados = list(s.trabajadores_asignados.all())
        n = len(asignados) or 1

        # ⬅️ monto con signo según el estado (descuento negativo, bono/adelanto positivo)
        firmado = _monto_firmado_por_estado(s.monto_mmoo, s.estado)
        parte = (Decimal(str(firmado)) / n).quantize(Decimal('0.01'))

        if asignados:
            for tec in asignados:
                if f_tecnico:
                    target = f_tecnico.lower()
                    full_name = ((tec.first_name or "") + " " +
                                 (tec.last_name or "")).strip().lower()
                    username = (tec.username or "").lower()
                    if target not in full_name and target not in username:
                        continue

                filas.append((
                    f"DU{s.du}",
                    s.id_claro or "",
                    s.region or "",
                    s.mes_produccion or "",
                    s.id_new or "",
                    s.detalle_tarea or "",
                    (tec.get_full_name() or tec.username) if tec else "",
                    s.fecha_aprobacion_supervisor.strftime(
                        "%d/%m/%Y %H:%M") if s.fecha_aprobacion_supervisor else "",
                    s.estado,
                    float(parte),
                ))
        else:
            filas.append((
                f"DU{s.du}",
                s.id_claro or "",
                s.region or "",
                s.mes_produccion or "",
                s.id_new or "",
                s.detalle_tarea or "",
                "",
                s.fecha_aprobacion_supervisor.strftime(
                    "%d/%m/%Y %H:%M") if s.fecha_aprobacion_supervisor else "",
                s.estado,
                float(parte),
            ))

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Producción"
    headers = [
        "DU", "ID CLARO", "REGIÓN", "MES PRODUCCIÓN", "ID NEW", "DETALLE TAREA",
        "TÉCNICO", "FECHA FIN", "STATUS", "MONTO TÉCNICO (CLP)"
    ]
    ws.append(headers)
    for row in filas:
        ws.append(row)

    # Auto ancho
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)

    # Formato numérico última columna
    last_col = len(headers)
    for col_cells in ws.iter_cols(min_col=last_col, max_col=last_col, min_row=2, values_only=False):
        for c in col_cells:
            c.number_format = '#,##0.00'

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="produccion_aprobada.xlsx"'
    wb.save(resp)
    return resp


@login_required
def produccion_totales_a_pagar(request):
    """
    Totales a pagar por TÉCNICO (MENSUAL) a partir de ServiciosCotizados aprobados
    + ajustes (bono/adelanto/descuento).

    - Estados producción: aprobado_supervisor / aprobado_pm / aprobado_finanzas
    - Ajustes: ajuste_bono (suma), ajuste_adelanto (suma), ajuste_descuento (resta)
    - Prorrateo: monto_mmoo / N_tecnicos
    - Filtros: f_month (YYYY-MM o 'Mes Año' o 'Mes de Año'), f_tecnico, f_project, f_region
    - Desglose: lista de DUs/ajustes con su subtotal prorrateado.
    """
    current_month = _yyyy_mm(timezone.localdate())

    f_month = (request.GET.get("f_month") or "").strip()
    f_tecnico = (request.GET.get("f_tecnico") or "").strip()
    f_project = (request.GET.get("f_project") or "").strip()
    f_region = (request.GET.get("f_region") or "").strip()
    cantidad = (request.GET.get("cantidad") or "10").strip().lower()

    # ANTES: month_to_use = f_month or current_month  (esto escondía septiembre al pasar a octubre)
    # AHORA: el mes solo afecta el HISTORIAL (abajo). Arriba NO se filtra por mes.
    month_to_use = f_month  # <-- clave

    # ---- Estados incluidos ----
    EST_PROD = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}
    AJ_POS = {"ajuste_bono", "ajuste_adelanto"}        # SUMAN
    AJ_NEG = {"ajuste_descuento"}                      # RESTA
    estados_ok = EST_PROD | AJ_POS | AJ_NEG

    # ===== 1) BASE SIN MES (servirá para el bloque superior) =====
    qs_base = (
        ServicioCotizado.objects
        .filter(estado__in=estados_ok)
        .prefetch_related("trabajadores_asignados")
        .order_by("-fecha_aprobacion_supervisor", "-id")
    )

    # Filtros de proyecto / región (se aplican a ambos bloques)
    if f_project:
        qs_base = qs_base.filter(
            Q(id_claro__icontains=f_project) |
            Q(id_new__icontains=f_project) |
            Q(detalle_tarea__icontains=f_project)
        )
    if f_region:
        qs_base = qs_base.filter(region__icontains=f_region)

    # ===== 2) HISTORIAL (con mes si viene) =====
    # Creamos un clon que sí aplicará el filtro de mes
    qs_hist = qs_base

    # ---- Filtro MES: acepta 'YYYY-MM', 'Mes Año' y 'Mes de Año' ----
    MONTHS_ES = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
    ]

    def _aliases_for_month(token: str) -> list[str]:
        token = (token or "").strip()
        if not token:
            return []
        out = set([token])

        # YYYY-MM o YYYY/M
        m = re.search(r"^\s*(\d{4})[/-](\d{1,2})\s*$", token)
        if m:
            y = int(m.group(1))
            mm = int(m.group(2))
            mm2 = str(mm).zfill(2)
            name = MONTHS_ES[mm-1]
            out.update({
                f"{y}-{mm2}", f"{y}/{mm}",
                f"{name} {y}", f"{name.capitalize()} {y}",
                f"{name} de {y}", f"{name.capitalize()} de {y}",
                f"{mm}-{y}", f"{mm2}-{y}", f"{mm}/{y}",
            })
            return list(out)

        # "mes (de) año"
        m = re.search(
            r"^\s*([a-záéíóúü]+)\s+(?:de\s+)?(\d{4})\s*$", token.lower())
        if m:
            name = m.group(1)
            y = int(m.group(2))
            idx = next((i for i, n in enumerate(MONTHS_ES)
                       if name.startswith(n[:3])), None)
            if idx is not None:
                mm = idx+1
                mm2 = str(mm).zfill(2)
                full = MONTHS_ES[idx]
                out.update({
                    f"{y}-{mm2}", f"{y}/{mm}",
                    f"{full} {y}", f"{full.capitalize()} {y}",
                    f"{full} de {y}", f"{full.capitalize()} de {y}",
                    f"{mm}-{y}", f"{mm2}-{y}", f"{mm}/{y}",
                })
            return list(out)

        return list(out)

    # SOLO al historial le aplicamos f_month
    if month_to_use:
        try:
            aliases = _month_aliases(month_to_use)  # si ya existe
        except NameError:
            aliases = _aliases_for_month(month_to_use)

        cond = Q()
        for a in aliases:
            cond |= Q(mes_produccion__icontains=a)
        qs_hist = qs_hist.filter(cond)

    # ===== 3) Agregador común (tu misma lógica, solo envuelta) =====
    agregados_tipo = dict[int, dict]  # hint

    def _agrega_por_tecnico(qs_src) -> list[dict]:
        agregados: agregados_tipo = {}
        for s in qs_src:
            asignados = list(s.trabajadores_asignados.all())
            if not asignados:
                continue

            n = len(asignados)
            base = Decimal(str(s.monto_mmoo or 0))

            # Normalizar signo según estado
            if s.estado in AJ_NEG:
                firmado = -abs(base)        # descuento/adelanto restan
            elif s.estado in AJ_POS:
                firmado = abs(base)         # bono suma
            else:
                firmado = base              # producción

            parte = (firmado / n).quantize(Decimal("0.01"))

            # Etiqueta DU
            du_label, project_id = _fmt_du_and_pid(s)

            for tec in asignados:
                # Filtro por técnico (nombre/username)
                if f_tecnico:
                    target = f_tecnico.lower()
                    full_name = ((tec.first_name or "") + " " +
                                 (tec.last_name or "")).strip().lower()
                    username = (tec.username or "").lower()
                    if target not in full_name and target not in username:
                        continue

                display_name = tec.get_full_name() or tec.username
                serv_month = (s.mes_produccion or "").strip()

                bucket = agregados.setdefault(tec.id, {
                    "tecnico": tec,
                    "display_name": display_name,
                    "week": serv_month or (month_to_use or ""),
                    "amount": Decimal("0.00"),
                    "details": [],
                    "status": "",
                    "receipt": None,
                })

                # Si mezcla meses en el mismo bucket → "(varios meses)"
                if bucket["week"] and serv_month and bucket["week"] != serv_month:
                    bucket["week"] = "(varios meses)"

                bucket["amount"] += parte
                bucket["details"].append({
                    "project_id": s.id_claro or s.id_new or du_label,
                    "du": du_label,
                    "subtotal": parte,
                })

        rows = list(agregados.values())
        rows.sort(key=lambda r: (
            (r["tecnico"].last_name or "").lower(),
            (r["tecnico"].first_name or "").lower(),
            (r["tecnico"].username or "").lower(),
        ))
        return rows

    # ===== 4) Construir datasets separados =====
    # Arriba (pendiente por pagar): SIN mes
    rows_top = _agrega_por_tecnico(qs_base)

    # Abajo (historial/paginación): CON mes si se solicitó
    rows_hist = _agrega_por_tecnico(qs_hist)

    # Paginación del bloque inferior
    if cantidad == "todos":
        pagina = rows_hist
    else:
        try:
            per_page = max(5, min(100, int(cantidad)))
        except ValueError:
            per_page = 10
            cantidad = "10"
        paginator = Paginator(rows_hist, per_page)
        page_number = request.GET.get("page") or 1
        pagina = paginator.get_page(page_number)

    filters_qs = urlencode({
        k: v for k, v in {
            "f_month": month_to_use,      # solo historial
            "f_tecnico": f_tecnico,
            "f_project": f_project,
            "f_region":  f_region,
            "cantidad":  cantidad,
        }.items() if v
    })

    return render(request, "operaciones/produccion_totales_admin.html", {
        "current_month": current_month,
        "selected_month": month_to_use or "",
        "top": rows_top,          # ← independiente del mes y de ?f_month=...
        "pagina": pagina,         # ← sí respeta el filtro de mes
        "cantidad": cantidad,
        "f_month": month_to_use or "",
        "f_tecnico": f_tecnico,
        "f_project": f_project,
        "f_region": f_region,
        "filters_qs": filters_qs,
    })


@login_required
def exportar_totales_produccion(request):
    """
    Exporta a Excel los totales a pagar (mensual) por técnico con detalle por DU/ajuste.
    Respeta filtros de mes (acepta YYYY-MM y alias), técnico, proyecto y región.
    """
    current_month = _yyyy_mm(timezone.localdate())

    f_month = (request.GET.get("f_month") or "").strip() or current_month
    f_tecnico = (request.GET.get("f_tecnico") or "").strip()
    f_project = (request.GET.get("f_project") or "").strip()
    f_region = (request.GET.get("f_region") or "").strip()

    # Query base con ajustes
    qs = (
        ServicioCotizado.objects
        .filter(estado__in=ESTADOS_PROD_Y_AJUSTES)
        .prefetch_related("trabajadores_asignados")
        .order_by("-fecha_aprobacion_supervisor", "-id")
    )

    # Filtro por mes con alias
    if f_month:
        cond_m = Q()
        for a in _month_aliases(f_month):
            cond_m |= Q(mes_produccion__icontains=a)
        qs = qs.filter(cond_m)

    if f_project:
        qs = qs.filter(
            Q(id_claro__icontains=f_project) |
            Q(id_new__icontains=f_project) |
            Q(detalle_tarea__icontains=f_project)
        )
    if f_region:
        qs = qs.filter(region__icontains=f_region)

    agregados: dict[int, dict] = {}
    for s in qs:
        asignados = list(s.trabajadores_asignados.all())
        if not asignados:
            continue
        n = len(asignados)

        firmado = _monto_firmado_por_estado(s.monto_mmoo, s.estado)
        parte = (Decimal(str(firmado)) / n).quantize(Decimal("0.01"))

        du_label, project_id = _fmt_du_and_pid(s)

        for tec in asignados:
            if f_tecnico:
                target = f_tecnico.lower()
                full_name = ((tec.first_name or "") + " " +
                             (tec.last_name or "")).strip().lower()
                username = (tec.username or "").lower()
                if target not in full_name and target not in username:
                    continue
            b = agregados.setdefault(tec.id, {
                "tecnico": tec,
                "week": f_month,
                "amount": Decimal("0.00"),
                "details": [],
            })
            b["amount"] += parte
            b["details"].append({
                "project_id": project_id,
                "du": du_label,
                "subtotal": parte,
            })

    rows = list(agregados.values())
    rows.sort(key=lambda r: (
        (r["tecnico"].last_name or "").lower(),
        (r["tecnico"].first_name or "").lower(),
        (r["tecnico"].username or "").lower(),
    ))

    # ---- Excel ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Totales a pagar"

    ws.append(["Técnico", "Mes", "Total (CLP)"])
    for r in rows:
        tech_name = r["tecnico"].get_full_name() or r["tecnico"].username
        ws.append([tech_name, r["week"], float(r["amount"])])

    # Hoja de detalle
    ws2 = wb.create_sheet("Detalle por DU")
    ws2.append(["Técnico", "Mes", "Proyecto (ID)", "DU", "Subtotal (CLP)"])
    for r in rows:
        tech_name = r["tecnico"].get_full_name() or r["tecnico"].username
        for d in r["details"]:
            ws2.append([tech_name, r["week"], d["project_id"],
                       d["du"], float(d["subtotal"])])

    # Anchos auto
    for sheet in (ws, ws2):
        for col in sheet.columns:
            max_len = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            sheet.column_dimensions[letter].width = min(
                max(12, max_len + 2), 60)

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="totales_a_pagar_mensual.xlsx"'
    wb.save(resp)
    return resp


ESTADOS_OK = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}


def _s3_client():
    return boto3.client(
        "s3",
        endpoint_url=getattr(settings, "AWS_S3_ENDPOINT_URL",
                             "https://s3.us-east-1.wasabisys.com"),
        region_name=getattr(settings, "AWS_S3_REGION_NAME", "us-east-1"),
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        config=Config(signature_version="s3v4", s3={
                      "addressing_style": "path"}),
        verify=getattr(settings, "AWS_S3_VERIFY", True),
    )


def _format_du(du) -> str:
    """Devuelve DU normalizado:
       - si ya viene con prefijo 'DU' -> lo deja igual
       - si viene solo número/texto  -> le agrega 'DU'
       - si viene vacío               -> '—'
    """
    s = str(du or '').strip()
    if not s:
        return '—'
    return s if s.upper().startswith('DU') else f'DU{s}'


def _fmt_du_and_pid(s) -> tuple[str, str]:
    """
    Devuelve (du_label, project_id) sin duplicar DU como proyecto y,
    si el registro es un AJUSTE y no tiene proyecto, usa una etiqueta
    legible: 'Descuento' / 'Bono' / 'Adelanto'.
    """
    du_label = _format_du(getattr(s, 'du', None))  # 'DU000001' o '—'

    # Proyecto real si existe
    project = (getattr(s, 'id_claro', '') or getattr(
        s, 'id_new', '') or '').strip()

    # Si no hay proyecto, y es ajuste, usamos el nombre del ajuste
    if not project or project in {'-', '—'}:
        estado = getattr(s, 'estado', '')
        ajuste_map = {
            'ajuste_descuento': 'Descuento',
            'ajuste_bono': 'Bono',
            'ajuste_adelanto': 'Adelanto',
        }
        project = ajuste_map.get(estado, '—')

    # Si hay DU, devolvemos DU + proyecto (si lo hay)
    if du_label != '—':
        return du_label, project

    # Si NO hay DU, usamos el proyecto como fallback en ambos
    return project, project


def _sync_monthly_totals(month: str, create_missing: bool = True) -> dict:
    aliases = _month_aliases(month)
    cond = Q()
    for a in aliases:
        cond |= Q(mes_produccion__icontains=a)

    totals: dict[int, Decimal] = {}

    servicios = (
        ServicioCotizado.objects
        .filter(estado__in=ESTADOS_PROD_Y_AJUSTES)   # ← incluir ajustes
        .filter(cond)
        .prefetch_related("trabajadores_asignados")
        .only("id", "du", "id_claro", "id_new", "monto_mmoo", "estado")
    )

    for s in servicios:
        asignados = list(s.trabajadores_asignados.all())
        if not asignados:
            continue
        n = len(asignados)
        firmado = _monto_firmado_por_estado(s.monto_mmoo, s.estado)
        parte = (Decimal(str(firmado)) / n).quantize(Decimal("0.01"))
        for tec in asignados:
            totals[tec.id] = totals.get(tec.id, Decimal("0.00")) + parte

    existing = {(mp.technician_id, mp.month): mp
                for mp in MonthlyPayment.objects.filter(month=month)}
    updated = created = deleted = 0

    for tech_id, amount in totals.items():
        key = (tech_id, month)
        mp = existing.pop(key, None)
        if mp:
            if mp.amount != amount:
                mp.amount = amount
                save_fields = ["amount", "updated_at"]
                if mp.status == "approved_user":
                    mp.status = "pending_payment"
                    save_fields.append("status")
                mp.save(update_fields=save_fields)
                updated += 1
        else:
            if create_missing and amount != 0:
                MonthlyPayment.objects.create(
                    technician_id=tech_id,
                    month=month,
                    amount=amount,
                    status="pending_user",
                )
                created += 1

    # NO borrar rechazados; mantenerlos para poder restablecerlos.
    for _key, mp in existing.items():
        if mp.status == "paid":
            # nunca tocar pagados
            continue

        if mp.status == "rejected_user":
            # conservar el registro rechazado, pero sincronizar monto (por si cambió)
            new_amount = totals.get(mp.technician_id, Decimal("0.00"))
            if mp.amount != new_amount:
                mp.amount = new_amount
                mp.save(update_fields=["amount", "updated_at"])
                updated += 1
            continue

        # para los demás estados no pagados (pending_user/approved_user/pending_payment):
        # si ya no hay total para ese técnico-mes, se elimina
        mp.delete()
        deleted += 1

    return {"updated": updated, "created": created, "deleted": deleted}


def _prev_month(yyyy_mm: str) -> str:
    y = int(yyyy_mm[:4])
    m = int(yyyy_mm[5:7])
    d = datetime(y, m, 1) - relativedelta(months=1)
    return f"{d.year:04d}-{d.month:02d}"


@login_required
def admin_monthly_payments(request):
    current_month = _yyyy_mm(timezone.localdate())
    today = timezone.localdate()
    day = today.day  # (ya no se usa para decidir qué mostrar)

    # ---------- Filtros del formulario ----------
    applied = (request.GET.get("filters") == "1")
    f_month_raw = (request.GET.get("f_month") or "").strip()
    f_tecnico = (request.GET.get("f_tecnico") or "").strip()
    f_project = (request.GET.get("f_project") or "").strip()
    f_region = (request.GET.get("f_region") or "").strip()
    f_du = (request.GET.get("f_du") or "").strip()
    f_paid_month = (request.GET.get("f_paid_month") or "").strip()
    cantidad = (request.GET.get("cantidad") or "10").strip().lower()

    # ---------- TOP: lógica de pendientes ----------
    prev_month = _prev_month(current_month)

    # Siempre sincroniza mes anterior Y mes actual
    _sync_monthly_totals(prev_month, create_missing=True)
    _sync_monthly_totals(current_month, create_missing=True)

    # Base: todos los no pagados con monto > 0
    top_qs = (
        MonthlyPayment.objects
        .filter(amount__gt=0)
        .exclude(status="paid")
        .select_related("technician")
    )

    # Mostrar SIEMPRE todos los meses pendientes hasta el mes actual (incluido)
    top_qs = top_qs.filter(month__lte=current_month)

    # Filtro por técnico en el bloque superior
    if f_tecnico:
        top_qs = top_qs.filter(
            Q(technician__first_name__icontains=f_tecnico) |
            Q(technician__last_name__icontains=f_tecnico) |
            Q(technician__username__icontains=f_tecnico)
        )

    top_qs = top_qs.order_by(
        "status", "month", "technician__first_name", "technician__last_name"
    )

    # Helper para armar el detalle por DU para un técnico/mes
    def _details_for(tech_id: int, month: str):
        aliases = _month_aliases(month)
        cond = Q()
        for a in aliases:
            cond |= Q(mes_produccion__icontains=a)
        detalles = []
        servicios = (
            ServicioCotizado.objects
            .filter(estado__in=ESTADOS_PROD_Y_AJUSTES)
            .filter(cond)
            .prefetch_related("trabajadores_asignados")
            .only("du", "id_claro", "id_new", "monto_mmoo", "estado")
        )
        for s in servicios:
            tecs = list(s.trabajadores_asignados.all())
            if not tecs or not any(t.id == tech_id for t in tecs):
                continue
            firmado = _monto_firmado_por_estado(s.monto_mmoo, s.estado)
            parte = (Decimal(str(firmado)) / len(tecs)
                     ).quantize(Decimal("0.01"))
            du_label, project_id = _fmt_du_and_pid(s)
            detalles.append(
                {"du": du_label, "project_id": project_id, "subtotal": parte})
        return detalles

    top = list(top_qs)
    for r in top:
        r.display_name = (r.technician.get_full_name()
                          or r.technician.username)
        r.week = r.month
        r.details = _details_for(r.technician_id, r.month)

    # ---------- HISTORIAL (pagados) con filtros opcionales ----------
    bottom_qs = (
        MonthlyPayment.objects
        .filter(status="paid")
        .select_related("technician")
        .order_by("-paid_month", "-month", "technician__first_name", "technician__last_name")
    )
    if applied and f_month_raw:
        bottom_qs = bottom_qs.filter(month=f_month_raw)
    if f_paid_month:
        bottom_qs = bottom_qs.filter(paid_month__icontains=f_paid_month)
    if f_tecnico:
        bottom_qs = bottom_qs.filter(
            Q(technician__first_name__icontains=f_tecnico) |
            Q(technician__last_name__icontains=f_tecnico) |
            Q(technician__username__icontains=f_tecnico)
        )

    # Filtro DU en memoria
    def _match_du(detalles, needle: str) -> bool:
        if not needle:
            return True
        n = needle.strip().lower()
        for d in detalles:
            du_text = (d.get("du") or "").lower()
            pid = (d.get("project_id") or "").lower()
            if n in du_text or n in pid:
                return True
            if du_text.startswith("du") and du_text[2:] == n:
                return True
        return False

    if f_du:
        tmp = []
        for mp in bottom_qs:
            dets = _details_for(mp.technician_id, mp.month)
            if _match_du(dets, f_du):
                mp.details = dets
                tmp.append(mp)
        bottom_list = tmp
    else:
        bottom_list = list(bottom_qs)

    # Paginación
    if cantidad == "todos":
        pagina = bottom_list
    else:
        try:
            per_page = max(5, min(100, int(cantidad)))
        except ValueError:
            per_page = 10
            cantidad = "10"
        paginator = Paginator(bottom_list, per_page)
        page_number = request.GET.get("page") or 1
        pagina = paginator.get_page(page_number)

    # Completar visibles
    page_items = list(pagina) if not isinstance(pagina, list) else pagina
    for r in page_items:
        r.display_name = (r.technician.get_full_name()
                          or r.technician.username)
        r.week = r.month
        if not hasattr(r, "details"):
            r.details = _details_for(r.technician_id, r.month)

    # El input “Mes de Pago” aparece vacío si no hay filtros aplicados
    f_month_input = f_month_raw if (applied and f_month_raw) else ""

    filters_qs = urlencode({
        k: v for k, v in {
            "filters": "1" if applied else "",
            "f_month": f_month_input,
            "f_tecnico": f_tecnico,
            "f_project": f_project,
            "f_region": f_region,
            "f_du": f_du,
            "f_paid_month": f_paid_month,
            "cantidad": cantidad,
        }.items() if v
    })

    # selected_month: muestra el mes actual en la cabecera
    return render(request, "operaciones/produccion_totales_admin.html", {
        "current_month": current_month,
        "selected_month": current_month,
        "top": top,
        "pagina": pagina,
        "cantidad": cantidad,
        "f_month": f_month_input,
        "f_tecnico": f_tecnico,
        "f_project": f_project,
        "f_region": f_region,
        "f_du": f_du,
        "f_paid_month": f_paid_month,
        "filters_qs": filters_qs,
    })


log = logging.getLogger('presign_receipt_monthly')


@require_POST
@login_required
def presign_receipt_monthly(request, pk: int):
    pay_id = pk
    log.debug('HIT presign_receipt_monthly pay_id=%s user=%s ajax=%s',
              pay_id, getattr(request.user, 'username', None),
              request.headers.get('X-Requested-With'))
    obj = get_object_or_404(MonthlyPayment, pk=pay_id)
    filename = request.POST.get(
        "filename") or f"receipt-{uuid.uuid4().hex}.pdf"
    base, ext = os.path.splitext(filename)
    ext = (ext or ".pdf").lower()
    key = f"operaciones/pagos/{obj.month}/{obj.technician_id}/receipt_{uuid.uuid4().hex}{ext}"

    s3 = _s3_client()
    fields = {"acl": "private", "success_action_status": "201"}
    conditions = [
        {"acl": "private"},
        {"success_action_status": "201"},
        ["content-length-range", 0, 25 * 1024 * 1024],
    ]
    post = s3.generate_presigned_post(
        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
        Key=key,
        Fields=fields,
        Conditions=conditions,
        ExpiresIn=600,
    )
    endpoint = settings.AWS_S3_ENDPOINT_URL.rstrip("/")
    bucket = settings.AWS_STORAGE_BUCKET_NAME
    post["url"] = f"{endpoint}/{bucket}"
    log.debug('presign OK url=%s', post["url"])
    return JsonResponse({"post": post, "key": key})


log = logging.getLogger('confirm_receipt_monthly')


@require_POST
@login_required
def confirm_receipt_monthly(request, pay_id: int = None, pk: int = None):
    # Acepta ambos nombres desde la URL
    pay_id = pay_id or pk
    log.debug('HIT confirm_receipt_monthly id=%s user=%s',
              pay_id, getattr(request.user, 'username', None))
    try:
        obj = get_object_or_404(MonthlyPayment, pk=pay_id)

        key = (request.POST.get("key") or "").strip()
        log.debug('confirm key=%s', key)
        if not key:
            return JsonResponse({"ok": False, "error": "missing key"}, status=400)

        obj.receipt.name = key
        obj.status = "paid"
        obj.paid_month = _yyyy_mm(timezone.localdate())
        obj.save(update_fields=["receipt", "status",
                 "paid_month", "updated_at"])

        log.debug('confirm OK receipt=%s status=paid paid_month=%s',
                  obj.receipt.name, obj.paid_month)
        return JsonResponse({"ok": True})

    except Exception as e:
        log.exception('confirm ERROR id=%s', pay_id)
        return JsonResponse({"ok": False, "error": str(e)}, status=500)


@require_POST
@login_required
@transaction.atomic
def admin_unpay_monthly(request, pk: int):
    mp = get_object_or_404(MonthlyPayment, pk=pk)

    # 🚩 Señales de que esperan JSON
    xrw = (request.headers.get('X-Requested-With') or
           request.META.get('HTTP_X_REQUESTED_WITH') or '')
    accept = (request.headers.get('Accept') or '')
    is_ajax = (
        # 👈 parámetro explícito
        request.GET.get('ajax') == '1'
        or xrw.lower() == 'xmlhttprequest'                   # case-insensitive
        or 'application/json' in accept.lower()
        or 'json' in accept.lower()
    )

    if mp.status != "paid":
        if is_ajax:
            return JsonResponse({"ok": False, "error": "only_paid_can_be_reverted"}, status=400)
        messages.info(
            request, "Sólo se puede revertir pagos en estado 'Pagado'.")
        return redirect("operaciones:admin_monthly_payments")

    # borrar recibo si existe (ignorar errores silenciosamente)
    try:
        if mp.receipt:
            mp.receipt.delete(save=False)
    except Exception:
        pass

    mp.receipt = None
    mp.paid_month = ""
    mp.status = "pending_payment"
    mp.save(update_fields=["receipt", "paid_month", "status", "updated_at"])

    if is_ajax:
        return JsonResponse({"ok": True})

    messages.success(request, "Pago revertido. Vuelve a 'Pendiente de pago'.")
    return redirect("operaciones:admin_monthly_payments")


@login_required
def user_monthly_payments(request):
    """
    Vista del técnico:
    - Sincroniza TODOS los meses (hasta el mes actual) donde el técnico tenga producción/ajustes.
    - Lista sus MonthlyPayment con desglose (DU/AJUSTE) y TOTAL.
    - Botones Aprobar/Rechazar: mes en curso desde el 26; meses anteriores, siempre.
    """
    today = timezone.localdate()
    current_month = _yyyy_mm(today)

    # === 1) Descubrir meses a sincronizar para ESTE técnico ===
    # Incluye producción aprobada + ajustes (bono/adelanto/descuento)
    qs_serv = (
        ServicioCotizado.objects
        .filter(estado__in=ESTADOS_PROD_Y_AJUSTES, trabajadores_asignados=request.user)
        .only("mes_produccion")
    )

    meses_detectados: set[str] = set()
    for s in qs_serv:
        txt = (s.mes_produccion or "").strip()
        if not txt:
            continue
        # Intenta normalizar a "YYYY-MM" (si falla, usa alias tal cual)
        norm = _yyyymm_from_mes_texto(txt) if ' ' in txt else txt
        if isinstance(norm, str) and len(norm) >= 7 and norm[4] == "-":
            meses_detectados.add(norm[:7])
        else:
            # fallback: si viene "2025-9" o similar, intenta normalizar
            m = re.match(r"^\s*(\d{4})[/-](\d{1,2})\s*$", txt)
            if m:
                y, mm = int(m.group(1)), int(m.group(2))
                meses_detectados.add(f"{y:04d}-{mm:02d}")

    # Garantiza al menos mes actual
    meses_detectados.add(current_month)

    # Mantén sólo meses <= actual
    meses_a_sync = sorted([m for m in meses_detectados if m <= current_month])

    # === 2) Sincronizar cada mes detectado para este técnico ===
    # Nota: _sync_monthly_totals crea/actualiza para todos los técnicos, lo cual está OK.
    for m in meses_a_sync:
        _sync_monthly_totals(m, create_missing=True)

    # === 3) Cargar mis MonthlyPayment (ahora sí deberían incluir septiembre, etc.) ===
    mine = (
        MonthlyPayment.objects
        .filter(technician=request.user)
        .order_by("-month")
    )

    # === 4) Preparar desglose y totales por mes sólo para ESTE usuario ===
    by_month = {mp.month for mp in mine}
    details: dict[str, list] = {}
    totals: dict[str, Decimal] = {}

    for m in by_month:
        aliases = _month_aliases(m)
        cond = Q()
        for a in aliases:
            cond |= Q(mes_produccion__icontains=a)

        servicios = (
            ServicioCotizado.objects
            .filter(estado__in=ESTADOS_PROD_Y_AJUSTES)
            .filter(cond)
            .prefetch_related("trabajadores_asignados")
            .only("du", "id_claro", "id_new", "monto_mmoo", "estado")
        )

        tot_m = Decimal("0.00")
        det_m = []

        for s in servicios:
            tecs = list(s.trabajadores_asignados.all())
            if not tecs or not any(t.id == request.user.id for t in tecs):
                continue

            firmado = _monto_firmado_por_estado(s.monto_mmoo, s.estado)
            parte = (Decimal(str(firmado)) / len(tecs)
                     ).quantize(Decimal("0.01"))
            du_label, project_id = _fmt_du_and_pid(s)

            det_m.append(
                {"du": du_label, "project_id": project_id, "subtotal": parte})
            tot_m += parte

        if det_m:
            details[m] = det_m
        totals[m] = tot_m

    # Inyecta 'details' y un 'display_amount' robusto
    for mp in mine:
        mp.details = details.get(mp.month, [])
        mp.display_amount = (
            Decimal(str(mp.amount)).quantize(Decimal("0.01"))
            if getattr(mp, "amount", None) is not None
            else totals.get(mp.month, Decimal("0.00")).quantize(Decimal("0.01"))
        )

    can_approve_today = today.day >= 26
    return render(request, "operaciones/pagos_usuario_mensual.html", {
        "current_month": current_month,
        "mine": mine,
        "can_approve_today": can_approve_today,
    })


# ---- Helper para detectar (año, mes) del MonthlyPayment ----

def _mp_period_ym(mp) -> tuple[int, int]:
    """
    Devuelve (year, month) del período del MonthlyPayment 'mp' intentando varios esquemas:
    - mp.year / mp.month        -> ints
    - mp.period (date/datetime) -> toma .year/.month
    - mp.month (str "YYYY-MM")  -> parsea
    Si no se puede, devuelve el mes actual (fallback seguro).
    """
    # 1) year / month int
    if hasattr(mp, "year") and hasattr(mp, "month"):
        try:
            y = int(getattr(mp, "year"))
            m = int(getattr(mp, "month"))
            if 1 <= m <= 12:
                return y, m
        except Exception:
            pass

    # 2) period date/datetime
    if hasattr(mp, "period"):
        per = getattr(mp, "period")
        try:
            if per is not None:
                return per.year, per.month
        except Exception:
            pass

    # 3) month string "YYYY-MM"
    if hasattr(mp, "month"):
        mm = getattr(mp, "month")
        if isinstance(mm, str) and len(mm) >= 7 and mm[4] == "-":
            try:
                y = int(mm[:4])
                m = int(mm[5:7])
                if 1 <= m <= 12:
                    return y, m
            except Exception:
                pass

    # Fallback: hoy
    today = timezone.localdate()
    return today.year, today.month


def _is_current_period(mp) -> bool:
    y, m = _mp_period_ym(mp)
    today = timezone.localdate()
    return (y == today.year) and (m == today.month)


def _is_past_period(mp) -> bool:
    y, m = _mp_period_ym(mp)
    today = timezone.localdate()
    # Comparación por (año, mes)
    return (y, m) < (today.year, today.month)


@login_required
@require_POST
@transaction.atomic
def user_approve_monthly(request, pk: int):
    """
    El técnico aprueba su monto mensual.
    - Mes en curso: permitido desde el día 26 inclusive.
    - Meses anteriores: siempre permitido (mientras esté pending_user).
    """
    mp = get_object_or_404(MonthlyPayment, pk=pk, technician=request.user)

    if mp.status != "pending_user":
        messages.info(
            request, "Sólo puedes aprobar cuando está 'Pendiente de mi aprobación'.")
        return redirect("operaciones:user_monthly_payments")

    today = timezone.localdate()
    if _is_current_period(mp):
        if today.day < 26:
            messages.info(
                request, "Sólo puedes aprobar/rechazar desde el día 26 de cada mes (mes en curso).")
            return redirect("operaciones:user_monthly_payments")
    elif not _is_past_period(mp):
        # Caso futuro (no debería ocurrir normalmente)
        messages.info(
            request, "Este período aún no está habilitado para aprobación.")
        return redirect("operaciones:user_monthly_payments")

    mp.reject_reason = ""
    mp.status = "pending_payment"
    mp.save(update_fields=["status", "reject_reason", "updated_at"])
    messages.success(request, "Monto aprobado. Esperando pago.")
    return redirect("operaciones:user_monthly_payments")


@login_required
@require_POST
@transaction.atomic
def user_reject_monthly(request, pk: int):
    """
    El técnico rechaza su monto mensual (debe indicar motivo).
    - Mes en curso: permitido desde el día 26 inclusive.
    - Meses anteriores: siempre permitido (mientras esté pending_user).
    """
    mp = get_object_or_404(MonthlyPayment, pk=pk, technician=request.user)
    reason = (request.POST.get("reason") or "").strip()

    if mp.status != "pending_user":
        messages.info(
            request, "Sólo puedes rechazar cuando está 'Pendiente de mi aprobación'.")
        return redirect("operaciones:user_monthly_payments")
    if not reason:
        messages.error(request, "Debes indicar un motivo.")
        return redirect("operaciones:user_monthly_payments")

    today = timezone.localdate()
    if _is_current_period(mp):
        if today.day < 26:
            messages.info(
                request, "Sólo puedes aprobar/rechazar desde el día 26 de cada mes (mes en curso).")
            return redirect("operaciones:user_monthly_payments")
    elif not _is_past_period(mp):
        # Caso futuro (no debería ocurrir normalmente)
        messages.info(
            request, "Este período aún no está habilitado para rechazo.")
        return redirect("operaciones:user_monthly_payments")

    mp.reject_reason = reason
    mp.status = "rejected_user"
    mp.save(update_fields=["status", "reject_reason", "updated_at"])
    messages.success(request, "Monto rechazado. Motivo guardado.")
    return redirect("operaciones:user_monthly_payments")


APROBADOS_OK = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}


# Asegúrate de tener estos importados en tu archivo real:
# from operaciones.models import ServicioCotizado
# from operaciones.const import APROBADOS_OK


def _yyyy_mm(dt) -> str:
    d = dt.date() if hasattr(dt, "date") else dt
    return f"{d.year:04d}-{d.month:02d}"


def _month_aliases(token: str) -> list[str]:
    s = (token or "").strip()
    if not s:
        return []
    m = re.match(r"^(\d{4})-(\d{1,2})$", s)
    if not m:
        return [s]
    y = int(m.group(1))
    mm = int(m.group(2))
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    if 1 <= mm <= 12:
        return [s, f"{meses[mm-1]} {y}"]
    return [s]


@login_required
def produccion_tecnico(request):
    current_month = _yyyy_mm(timezone.localdate())

    q_text = (request.GET.get("id_claro") or "").strip()
    q_month = (request.GET.get("mes_produccion") or "").strip()

    # 🔹 ahora usamos raw_cantidad y luego la normalizamos / limitamos a 100
    raw_cantidad = (request.GET.get("cantidad") or "10").strip().lower()

    # 👉 incluir ajustes para que se muestren al usuario
    AJUSTES = ["ajuste_bono", "ajuste_adelanto", "ajuste_descuento"]

    qs = (
        ServicioCotizado.objects
        .filter(
            Q(estado__in=APROBADOS_OK) | Q(estado__in=AJUSTES),
            trabajadores_asignados=request.user
        )
        .order_by("-fecha_aprobacion_supervisor", "-id")
        .prefetch_related("trabajadores_asignados")
    )

    # ====== Encabezado (mes actual) ======
    _y = int(current_month[:4])
    _m = int(current_month[5:7])
    _meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
              "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    current_month_label = f"{_meses[_m-1]} {_y}"

    total_current_month = Decimal("0.00")
    _cond_mes = Q()
    for alias in _month_aliases(current_month):
        _cond_mes |= Q(mes_produccion__icontains=alias)
    for s in qs.filter(_cond_mes):
        asignados_cm = list(s.trabajadores_asignados.all())
        if not asignados_cm:
            continue
        parte_cm = (Decimal(str(s.monto_mmoo or 0)) /
                    len(asignados_cm)).quantize(Decimal("0.01"))
        total_current_month += parte_cm
    # =====================================

    if q_month:
        cond = Q()
        for alias in _month_aliases(q_month):
            cond |= Q(mes_produccion__icontains=alias)
        qs = qs.filter(cond)

    if q_text:
        qs = qs.filter(
            Q(id_claro__icontains=q_text) |
            Q(id_new__icontains=q_text) |
            Q(detalle_tarea__icontains=q_text)
        )

    # === Construcción de filas (primero mes actual, luego septiembre, agosto, etc.) ===
    filas, total_estimado = [], Decimal("0.00")

    qs_mes_actual = qs.filter(_cond_mes)
    qs_otros = qs.exclude(_cond_mes)

    # Parser robusto para "mes_produccion": "YYYY-MM" o "Octubre 2025"
    def _ym_from_mes(txt: str):
        import re
        if not txt:
            return (0, 0)
        t = str(txt).strip()
        # 1) Formato "YYYY-MM"
        m = re.match(r"^\s*(\d{4})-(\d{1,2})\s*$", t)
        if m:
            y, mm = int(m.group(1)), int(m.group(2))
            return (y, mm)
        # 2) Formato "Mes YYYY" en español
        mes_map = {
            "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
            "julio": 7, "agosto": 8, "septiembre": 9, "setiembre": 9,
            "octubre": 10, "noviembre": 11, "diciembre": 12
        }
        m2 = re.match(r"^\s*([A-Za-zÁÉÍÓÚáéíóúñÑ]+)\s+(\d{4})\s*$", t)
        if m2:
            mes_txt = m2.group(1).lower()
            y = int(m2.group(2))
            mm = mes_map.get(mes_txt, 0)
            return (y, mm)
        return (0, 0)

    def _push(s):
        nonlocal total_estimado
        asignados = list(s.trabajadores_asignados.all())
        if not asignados:
            return
        parte = (Decimal(str(s.monto_mmoo or 0)) /
                 len(asignados)).quantize(Decimal("0.01"))
        filas.append({
            "id": s.id,
            "du": s.du,
            "id_claro": s.id_claro or "",
            "id_new": s.id_new or "",
            "region": s.region or "",
            "mes_produccion": s.mes_produccion or "",
            "detalle_tarea": s.detalle_tarea or "",
            "estado": s.estado or "",
            "monto_tecnico": parte,
        })
        total_estimado += parte

    # 1) Mes actual (respetando el order_by original del queryset)
    for s in qs_mes_actual:
        _push(s)

    # 2) Resto de meses, ordenados de más reciente a más antiguo por mes_produccion
    otros_ordenados = sorted(
        list(qs_otros),
        key=lambda s: _ym_from_mes(s.mes_produccion),
        reverse=True
    )
    for s in otros_ordenados:
        _push(s)
    # ====================================================================

    # 🔹 Lógica de paginación: máx 100, igual que Mis Rendiciones
    if raw_cantidad == "todos":
        per_page = 100
        cantidad = "100"
    else:
        try:
            per_page = int(raw_cantidad)
        except (TypeError, ValueError):
            per_page = 10
            cantidad = "10"
        else:
            if per_page < 1:
                per_page = 10
                cantidad = "10"
            elif per_page > 100:
                per_page = 100
                cantidad = "100"
            else:
                cantidad = raw_cantidad

    paginator = Paginator(filas, per_page)
    pagina = paginator.get_page(request.GET.get("page") or 1)

    return render(request, "operaciones/produccion_tecnico.html", {
        "current_month": current_month,
        "current_month_label": current_month_label,
        "total_current_month": total_current_month,
        "pagina": pagina,
        "cantidad": cantidad,
        "id_claro": q_text,
        "mes_produccion": q_month,
        "total_estimado": total_estimado,
    })


logger = logging.getLogger(__name__)


@login_required
@rol_requerido('usuario')
def exportar_produccion_pdf(request):
    try:
        usuario = request.user
        id_new = request.GET.get("id_new", "")
        mes_produccion = request.GET.get("mes_produccion", "")
        filtro_pdf = request.GET.get("filtro_pdf", "mes_actual")

        # Traducción manual de meses
        meses_es = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        now_local = timezone.localtime()
        mes_actual = f"{meses_es[now.month]} {now.year}"

        # Texto de filtro
        if filtro_pdf == "mes_actual":
            filtro_seleccionado = f"Solo mes actual: {mes_actual}"
        elif filtro_pdf == "filtro_actual":
            filtro_seleccionado = f"Con filtros aplicados: {mes_produccion}" if mes_produccion else "Con filtros aplicados"
        else:
            filtro_seleccionado = "Toda la producción"

        # Query base
        servicios = ServicioCotizado.objects.filter(
            trabajadores_asignados=usuario,
            estado='aprobado_supervisor'
        )

        # Filtro según selección
        if filtro_pdf == "filtro_actual":
            if id_new:
                servicios = servicios.filter(id_new__icontains=id_new)
            if mes_produccion:
                servicios = servicios.filter(
                    mes_produccion__icontains=mes_produccion)
        elif filtro_pdf == "mes_actual":
            servicios = servicios.filter(mes_produccion__iexact=mes_actual)

        # Si no hay datos, lanzamos excepción
        if not servicios.exists():
            raise ValueError("No hay datos para exportar.")

        # Datos PDF
        produccion_data = []
        total_produccion = Decimal("0.0")
        for servicio in servicios:
            total_mmoo = servicio.monto_mmoo or Decimal("0.0")
            total_tecnicos = servicio.trabajadores_asignados.count()
            monto_tecnico = total_mmoo / \
                total_tecnicos if total_tecnicos else Decimal("0.0")

            produccion_data.append([
                f"DU{servicio.du}",
                servicio.id_new or "-",
                Paragraph(servicio.detalle_tarea or "-", ParagraphStyle(
                    'detalle_style', fontSize=9, leading=11, alignment=0)),
                f"{monto_tecnico:,.0f}".replace(",", ".")
            ])

            total_produccion += monto_tecnico

        # Generación PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4,
                                topMargin=50, bottomMargin=50)
        elements = []
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="CenterTitle",
                   alignment=1, fontSize=16, spaceAfter=20))

        # Títulos
        elements.append(Paragraph(
            f"Producción del Técnico: {usuario.get_full_name()}", styles["CenterTitle"]))
        elements.append(Paragraph(
            f"<b>Total Producción:</b> ${total_produccion:,.0f} CLP".replace(",", "."), styles["Normal"]))
        elements.append(Paragraph(
            f"<i>El total corresponde a la selección:</i> {filtro_seleccionado}.", styles["Normal"]))
        elements.append(Paragraph(
            f"<b>Fecha de generación:</b> {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
        elements.append(Spacer(1, 12))

        # Tabla
        data = [["DU", "ID NEW", "Detalle",
                 "Producción (CLP)"]] + produccion_data
        table = Table(data, colWidths=[70, 100, 300, 80])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0e7490")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.whitesmoke, colors.lightgrey]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.gray),
        ]))
        elements.append(table)

        # Firma
        elements.append(Spacer(1, 40))
        elements.append(
            Paragraph("<b>Firma del Técnico:</b>", styles["Normal"]))
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            f"__________________________<br/>{usuario.get_full_name()}", styles["Normal"]))

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="produccion.pdf"'
        return response

    except Exception as e:
        logger.error(f"Error exportando PDF: {e}")
        return HttpResponse(f"Error generando PDF: {e}", status=500)


User = get_user_model()


# Helper: siguiente DU disponible (correlativo, sin duplicar)


def _next_du():
    last_num = (
        ServicioCotizado.objects
        .exclude(du__isnull=True)
        .exclude(du__exact="")
        .annotate(
            du_num=Cast(
                Replace("du", Value("DU"), Value("")),
                IntegerField()
            )
        )
        .aggregate(m=Max("du_num"))
        .get("m") or 0
    )
    return int(last_num) + 1


@require_POST
@login_required
@transaction.atomic
def crear_ajuste(request):
    """
    Crea un ajuste (bono / adelanto / descuento) para uno o más técnicos.
    - Si la petición es AJAX (X-Requested-With / Accept JSON / ?ajax=1):
        * setea messages igualmente (en sesión)
        * responde JSON con {'ok': ..., 'redirect': <url>}
        * el frontend debe hacer window.location = redirect
    - Si NO es AJAX: PRG clásico -> setea messages y redirect.
    """
    # ¿Es AJAX?
    xrw = (request.headers.get('X-Requested-With')
           or request.META.get('HTTP_X_REQUESTED_WITH') or '')
    accept = (request.headers.get('Accept') or '')
    is_ajax = (
        request.GET.get('ajax') == '1'
        or xrw.lower() == 'xmlhttprequest'
        or 'application/json' in accept.lower()
        or 'json' in accept.lower()
    )
    go_back = reverse('operaciones:produccion_admin')

    try:
        User = get_user_model()

        tipo = (request.POST.get('tipo') or '').strip()
        mes_texto = (request.POST.get('mes_texto') or '').strip()
        detalle = (request.POST.get('detalle') or '').strip()
        monto = Decimal(str(request.POST.get('monto') or '0')
                        ).quantize(Decimal('0.01'))
        asignados = request.POST.getlist('asignados')

        # Validaciones
        if not tipo or not mes_texto or not detalle or not asignados:
            msg = 'Faltan campos obligatorios.'
            messages.error(request, msg)
            if is_ajax:
                return JsonResponse({'ok': False, 'error': msg, 'redirect': go_back}, status=400)
            return redirect(go_back)

        if monto <= 0:
            msg = 'El monto debe ser mayor a 0.'
            messages.error(request, msg)
            if is_ajax:
                return JsonResponse({'ok': False, 'error': msg, 'redirect': go_back}, status=400)
            return redirect(go_back)

        if tipo not in {'ajuste_bono', 'ajuste_adelanto', 'ajuste_descuento'}:
            msg = 'Tipo inválido.'
            messages.error(request, msg)
            if is_ajax:
                return JsonResponse({'ok': False, 'error': msg, 'redirect': go_back}, status=400)
            return redirect(go_back)

        # Regla de signo: sólo descuento resta; bono y adelanto suman
        monto_mmoo = -monto if tipo == 'ajuste_descuento' else monto

        created_ids = []
        for uid in asignados:
            user = User.objects.filter(id=uid).first()
            if not user:
                continue

            du_text = f"{_next_du():08d}"  # DU correlativo "DU00000047"

            s = ServicioCotizado.objects.create(
                du=du_text,
                id_claro='-',
                region='-',
                mes_produccion=mes_texto,
                id_new='-',
                detalle_tarea=detalle,
                monto_cotizado=0,
                monto_mmoo=monto_mmoo,
                estado=tipo,
                fecha_aprobacion_supervisor=None,
            )
            s.trabajadores_asignados.add(user)
            created_ids.append(s.id)

        # OK: colocar message y responder según modo
        messages.success(request, 'Ajuste(s) creado(s) correctamente.')
        if is_ajax:
            return JsonResponse({'ok': True, 'created': created_ids, 'redirect': go_back})
        return redirect(go_back)

    except Exception as e:
        # Error inesperado -> message + redirect/JSON
        err = f'No se pudo crear el/los ajuste(s): {e}'
        messages.error(request, err)
        if is_ajax:
            return JsonResponse({'ok': False, 'error': err, 'redirect': go_back}, status=500)
        return redirect(go_back)


AJUSTES_SET = {'ajuste_bono', 'ajuste_adelanto', 'ajuste_descuento'}


def _yyyymm_from_mes_texto(txt: str) -> str:
    """
    'Octubre 2025' -> '2025-10' (best-effort). Si falla, retorna ''.
    """
    if not txt:
        return ''
    meses = ["enero", "febrero", "marzo", "abril", "mayo", "junio",
             "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    s = txt.strip().lower()
    m = re.search(r'([a-záéíóúü]+)\s+(\d{4})', s)
    if not m:
        return ''
    name, y = m.group(1), int(m.group(2))
    try:
        mm = next(i+1 for i, n in enumerate(meses) if name.startswith(n[:3]))
    except StopIteration:
        return ''
    return f"{y:04d}-{mm:02d}"


@login_required
@require_http_methods(["GET", "POST"])
@transaction.atomic
def editar_ajuste(request, pk: int):
    """
    GET  -> muestra la página de edición (misma UI que 'nuevo ajuste', pre-rellena valores)
    POST -> actualiza el ajuste y redirige a la lista con un flash
    """
    s = get_object_or_404(
        ServicioCotizado,
        pk=pk,
        estado__in=AJUSTES_SET
    )
    User = get_user_model()

    if request.method == "GET":
        # monto positivo para la UI
        monto_pos = Decimal(str(abs(s.monto_mmoo or 0))
                            ).quantize(Decimal('0.01'))
        init = {
            "id": s.id,
            "tipo": s.estado,  # ajuste_bono / ajuste_adelanto / ajuste_descuento
            "mes_texto": s.mes_produccion or "",
            "mes_yyyy_mm": _yyyymm_from_mes_texto(s.mes_produccion or ""),
            "detalle": s.detalle_tarea or "",
            "monto": str(monto_pos),
            "asignados_ids": list(s.trabajadores_asignados.values_list('id', flat=True)),
        }
        usuarios = User.objects.filter(is_active=True).order_by(
            "first_name", "last_name", "username"
        )

        # Usamos el mismo template de "nuevo", pero en modo edición
        return render(request, "operaciones/ajuste_nuevo.html", {
            "usuarios": usuarios,
            "modo": "edit",
            "initial": init,
        })

    # ===== POST -> guardar cambios =====
    tipo = (request.POST.get('tipo') or '').strip()
    mes_texto = (request.POST.get('mes_texto') or '').strip()
    detalle = (request.POST.get('detalle') or '').strip()
    asignados = request.POST.getlist('asignados')

    try:
        monto = Decimal(str(request.POST.get('monto') or '0')
                        ).quantize(Decimal('0.01'))
    except Exception:
        return HttpResponseBadRequest("Monto inválido")

    # Validaciones
    if tipo not in AJUSTES_SET:
        return HttpResponseBadRequest("Tipo inválido")
    if not mes_texto or not detalle or not asignados:
        return HttpResponseBadRequest("Faltan campos")
    if monto <= 0:
        return HttpResponseBadRequest("El monto debe ser > 0")

    # Signo y campos
    s.monto_mmoo = -monto if tipo == 'ajuste_descuento' else monto
    s.estado = tipo
    s.mes_produccion = mes_texto
    s.detalle_tarea = detalle

    users = list(User.objects.filter(id__in=asignados, is_active=True))
    if not users:
        return HttpResponseBadRequest("Asignados inválidos")

    # ⬇️ Guardar SOLO con campos existentes
    fields = ["monto_mmoo", "estado", "mes_produccion", "detalle_tarea"]
    # agrega 'updated_at' solo si tu modelo realmente lo tiene
    try:
        if any(getattr(f, "name", "") == "updated_at" for f in s._meta.get_fields()):
            fields.append("updated_at")
    except Exception:
        pass

    s.save(update_fields=fields)
    s.trabajadores_asignados.set(users)

    messages.success(request, "Ajuste actualizado.")
    return redirect(f"{reverse('operaciones:produccion_admin')}?flash=ajuste_edit_ok")


AJUSTE_ESTADOS = {"ajuste_bono", "ajuste_adelanto", "ajuste_descuento"}


def _is_admin(user):
    # Ajusta si tu flag de admin es otro
    return user.is_superuser or getattr(user, "es_admin_general", False)


def _bad_request(msg):
    return JsonResponse({"ok": False, "error": msg}, status=400)


@login_required
def eliminar_ajuste(request, pk: int):
    """
    Procesa el botón 'Eliminar' del modal (form POST).
    Valida permisos, tipo de ajuste y que no esté pagado.
    """
    if request.method != "POST":
        return HttpResponseBadRequest("Método inválido.")

    s = get_object_or_404(ServicioCotizado, pk=pk)

    if not _is_admin(request.user):
        return HttpResponseForbidden("No autorizado.")
    if s.estado not in AJUSTE_ESTADOS:
        return HttpResponseBadRequest("El registro no es un ajuste.")
    if getattr(s, "pagado", False):
        messages.error(
            request, "No puedes eliminar un ajuste que ya fue pagado.")
        return redirect(reverse("operaciones:produccion_admin"))

    s.delete()
    # Lleva un flash simple por querystring como usas en JS (opcionalmente usa messages)
    url = reverse("operaciones:produccion_admin")
    return redirect(f"{url}?flash=ajuste_del_ok")


@login_required
def ajuste_nuevo(request):
    """
    Renderiza una página para crear un Bono/Adelanto/Descuento sin usar modal.
    El envío se hace por fetch POST al endpoint existente 'operaciones:crear_ajuste'
    (que ya usabas desde el modal).
    """
    usuarios = User.objects.filter(is_active=True).order_by(
        "first_name", "last_name", "username")
    current_month = timezone.localdate().strftime("%Y-%m")
    return render(request, "operaciones/ajuste_nuevo.html", {
        "usuarios": usuarios,
        "current_month": current_month,
    })


@require_POST
@login_required
@transaction.atomic
def admin_restore_user_pending_monthly(request, pk: int):
    """
    Permite al admin restaurar un MonthlyPayment rechazado por el técnico
    a 'pending_user' para que el técnico vuelva a aprobar/rechazar.
    """
    mp = get_object_or_404(MonthlyPayment, pk=pk)

    # Señales de que esperan JSON (igual que en otras vistas)
    xrw = (request.headers.get('X-Requested-With')
           or request.META.get('HTTP_X_REQUESTED_WITH') or '')
    accept = (request.headers.get('Accept') or '')
    is_ajax = (
        request.GET.get('ajax') == '1'
        or xrw.lower() == 'xmlhttprequest'
        or 'application/json' in accept.lower()
        or 'json' in accept.lower()
    )

    if mp.status != "rejected_user":
        if is_ajax:
            return JsonResponse({"ok": False, "error": "only_rejected_can_be_restored"}, status=400)
        messages.info(
            request, "Sólo puedes restaurar pagos en estado 'Rechazado por el técnico'.")
        return redirect("operaciones:admin_monthly_payments")

    # No borramos el motivo; queda como historial. (Si prefieres, limpia: mp.reject_reason="")
    mp.status = "pending_user"
    mp.save(update_fields=["status", "updated_at"])

    if is_ajax:
        return JsonResponse({"ok": True})

    messages.success(
        request, "Restaurado a 'Pendiente de aprobación del técnico'.")
    return redirect("operaciones:admin_monthly_payments")
