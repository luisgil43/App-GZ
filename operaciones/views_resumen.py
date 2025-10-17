# operaciones/views_resumen.py
from decimal import Decimal
from urllib.parse import urlencode

from django.db.models import Count, Sum, Value, DecimalField, Max
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import render

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .models import ServicioCotizado

# Orden y etiquetas “amigables” exactamente en el orden que quieres verlos
ESTADOS_ORDER = [
    "cotizado",
    "aprobado_pendiente",
    "asignado",
    "en_progreso",
    "en_revision_supervisor",
    "rechazado_supervisor",
    "aprobado_supervisor",
]
ESTADOS_LABEL = dict(ServicioCotizado.ESTADOS)  # {'cotizado': 'Cotizado', ...}


def _build_resumen_data(request):
    """
    Devuelve:
      meses_disponibles, mes_sel, secciones, total_global_count, total_global_uf
    Corrige el caso 'Todos': si el parámetro ?mes= existe pero viene vacío,
    NO selecciona el último mes por defecto.
    """
    # 1) Meses disponibles (ordenados por última creación)
    meses_qs = (
        ServicioCotizado.objects
        .values("mes_produccion")
        .annotate(last_created=Max("fecha_creacion"))
        .order_by("-last_created")
    )
    meses_disponibles = [m["mes_produccion"] for m in meses_qs if m["mes_produccion"]]

    # 2) Selección de mes
    if "mes" in request.GET:  # el usuario hizo una elección (incluye vacío = Todos)
        mes_sel = (request.GET.get("mes") or "").strip()
    else:
        # Sin parámetro en la URL: por usabilidad, si existen meses, toma el último.
        mes_sel = meses_disponibles[0] if meses_disponibles else ""

    # 3) Base filtrada por mes si corresponde
    base = ServicioCotizado.objects.all()
    if mes_sel:
        base = base.filter(mes_produccion=mes_sel)

    # 4) Agregado por estado
    agregados = (
        base.values("estado")
        .annotate(
            total=Count("id"),
            total_uf=Coalesce(
                Sum("monto_cotizado"),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=12, decimal_places=2)),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
    )
    agg_by_estado = {a["estado"]: a for a in agregados}

    # 5) Items por estado (para el desplegable) — SIN la columna 'detalle'
    qs_items = (
        base.values("estado", "du", "id_claro", "monto_cotizado")
            .order_by("du")
    )
    items_by_estado = {k: [] for k in ESTADOS_ORDER}
    for it in qs_items:
        estado = it["estado"]
        if estado in items_by_estado:
            items_by_estado[estado].append(it)

    # 6) Estructura final
    secciones = []
    for estado in ESTADOS_ORDER:
        agg = agg_by_estado.get(estado, {"total": 0, "total_uf": Decimal("0.00")})
        secciones.append({
            "key": estado,
            "label": ESTADOS_LABEL.get(estado, estado.replace("_", " ").title()),
            "count": agg["total"],
            "total_uf": agg["total_uf"] or Decimal("0.00"),
            "items": items_by_estado.get(estado, []),
        })

    total_global_count = sum(s["count"] for s in secciones)
    total_global_uf = sum(s["total_uf"] for s in secciones)

    return meses_disponibles, mes_sel, secciones, total_global_count, total_global_uf


def resumen_operativo(request):
    meses_disponibles, mes_sel, secciones, total_count, total_uf = _build_resumen_data(request)

    contexto = {
        "meses_disponibles": meses_disponibles,
        "mes_sel": mes_sel,
        "secciones": secciones,
        "total_global_count": total_count,
        "total_global_uf": total_uf,
        # Para que el botón de export mantenga los filtros actuales
        "current_query": urlencode({"mes": mes_sel}) if mes_sel or ("mes" in request.GET) else "",
    }
    return render(request, "operaciones/resumen_operativo.html", contexto)


def export_resumen_operativo_xlsx(request):
    """
    Exporta a Excel el mismo resumen que se ve en pantalla, respetando ?mes=.
    """
    meses_disponibles, mes_sel, secciones, total_count, total_uf = _build_resumen_data(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    fila = 1

    def write(row, col, value, bold=False):
        c = ws.cell(row=row, column=col, value=value)
        if bold:
            c.font = c.font.copy(bold=True)
        return c

    # Encabezado
    titulo = "Resumen Operativo — Proyectos por estado"
    write(fila, 1, titulo, bold=True)
    fila += 1
    write(fila, 1, f"Mes: {mes_sel if (mes_sel or 'mes' in request.GET) else 'Todos'}")
    fila += 1
    write(fila, 1, f"Total proyectos: {total_count}")
    write(fila, 3, float(total_uf))  # Excel prefiere float; Decimal también funciona, pero así evitamos formato extraño
    write(fila, 2, "Total UF:")
    fila += 2

    # Por cada sección
    for s in secciones:
        # Solo imprimimos secciones con datos (opcional; comenta si quieres todas)
        if not s["count"]:
            continue

        # Título de sección
        write(fila, 1, s["label"], bold=True)
        write(fila, 2, "Proyectos:")
        write(fila, 3, s["count"])
        write(fila, 4, "Total UF:")
        write(fila, 5, float(s["total_uf"]))
        fila += 1

        # Tabla
        headers = ["DU", "ID Claro", "UF"]
        for idx, h in enumerate(headers, start=1):
            write(fila, idx, h, bold=True)
        fila += 1

        for it in s["items"]:
            write(fila, 1, f"DU{it['du']}")
            write(fila, 2, it.get("id_claro") or "")
            write(fila, 3, float(it.get("monto_cotizado") or 0))
            fila += 1

        # Subtotal
        write(fila, 2, f"Subtotal UF ({s['label']})", bold=True)
        write(fila, 3, float(s["total_uf"]), bold=True)
        fila += 2

    # Auto ancho de columnas
    for col in range(1, 6):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=fila):
            val = row[0].value
            if val is None:
                continue
            val_str = str(val)
            max_len = max(max_len, len(val_str))
        ws.column_dimensions[get_column_letter(col)].width = max(10, min(60, max_len + 2))

    # Responder
    nombre_mes = (mes_sel or "Todos").replace(" ", "_")
    filename = f"ResumenOperativo_{nombre_mes}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



# === Comparativas por MES/AÑO usando mes_produccion (MISMA URL/NOMBRE) ===
from django.http import JsonResponse
from django.db.models.functions import Substr, Length

import re
import json

def _meses_disponibles_ordenados():
    """
    Lista de mes_produccion distintos, ordenados por última creación (reciente primero).
    Ej: ["Octubre 2025", "Septiembre 2025", ...]
    """
    qs = (
        ServicioCotizado.objects
        .values("mes_produccion")
        .annotate(last_created=Max("fecha_creacion"))
        .order_by("-last_created")
    )
    return [r["mes_produccion"] for r in qs if r["mes_produccion"]]

def _anios_disponibles_ordenados():
    """
    Extrae año desde el final de mes_produccion (… 'Julio 2025') y devuelve
    lista única ordenada por recencia (usando last_created).
    """
    year_expr = Substr("mes_produccion", Length("mes_produccion") - 3, 4)
    qs = (
        ServicioCotizado.objects
        .annotate(year=year_expr)
        .values("year")
        .annotate(last_created=Max("fecha_creacion"))
        .order_by("-last_created")
    )
    anios = [r["year"] for r in qs if r["year"] and re.fullmatch(r"\d{4}", r["year"])]
    seen = set(); out = []
    for y in anios:
        if y not in seen:
            seen.add(y); out.append(y)
    return out

def _agg_por_mes(base_qs, metric):
    """
    Agrupa por mes_produccion -> dict {mes: valor}
    """
    agg = (
        base_qs.values("mes_produccion")
        .annotate(
            proyectos=Count("id"),
            total_uf=Coalesce(
                Sum("monto_cotizado"),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=12, decimal_places=2)),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
        .order_by()
    )
    if metric == "count":
        return {r["mes_produccion"]: int(r["proyectos"]) for r in agg}
    return {r["mes_produccion"]: float(r["total_uf"]) for r in agg}

def _agg_por_anio(base_qs, metric):
    """
    Agrupa por año (extraído de mes_produccion) -> dict {YYYY: valor}
    """
    year_expr = Substr("mes_produccion", Length("mes_produccion") - 3, 4)
    agg = (
        base_qs.annotate(year=year_expr)
        .values("year")
        .annotate(
            proyectos=Count("id"),
            total_uf=Coalesce(
                Sum("monto_cotizado"),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=12, decimal_places=2)),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            ),
        )
        .order_by()
    )
    agg = [r for r in agg if r["year"] and re.fullmatch(r"\d{4}", r["year"])]
    if metric == "count":
        return {r["year"]: int(r["proyectos"]) for r in agg}
    return {r["year"]: float(r["total_uf"]) for r in agg}

def _auto_prev_list(full_order_list, selected_list):
    """
    Calcula B como el período anterior a A dentro del orden maestro (reciente primero).
    Mantiene la misma longitud (si hay elementos suficientes).
    """
    if not selected_list:
        return []
    idxs = [full_order_list.index(x) for x in selected_list if x in full_order_list]
    if not idxs:
        return []
    start = min(idxs); end = max(idxs)
    length = (end - start) + 1
    b_start = end + 1
    b_end = b_start + length - 1
    return [full_order_list[i] for i in range(b_start, min(b_end + 1, len(full_order_list)))]

def comparativas_productividad(request):
    """
    (MISMO ENDPOINT) Comparativa por mes/año usando mes_produccion.

    Parámetros GET:
      g=month|year (default: month)
      m=count|uf (default: uf)
      estados=<multi>
      auto=1|0
      A=<multi>  (meses/años según g)  -> SERIE A (azul)
      B=<multi>  (meses/años según g)  -> SERIE B (rojo)
    Empareja por ÍNDICE: Par 1 = A[0] vs B[0], Par 2 = A[1] vs B[1], etc.
    El eje X muestra los pares y los tooltips enseñan el período real.
    """
    g = (request.GET.get("g") or "month").lower()
    m = (request.GET.get("m") or "uf").lower()
    estados = request.GET.getlist("estado") or request.GET.getlist("estados")

    # 'auto' por defecto solo en primera carga
    auto = (request.GET.get("auto") == "1") if request.GET else True

    # Listas maestras (reciente -> antiguo)
    master_months = _meses_disponibles_ordenados()
    master_years = _anios_disponibles_ordenados()

    if g == "year":
        master = master_years
        g = "year"
    else:
        master = master_months
        g = "month"

    # Selecciones A/B
    selA = [x for x in request.GET.getlist("A") if x in master]
    if not selA:
        selA = master[:3]

    selB = [x for x in request.GET.getlist("B") if x in master]
    if auto and not selB:
        selB = _auto_prev_list(master, selA)

    # Orden cronológico (antiguo -> reciente) y listas ordenadas para A y B
    chronological = list(reversed(master))
    orderedA = [x for x in chronological if x in selA]
    orderedB = [x for x in chronological if x in selB]

    # Base filtrada por estados
    base = ServicioCotizado.objects.all()
    if estados:
        base = base.filter(estado__in=estados)

    # Agregaciones globales por periodo (tomaremos por key)
    if g == "year":
        agg_all = _agg_por_anio(base, m)
    else:
        agg_all = _agg_por_mes(base, m)

    # Construir pares por índice
    n = max(len(orderedA), len(orderedB)) or 0
    labelsA = [orderedA[i] if i < len(orderedA) else "" for i in range(n)]
    labelsB = [orderedB[i] if i < len(orderedB) else "" for i in range(n)]
    # 👇 ahora sí podemos usar labelsA/labelsB
    labels_pairs = [f"{labelsA[i] or '—'} vs {labelsB[i] or '—'}" for i in range(n)]

    serieA = [float(agg_all.get(labelsA[i], 0)) for i in range(n)]
    serieB = [float(agg_all.get(labelsB[i], 0)) for i in range(n)]

    # Totales y KPIs
    totalA = float(sum(serieA))
    totalB = float(sum(serieB))
    delta = totalA - totalB
    pct = (delta / totalB * 100.0) if totalB else None

    # Tabla por par
    table_rows = []
    for i in range(n):
        a = serieA[i]; b = serieB[i]
        d = a - b
        p = (d / b * 100.0) if b else None
        table_rows.append({
            "label_a": labelsA[i],
            "a": a,
            "label_b": labelsB[i],
            "b": b,
            "delta": d,
            "pct": p,
        })

    context = {
        # Controles
        "g": g,
        "m": m,
        "estado": "",
        "estados_opciones": list(ServicioCotizado.ESTADOS),
        "estados_sel": estados,
        "meses_disponibles": master_months,
        "anios_disponibles": master_years,
        "selA": selA,
        "selB": selB,
        "auto": "1" if auto else "0",

        # Datos para Chart.js
        "labels_json": json.dumps(labels_pairs),
        "cur_values_json": json.dumps(serieA),   # A (azul)
        "prev_values_json": json.dumps(serieB),  # B (rojo)
        "labelsA_json": json.dumps(labelsA),
        "labelsB_json": json.dumps(labelsB),

        # KPIs
        "cur_total": totalA,
        "prev_total": totalB,
        "delta": delta,
        "pct": pct,

        # Tabla
        "table_rows": table_rows,
    }
    return render(request, "operaciones/analytics_comparativas.html", context)