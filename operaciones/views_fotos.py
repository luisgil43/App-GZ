# operaciones/views_fotos.py
import io
import mimetypes
import os
import re
import unicodedata
import uuid
from datetime import datetime, timezone
from pathlib import Path
from tempfile import NamedTemporaryFile

import boto3
import xlsxwriter
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.db.models import Count  # ya lo tienes importado arriba
from django.db.models import Case, IntegerField, Max, Prefetch, Value, When
from django.http import (FileResponse, HttpResponse, HttpResponseBadRequest,
                         JsonResponse)
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.utils.text import slugify
from django.views.decorators.http import require_POST
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.workbook import Workbook
from PIL import ExifTags, Image, ImageOps
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
from usuarios.decoradores import rol_requerido

from .models import \
    SitioMovil  # para traer ID SITES, dirección, comuna, región
from .models import (EvidenciaFoto, RequisitoFoto, ServicioCotizado,
                     SesionFotos, SesionFotoTecnico, _site_name_for)

# ========== Helpers ==========




def _get_or_create_sesion(servicio: ServicioCotizado) -> SesionFotos:
    sesion, _ = SesionFotos.objects.get_or_create(servicio=servicio)

    existentes = {a.tecnico_id for a in sesion.asignaciones.all()}
    canon_asig = (
        sesion.asignaciones
        .filter(requisitos__isnull=False)
        .annotate(n=Count("requisitos"))
        .filter(n__gt=0)
        .first()
    )
    canon_reqs = []
    if canon_asig:
        canon_reqs = list(
            RequisitoFoto.objects
            .filter(tecnico_sesion=canon_asig, activo=True)  # 👈 SOLO activos
            .order_by("orden", "id")
        )

    for user in servicio.trabajadores_asignados.all():
        if user.id in existentes:
            continue
        a_nueva = SesionFotoTecnico.objects.create(
            sesion=sesion, tecnico=user, estado='asignado'
        )
        if canon_reqs:
            existing_norms = set(
                _norm_title(t) for t in
                RequisitoFoto.objects.filter(tecnico_sesion=a_nueva).values_list("titulo", flat=True)
            )
            to_create = []
            for r in canon_reqs:
                key = _norm_title(r.titulo)
                if key in existing_norms:
                    continue
                to_create.append(RequisitoFoto(
                    tecnico_sesion=a_nueva,
                    titulo=r.titulo,
                    descripcion=r.descripcion,
                    obligatorio=r.obligatorio,
                    orden=r.orden,
                    activo=True,
                ))
                existing_norms.add(key)
            if to_create:
                for r in to_create:
                    r.titulo_norm = _norm_title(r.titulo)
                RequisitoFoto.objects.bulk_create(to_create, ignore_conflicts=True)

        # Limpieza por si ya existían duplicados
        _dedupe_requisitos(a_nueva)

    # Limpia asignaciones huérfanas si se desasignó un técnico
    actuales = {u.id for u in servicio.trabajadores_asignados.all()}
    sesion.asignaciones.exclude(tecnico_id__in=actuales).delete()
    return sesion
# ============================
# SUPERVISOR — CONFIGURAR REQUISITOS
# ============================


# views.py


# ========== Helpers ==========

import re
import unicodedata

from django.db import transaction
from django.db.models import Count


def _norm_title(s: str) -> str:
    """
    Normaliza títulos para comparar:
    - trim
    - minúsculas
    - quita tildes
    - colapsa espacios internos
    """
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s


def _dedupe_requisitos(asig, allowed_norms: set | None = None):
    """
    En una asignación, deja a lo más 1 requisito activo por título normalizado.
    - Si allowed_norms es None: deduplica “clásico” (mantiene 1 por norma).
    - Si allowed_norms es un set: SOLO pueden quedar activos los de esas normas.
      Los demás se desactivan, respetando así los borrados desde el formulario.
    """
    reqs = list(
        RequisitoFoto.objects.filter(tecnico_sesion=asig).order_by("orden", "id")
    )

    # Elegimos el “ganador” por norma: preferimos (activo=True, menor orden, menor id)
    best_for_norm: dict[str, RequisitoFoto] = {}
    for r in reqs:
        norm = _norm_title(r.titulo)
        if allowed_norms is not None and norm not in allowed_norms:
            # Este título fue eliminado por el supervisor: no puede quedar activo
            continue
        cur = best_for_norm.get(norm)
        rank = (bool(getattr(r, "activo", True)), r.orden, r.id)
        if cur is None:
            best_for_norm[norm] = r
        else:
            cur_rank = (bool(getattr(cur, "activo", True)), cur.orden, cur.id)
            if rank < cur_rank:
                best_for_norm[norm] = r

    # Activar solo el ganador y desactivar el resto (y todo lo fuera de allowed_norms si aplica)
    to_activate, to_deactivate = [], []
    for r in reqs:
        norm = _norm_title(r.titulo)
        keep_active = (allowed_norms is None or norm in allowed_norms) and (best_for_norm.get(norm) is r)
        if keep_active:
            if hasattr(r, "activo") and not r.activo:
                r.activo = True
                to_activate.append(r)
        else:
            if getattr(r, "activo", True):
                r.activo = False
                to_deactivate.append(r)

    if to_activate:
        for r in to_activate:
            r.save(update_fields=["activo"])
    if to_deactivate:
        for r in to_deactivate:
            r.save(update_fields=["activo"])



@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def configurar_requisitos(request, servicio_id):
    servicio = get_object_or_404(ServicioCotizado, pk=servicio_id)
    sesion = _get_or_create_sesion(servicio)

    asignaciones = list(
        sesion.asignaciones
        .select_related("tecnico")
        .prefetch_related(
            Prefetch(
                "requisitos",
                queryset=RequisitoFoto.objects.filter(activo=True).order_by("orden", "id")
            )
        )
    )

    canonical = []
    if asignaciones and asignaciones[0].requisitos.exists():
        canonical = list(asignaciones[0].requisitos.all())

    if request.method == "POST":
        try:
            with transaction.atomic():
                sesion.proyecto_especial = bool(request.POST.get("proyecto_especial"))
                sesion.save(update_fields=["proyecto_especial"])

                ids    = request.POST.getlist("id[]")
                names  = request.POST.getlist("name[]")
                orders = request.POST.getlist("order[]")
                mand   = request.POST.getlist("mandatory[]")

                # Normalizamos el payload del formulario
                desired = []  # [(id_str_or_empty, orden, titulo, obligatorio, norm)]
                for i, nm in enumerate(names):
                    titulo = (nm or "").strip()
                    if not titulo:
                        continue
                    try:
                        orden = int(orders[i]) if i < len(orders) else i
                    except Exception:
                        orden = i
                    obligatorio = (mand[i] == "1") if i < len(mand) else True
                    rid = ids[i] if i < len(ids) else ""
                    desired.append((rid, orden, titulo, obligatorio, _norm_title(titulo)))

                # Conjunto de normas que DEBEN quedar activas tras guardar
                desired_norms = {x[4] for x in desired}

                for a in asignaciones:
                    existentes = list(RequisitoFoto.objects.filter(tecnico_sesion=a))
                    by_id = {str(r.id): r for r in existentes}

                    # Mapa por norma para merges/colisiones (el “mejor” por (activo, orden, id))
                    by_norm = {}
                    for r in existentes:
                        key = _norm_title(r.titulo)
                        cur = by_norm.get(key)
                        if cur is None:
                            by_norm[key] = r
                        else:
                            if (getattr(r, "activo", True), r.orden, r.id) < (getattr(cur, "activo", True), cur.orden, cur.id):
                                by_norm[key] = r

                    vistos_ids = set()

                    for rid, orden, titulo, obligatorio, norm in desired:
                        if rid and rid in by_id:
                            r = by_id[rid]
                            # Si el nuevo título colisiona con otro por norma, mergeamos
                            collision = by_norm.get(norm)
                            if collision and collision.id != r.id:
                                c = collision
                                upds = []
                                if c.obligatorio != obligatorio:
                                    c.obligatorio = obligatorio; upds.append("obligatorio")
                                if c.orden != orden:
                                    c.orden = orden; upds.append("orden")
                                # ⬇️ aseguramos normalizado correcto
                                if getattr(c, "titulo_norm", None) != norm:
                                    c.titulo_norm = norm; upds.append("titulo_norm")
                                if hasattr(c, "activo") and not c.activo:
                                    c.activo = True; upds.append("activo")
                                if upds:
                                    c.save(update_fields=upds)
                                if getattr(r, "activo", True):
                                    r.activo = False
                                    r.save(update_fields=["activo"])
                                vistos_ids.add(str(c.id))
                                by_norm[norm] = c
                                continue

                            cambios = []
                            if r.titulo != titulo:
                                r.titulo = titulo; cambios.append("titulo")
                            if r.orden != orden:
                                r.orden = orden; cambios.append("orden")
                            if r.obligatorio != obligatorio:
                                r.obligatorio = obligatorio; cambios.append("obligatorio")
                            # ⬇️ actualizar siempre el normalizado cuando cambie (o por saneo)
                            if getattr(r, "titulo_norm", None) != norm:
                                r.titulo_norm = norm; cambios.append("titulo_norm")
                            if hasattr(r, "activo") and not r.activo:
                                r.activo = True; cambios.append("activo")
                            if cambios:
                                r.save(update_fields=cambios)
                            vistos_ids.add(str(r.id))
                            by_norm[norm] = r
                        else:
                            exist = by_norm.get(norm)
                            if exist:
                                cambios = []
                                if exist.obligatorio != obligatorio:
                                    exist.obligatorio = obligatorio; cambios.append("obligatorio")
                                if exist.orden != orden:
                                    exist.orden = orden; cambios.append("orden")
                                if getattr(exist, "titulo_norm", None) != norm:
                                    exist.titulo_norm = norm; cambios.append("titulo_norm")
                                if hasattr(exist, "activo") and not exist.activo:
                                    exist.activo = True; cambios.append("activo")
                                if cambios:
                                    exist.save(update_fields=cambios)
                                vistos_ids.add(str(exist.id))
                            else:
                                nuevo = RequisitoFoto.objects.create(
                                    tecnico_sesion=a,
                                    titulo=titulo,
                                    titulo_norm=norm,   # ⬅️ NUEVO
                                    descripcion="",
                                    obligatorio=obligatorio,
                                    orden=orden,
                                    activo=True,
                                )
                                vistos_ids.add(str(nuevo.id))
                                by_norm[norm] = nuevo

                    # Desactivar (soft-delete) TODO lo no presente en el formulario
                    for r in existentes:
                        if str(r.id) not in vistos_ids and getattr(r, "activo", True):
                            r.activo = False
                            r.save(update_fields=["activo"])

                    # Dedupe final, pero SOLO sobre las normas permitidas
                    _dedupe_requisitos(a, allowed_norms=desired_norms)

            messages.success(request, "Requerimientos actualizados.")
            return redirect('operaciones:listar_servicios_supervisor')

        except Exception as e:
            messages.error(request, f"No se pudo guardar la lista: {e}")

    return render(
        request,
        "operaciones/fotos_configurar_requisitos.html",
        {
            "servicio": servicio,
            "sesion": sesion,
            "requirements": canonical,   # activos, ordenados
            "is_special": bool(sesion.proyecto_especial),
        },
    )

@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def import_requirements_page(request, servicio_id):
    servicio = get_object_or_404(ServicioCotizado, pk=servicio_id)
    sesion = _get_or_create_sesion(servicio)
    return render(
        request,
        "operaciones/fotos_importar_requisitos.html",
        {"servicio": servicio, "sesion": sesion},
    )


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def download_requirements_template(request, servicio_id, ext):
    ext = (ext or "").lower()
    filename_base = f"requisitos_template_servicio_{servicio_id}"

    if ext == "csv":
        content = (
            "name,order,mandatory\n"
            "Puerta frontal,0,1\n"
            "Tablero eléctrico,1,1\n"
            "Panorámica del sitio,2,0\n"
        )
        resp = HttpResponse(content, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
        return resp

    if ext in ("xlsx", "xls"):
        from io import BytesIO

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Requisitos"
        ws.append(["name", "order", "mandatory"])
        ws.append(["Puerta frontal", 0, 1])
        ws.append(["Tablero eléctrico", 1, 1])
        ws.append(["Panorámica del sitio", 2, 0])
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    return HttpResponseBadRequest("Formato no soportado. Usa csv o xlsx.")


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def importar_requisitos(request, servicio_id):
    import csv
    import io

    from openpyxl import load_workbook

    servicio = get_object_or_404(ServicioCotizado, pk=servicio_id)
    sesion = _get_or_create_sesion(servicio)

    f = request.FILES.get("file")
    if not f:
        messages.error(request, "Selecciona un archivo CSV o XLSX.")
        return redirect('operaciones:fotos_import_requirements_page', servicio_id=servicio_id)

    ext = (f.name.rsplit(".", 1)[-1] or "").lower()
    normalized = []  # [(order, name, mandatory), ...]

    # ---- Parseo del archivo
    try:
        if ext == "csv":
            raw = f.read().decode("utf-8", errors="ignore")
            lines = raw.splitlines()
            first = lines[0].lower() if lines else ""
            if "name" in first:
                reader = csv.DictReader(io.StringIO(raw))
                for row in reader:
                    name = (row.get("name") or "").strip()
                    if not name:
                        continue
                    order = int(row.get("order")) if (row.get("order") or "").strip().isdigit() else len(normalized)
                    mval = str(row.get("mandatory") or "1").strip().lower()
                    mandatory = mval in ("1", "true", "si", "sí", "yes", "y")
                    normalized.append((order, name, mandatory))
            else:
                reader = csv.reader(io.StringIO(raw))
                for row in reader:
                    if not row:
                        continue
                    name = (row[0] or "").strip()
                    if not name:
                        continue
                    normalized.append((len(normalized), name, True))

        elif ext in ("xlsx", "xls"):
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                messages.warning(request, "La planilla está vacía.")
                return redirect('operaciones:fotos_import_requirements_page', servicio_id=servicio_id)

            header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
            headered = "name" in header
            start = 1 if headered else 0

            if headered:
                i_name = header.index("name")
                i_order = header.index("order") if "order" in header else None
                i_mand = header.index("mandatory") if "mandatory" in header else None

                for r in rows[start:]:
                    name = (str(r[i_name]) if i_name < len(r) and r[i_name] is not None else "").strip()
                    if not name:
                        continue
                    if i_order is not None and i_order < len(r) and r[i_order] not in (None, ""):
                        try:
                            order = int(r[i_order])
                        except Exception:
                            order = len(normalized)
                    else:
                        order = len(normalized)

                    if i_mand is not None and i_mand < len(r) and r[i_mand] not in (None, ""):
                        mval = str(r[i_mand]).strip().lower()
                        mandatory = mval in ("1", "true", "si", "sí", "yes", "y")
                    else:
                        mandatory = True
                    normalized.append((order, name, mandatory))
            else:
                for r in rows:
                    if not r:
                        continue
                    name = (str(r[0]) if r[0] is not None else "").strip()
                    if not name:
                        continue
                    normalized.append((len(normalized), name, True))
        else:
            messages.error(request, "Tipo de archivo no soportado. Usa .csv o .xlsx.")
            return redirect('operaciones:fotos_import_requirements_page', servicio_id=servicio_id)

    except Exception as e:
        messages.error(request, f"No se pudo leer el archivo: {e}")
        return redirect('operaciones:fotos_import_requirements_page', servicio_id=servicio_id)

    if not normalized:
        messages.warning(request, "No se encontraron filas válidas.")
        return redirect('operaciones:fotos_import_requirements_page', servicio_id=servicio_id)

    # ---- Anexar / Reactivar (respetando índice único por (tecnico_sesion, titulo_norm))
    try:
        with transaction.atomic():
            asignaciones = list(
                sesion.asignaciones
                .select_related("tecnico")
                .prefetch_related("requisitos")
            )

            total_creados = 0
            total_omitidos = 0
            total_reactivados = 0

            for a in asignaciones:
                # punto de partida = último orden + 1 (solo activos si existe el flag)
                qs_max = RequisitoFoto.objects.filter(tecnico_sesion=a)
                if hasattr(RequisitoFoto, "activo"):
                    qs_max = qs_max.filter(activo=True)
                max_orden = qs_max.aggregate(m=Max("orden")).get("m")
                base = (max_orden or -1) + 1

                # mapa por nombre normalizado -> requisito existente (activo o no)
                existentes_map = {}
                for x in RequisitoFoto.objects.filter(tecnico_sesion=a).only("id", "titulo", "titulo_norm", "obligatorio", "orden"):
                    existentes_map[_norm_title(x.titulo)] = x

                nuevos = []
                omitidos_local = 0
                reactivados_local = 0

                # respetamos el orden del archivo, pero “desplazado” al final
                for i, (_o, name, mandatory) in enumerate(normalized):
                    key = _norm_title(name)
                    exist = existentes_map.get(key)

                    if exist:
                        # si el modelo tiene 'activo' y está desactivado, reactivar + actualizar
                        if hasattr(exist, "activo") and getattr(exist, "activo") is False:
                            cambios = []
                            if exist.obligatorio != mandatory:
                                exist.obligatorio = mandatory; cambios.append("obligatorio")
                            if exist.orden != (base + i):
                                exist.orden = (base + i); cambios.append("orden")
                            # ⬇️ asegurar que el normalizado coincide con el import
                            if getattr(exist, "titulo_norm", None) != key:
                                exist.titulo_norm = key; cambios.append("titulo_norm")
                            exist.activo = True; cambios.append("activo")
                            if cambios:
                                exist.save(update_fields=cambios)
                            reactivados_local += 1
                        else:
                            # ya existe activo con ese nombre -> lo omitimos (no duplicamos)
                            # (y, de paso, saneamos titulo_norm si está desalineado)
                            if getattr(exist, "titulo_norm", None) != key:
                                exist.titulo_norm = key
                                exist.save(update_fields=["titulo_norm"])
                            omitidos_local += 1
                        continue

                    # crear nuevo al final (con titulo_norm)
                    nuevos.append(
                        RequisitoFoto(
                            tecnico_sesion=a,
                            titulo=name,
                            titulo_norm=key,   # ⬅️ clave para el índice único
                            descripcion="",
                            obligatorio=mandatory,
                            orden=base + i,
                        )
                    )
                    # registra en el mapa para evitar duplicados dentro del mismo import
                    existentes_map[key] = True  # marcador simple

                if nuevos:
                    RequisitoFoto.objects.bulk_create(nuevos)
                    total_creados += len(nuevos)
                total_omitidos += omitidos_local
                total_reactivados += reactivados_local

        msg = f"Importados: {total_creados} agregados"
        if total_reactivados:
            msg += f", {total_reactivados} reactivados"
        if total_omitidos:
            msg += f", {total_omitidos} omitidos por duplicado"
        msg += "."
        messages.success(request, msg)
        return redirect('operaciones:fotos_configurar_requisitos', servicio_id=servicio_id)

    except Exception as e:
        messages.error(request, f"No se pudo aplicar la importación: {e}")
        return redirect('operaciones:fotos_import_requirements_page', servicio_id=servicio_id)


# ============================
# UTIL: normalizador + EXIF + HEIC
# ============================





def _to_jpeg_if_needed(uploaded_file):
    uploaded_file.seek(0)
    im = Image.open(uploaded_file)
    fmt = (im.format or "").upper()
    exif = im.info.get("exif")

    if fmt in {"HEIC", "HEIF"}:
        bio = io.BytesIO()
        im = im.convert("RGB")
        if exif:
            im.save(bio, format="JPEG", quality=92, exif=exif)
        else:
            im.save(bio, format="JPEG", quality=92)
        bio.seek(0)
        name = (uploaded_file.name.rsplit(".", 1)[0]) + ".jpg"
        return ContentFile(bio.read(), name=name)

    uploaded_file.seek(0)
    return uploaded_file


def _exif_to_latlng_taken_at(image):
    try:
        exif = getattr(image, "_getexif", lambda: None)()
        if not exif:
            return None, None, None

        tagmap = {ExifTags.TAGS.get(k, k): v for k, v in exif.items()}
        dt_raw = tagmap.get("DateTimeOriginal") or tagmap.get("DateTime")
        taken_at = None
        if dt_raw:
            from datetime import datetime
            try:
                taken_at = timezone.make_aware(
                    datetime.strptime(dt_raw, "%Y:%m:%d %H:%M:%S")
                )
            except Exception:
                taken_at = None

        gps_info = tagmap.get("GPSInfo")
        if not gps_info:
            return None, None, taken_at

        def _ratio_to_float(r):
            try:
                return float(r[0]) / float(r[1])
            except Exception:
                return float(r)

        def _dms_to_deg(dms, ref):
            deg = _ratio_to_float(dms[0])
            minutes = _ratio_to_float(dms[1])
            seconds = _ratio_to_float(dms[2])
            value = deg + (minutes / 60.0) + (seconds / 3600.0)
            if ref in ['S', 'W']:
                value = -value
            return value

        gps_tagmap = {ExifTags.GPSTAGS.get(
            k, k): v for k, v in gps_info.items()}
        lat = lng = None
        if all(k in gps_tagmap for k in ["GPSLatitude", "GPSLatitudeRef", "GPSLongitude", "GPSLongitudeRef"]):
            lat = _dms_to_deg(
                gps_tagmap["GPSLatitude"], gps_tagmap["GPSLatitudeRef"])
            lng = _dms_to_deg(
                gps_tagmap["GPSLongitude"], gps_tagmap["GPSLongitudeRef"])

        return lat, lng, taken_at
    except Exception:
        return None, None, None


SAFE_PREFIX = settings.DIRECT_UPLOADS_SAFE_PREFIX.rstrip("/") + "/"
ALLOWED_CT = ("image/jpeg", "image/png", "image/webp",
              "image/heic", "image/heif")


def _safe_key(filename: str, subdir: str = "") -> str:
    base = os.path.basename(filename or "upload")
    # sanea nombre
    base = re.sub(r"[^A-Za-z0-9._-]+", "_", base)
    ext = (os.path.splitext(base)[1] or ".jpg").lower()
    if ext not in (".jpg", ".jpeg", ".png", ".webp", ".heic", ".heif"):
        # fuerza extensión razonable
        ext = ".jpg"
    today = datetime.now(timezone.utc).strftime("%Y/%m/%d")
    uid = uuid.uuid4().hex[:12]
    sub = (subdir.strip("/") + "/") if subdir else ""
    return f"{SAFE_PREFIX}{sub}{today}/{uid}_{base.split('.')[0][:40]}{ext}"


@login_required
@require_POST
def presign_put(request, asig_id: int):
    if not settings.DIRECT_UPLOADS_ENABLED:
        return JsonResponse({"ok": False, "error": "Direct uploads disabled."}, status=400)

    a = get_object_or_404(SesionFotoTecnico, pk=asig_id, tecnico=request.user)

    filename = request.POST.get("filename") or ""
    content_type = request.POST.get("content_type") or mimetypes.guess_type(
        filename)[0] or "application/octet-stream"
    size_bytes = int(request.POST.get("size_bytes") or 0)

    if not content_type.startswith("image/") or content_type.lower() not in ALLOWED_CT:
        return JsonResponse({"ok": False, "error": "Tipo de archivo no permitido."}, status=400)

    max_bytes = settings.DIRECT_UPLOADS_MAX_MB * 1024 * 1024
    if size_bytes <= 0 or size_bytes > max_bytes:
        return JsonResponse({"ok": False, "error": f"Tamaño inválido (máximo {settings.DIRECT_UPLOADS_MAX_MB}MB)."}, status=400)

    key = _safe_key(filename, subdir=f"asig_{a.pk}")

    # Presign PUT (objeto será privado por defecto; no ACL)
    s3 = wasabi_client()
    url = s3.generate_presigned_url(
        ClientMethod="put_object",
        Params={
            "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
            "Key": key,
            "ContentType": content_type,
        },
        ExpiresIn=300,  # 5 minutos
    )
    head = s3.head_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=key)

    return JsonResponse({"ok": True, "url": url, "key": key, "content_type": content_type, "expires_in": 300})


@login_required
@require_POST
def finalize_upload(request, asig_id: int):
    """
    Finaliza una subida (direct-to-S3): valida el objeto, crea EvidenciaFoto
    y devuelve también el estado actual (faltantes por norma y can_finish)
    considerando SOLO requisitos ACTivos y agrupados por norma de título.
    """
    a = get_object_or_404(SesionFotoTecnico, pk=asig_id, tecnico=request.user)
    s = a.sesion

    # ========================= Helpers locales =========================
    # Límite global de "extras" (None => ilimitado)
    EXTRA_MAX = None
    # EXTRA_MAX = 400

    def _extra_count(sesion):
        return EvidenciaFoto.objects.filter(
            tecnico_sesion__sesion=sesion, requisito__isnull=True
        ).count()

    def _limit_reached(sesion, adding: int = 1) -> bool:
        if EXTRA_MAX is None:
            return False
        return _extra_count(sesion) + adding > EXTRA_MAX

    def _extras_left(sesion):
        if EXTRA_MAX is None:
            return None  # ilimitado
        return max(0, EXTRA_MAX - _extra_count(sesion))

    # ---- normalización y cómputo de faltantes/can_finish (solo activos) ----
    import re
    import unicodedata

    def _norm_title(s: str) -> str:
        s = (s or "").strip().lower()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
        s = re.sub(r"\s+", " ", s)
        return s

    def _canon_requisitos_por_norma():
        """
        Dict norm -> bloque canónico del requisito ACTIVO en la sesión:
        {"id","titulo","obligatorio","orden","ids": set(ids_equivalentes)}
        """
        canon_by_norm = {}
        qs = (RequisitoFoto.objects
              .filter(tecnico_sesion__sesion=s, activo=True)
              .values("id", "titulo", "obligatorio", "orden"))
        for r in qs:
            norm = _norm_title(r["titulo"])
            b = canon_by_norm.get(norm)
            if not b:
                canon_by_norm[norm] = {
                    "id": r["id"], "titulo": r["titulo"], "obligatorio": r["obligatorio"],
                    "orden": r["orden"], "ids": {r["id"]}
                }
            else:
                b["ids"].add(r["id"])
                # Conserva el “mejor” por (orden, id)
                if (r["orden"], r["id"]) < (b["orden"], b["id"]):
                    b.update({"id": r["id"], "titulo": r["titulo"],
                              "obligatorio": r["obligatorio"], "orden": r["orden"]})
        return canon_by_norm

    def _global_done_por_norma(canon_by_norm: dict):
        """
        Marca como done si existe al menos UNA evidencia para cualquiera de los IDs del bloque.
        """
        done = {norm: False for norm in canon_by_norm.keys()}
        if not canon_by_norm:
            return done

        all_ids = [rid for b in canon_by_norm.values() for rid in b["ids"]]
        ids_with_ev = set(
            EvidenciaFoto.objects
            .filter(requisito_id__in=all_ids)
            .values_list("requisito_id", flat=True)
        )
        for norm, b in canon_by_norm.items():
            if any(rid in ids_with_ev for rid in b["ids"]):
                done[norm] = True
        return done

    def _compute_faltantes_y_can_finish():
        canon = _canon_requisitos_por_norma()
        done = _global_done_por_norma(canon)

        faltantes = []
        for _, b in sorted(canon.items(), key=lambda x: (x[1]["orden"], x[1]["id"])):
            if b["obligatorio"] and not done.get(_, False):
                faltantes.append(b["titulo"])

        pendientes_aceptar = s.asignaciones.filter(estado="asignado").exists()
        can_finish = (len(faltantes) == 0) and (not pendientes_aceptar)
        return faltantes, can_finish
    # ==================================================================

    key = request.POST.get("key") or ""
    req_id = request.POST.get("req_id") or None
    nota = (request.POST.get("nota") or "").strip()
    lat = request.POST.get("lat") or None
    lng = request.POST.get("lng") or None
    acc = request.POST.get("acc") or None
    taken = request.POST.get("client_taken_at") or ""
    titulo_manual = (request.POST.get("titulo_manual") or "").strip()
    direccion_manual = (request.POST.get("direccion_manual") or "").strip()

    # Validaciones
    if not key or (".." in key) or (not key.startswith(SAFE_PREFIX)):
        return JsonResponse({"ok": False, "error": "Key inválida."}, status=400)

    # Límite global configurable de extras (si no hay requisito)
    if not req_id and _limit_reached(s, adding=1):
        lim = EXTRA_MAX
        return JsonResponse(
            {"ok": False, "error": f"Límite alcanzado: máximo {lim} fotos extra por proyecto."},
            status=400
        )

    # Verifica que el objeto exista y sea imagen
    s3 = wasabi_client()
    try:
        head = s3.head_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=key)
        ct = (head.get("ContentType") or "").lower()
        if not ct.startswith("image/"):
            return JsonResponse({"ok": False, "error": "El objeto subido no es imagen."}, status=400)
    except Exception:
        return JsonResponse({"ok": False, "error": "No se encontró el objeto en Wasabi."}, status=400)

    taken_dt = parse_datetime(taken) if taken else None

    # Si llega un req_id que quedó inactivo o fue mergeado, lo toleramos:
    # - si existe activo con el mismo título normalizado, la foto igual contará
    #   porque el cálculo de faltantes es por norma (no por id puntual).
    with transaction.atomic():
        ev = EvidenciaFoto(
            tecnico_sesion=a,
            requisito_id=(int(req_id) if req_id else None),
            nota=nota,
            lat=lat or None, lng=lng or None, gps_accuracy_m=acc or None,
            client_taken_at=taken_dt,
            titulo_manual=titulo_manual,
            direccion_manual=direccion_manual or "",
        )
        # ⚠️ no subas: asigna directamente el path en el campo
        ev.imagen.name = key
        ev.save()

    # URL (firmada si corresponde)
    try:
        url = default_storage.url(key)
    except Exception:
        url = ""

    titulo = ev.requisito.titulo if ev.requisito_id else (ev.titulo_manual or "Extra")
    fecha_txt = (ev.client_taken_at or ev.tomada_en).strftime("%Y-%m-%d %H:%M")

    # Estado actualizado tras esta subida
    faltantes, can_finish = _compute_faltantes_y_can_finish()

    return JsonResponse({
        "ok": True,
        "evidencia": {
            "id": ev.id, "url": url, "titulo": titulo, "fecha": fecha_txt,
            "lat": ev.lat, "lng": ev.lng, "acc": ev.gps_accuracy_m,
            "req_id": ev.requisito_id,
        },
        "extras_left": _extras_left(s),
        "max_extra": EXTRA_MAX,
        "faltantes_global": faltantes,   # <- para banner
        "can_finish": can_finish,        # <- para botón
    })

# operaciones/views_direct_uploads.py


def _wasabi_client():
    return boto3.client(
        "s3",
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        endpoint_url=settings.AWS_S3_ENDPOINT_URL,
        region_name=settings.AWS_S3_REGION_NAME,
    )


@login_required
@rol_requerido('usuario')
def upload_evidencias_fotos(request, pk):
    a = get_object_or_404(SesionFotoTecnico, pk=pk, tecnico=request.user)
    s = a.sesion
    servicio = s.servicio

    puede_subir = (a.estado == "en_proceso") or (
        a.estado == "rechazado_supervisor" and a.reintento_habilitado
    )

    EXTRA_MAX = None  # o un entero si quieres tope

    def _extra_count(sesion):
        return EvidenciaFoto.objects.filter(
            tecnico_sesion__sesion=sesion, requisito__isnull=True
        ).count()

    def _extras_left(sesion):
        if EXTRA_MAX is None:
            return None
        return max(0, EXTRA_MAX - _extra_count(sesion))

    if request.method == "POST":
        if not puede_subir:
            messages.info(request, "La asignación no está abierta para subir fotos.")
            return redirect("operaciones:fotos_upload", pk=a.pk)

        req_id = request.POST.get("req_id") or None
        nota = (request.POST.get("nota") or "").strip()
        lat = request.POST.get("lat") or None
        lng = request.POST.get("lng") or None
        acc = request.POST.get("acc") or None
        taken = request.POST.get("client_taken_at")
        taken_dt = parse_datetime(taken) if taken else None
        titulo_manual = (request.POST.get("titulo_manual") or "").strip()
        direccion_manual = (request.POST.get("direccion_manual") or "").strip()

        if s.proyecto_especial and not req_id and not titulo_manual:
            messages.error(request, "Ingresa un Título (proyecto especial).")
            return redirect("operaciones:fotos_upload", pk=a.pk)

        files = request.FILES.getlist("imagenes[]")

        if not req_id and EXTRA_MAX is not None:
            total_extra = _extra_count(s)
            cupo = max(0, EXTRA_MAX - total_extra)
            if cupo <= 0:
                messages.error(request, f"Ya se alcanzó el máximo de {EXTRA_MAX} fotos extra del proyecto.")
                return redirect("operaciones:fotos_upload", pk=a.pk)
            if len(files) > cupo:
                files = files[:cupo]
                messages.warning(request, f"Solo se aceptaron {cupo} foto(s) extra (límite {EXTRA_MAX} por proyecto).")

        n = 0
        for f in files:
            f_conv = _to_jpeg_if_needed(f)
            try:
                f_conv.seek(0)
                im = Image.open(f_conv)
                exif_lat, exif_lng, exif_dt = _exif_to_latlng_taken_at(im)
            except Exception:
                exif_lat = exif_lng = exif_dt = None
            finally:
                f_conv.seek(0)

            use_lat = lat or exif_lat
            use_lng = lng or exif_lng
            use_taken = taken_dt or exif_dt

            a.evidencias.create(
                requisito_id=req_id,
                imagen=f_conv,
                nota=nota,
                lat=use_lat, lng=use_lng, gps_accuracy_m=acc,
                client_taken_at=use_taken,
                titulo_manual=titulo_manual,
                direccion_manual=direccion_manual or "",
            )
            n += 1

        messages.success(request, f"{n} foto(s) subidas." if n else "No seleccionaste archivos.")
        return redirect("operaciones:fotos_upload", pk=a.pk)

    # ---------- GET: SYNC de requisitos ACTIVOS a nivel de sesión ----------
    # En vez de clonar solo cuando no hay ninguno, sincronizamos siempre lo faltante.
    # Canon = títulos activos de TODA la sesión, únicos por título normalizado.
    canon_map = {}  # norm -> (titulo, obligatorio, orden)
    for r in (RequisitoFoto.objects
              .filter(tecnico_sesion__sesion=s, activo=True)
              .order_by("orden", "id")
              .values("titulo", "obligatorio", "orden")):
        k = _norm_title(r["titulo"])
        if k not in canon_map:
            canon_map[k] = (r["titulo"], r["obligatorio"], r["orden"])

    mine = { _norm_title(x.titulo): x
             for x in RequisitoFoto.objects.filter(tecnico_sesion=a) }

    to_create = []
    for k, (titulo, oblig, orden) in canon_map.items():
        if k not in mine:
            to_create.append(RequisitoFoto(
                tecnico_sesion=a,
                titulo=titulo,
                descripcion="",
                obligatorio=oblig,
                orden=orden,
                activo=True,
            ))
        else:
            obj = mine[k]
            upds = []
            if hasattr(obj, "activo") and not obj.activo:
                obj.activo = True; upds.append("activo")
            if obj.obligatorio != oblig:
                obj.obligatorio = oblig; upds.append("obligatorio")
            if obj.orden != orden:
                obj.orden = orden; upds.append("orden")
            if upds:
                obj.save(update_fields=upds)

    if to_create:
        for r in to_create:
            r.titulo_norm = _norm_title(r.titulo)
        RequisitoFoto.objects.bulk_create(to_create, ignore_conflicts=True)
    # ----------------------------------------------------------------------

    # Requisitos de ESTA asignación (solo activos) con conteo "mío"
    requisitos = (
        a.requisitos
         .filter(activo=True)
         .annotate(uploaded=Count("evidencias"))
         .order_by("orden", "id")
    )

    # ¿Qué títulos ya tiene *cualquier* técnico del equipo? (match por título)
    titles_done = set(
        _norm_title(t) for t in
        EvidenciaFoto.objects
        .filter(tecnico_sesion__sesion=s, requisito_id__isnull=False)
        .values_list("requisito__titulo", flat=True)
        .distinct()
    )

    requisitos = list(requisitos)
    for r in requisitos:
        setattr(r, "completed_global", _norm_title(r.titulo) in titles_done)

    locked_ids = [r.id for r in requisitos if r.completed_global and r.uploaded == 0]

    # Faltantes globales: contar solo requisitos activos y obligatorios
    req_titles = (
        RequisitoFoto.objects
        .filter(tecnico_sesion__sesion=s, obligatorio=True, activo=True)
        .values_list("titulo", flat=True)
    )
    taken_titles = (
        s.asignaciones
         .values("requisitos__titulo")
         .annotate(c=Count("requisitos__evidencias"))
         .filter(c__gt=0)
         .values_list("requisitos__titulo", flat=True)
    )
    required_set = {_norm_title(t) for t in req_titles if t}
    taken_titles_set = {_norm_title(t) for t in taken_titles if t}
    faltantes_global = sorted(required_set - taken_titles_set)

    # ¿quién no ha aceptado?
    pendientes_aceptar = []
    for asg in s.asignaciones.select_related("tecnico"):
        accepted = bool(asg.aceptado_en) or asg.estado != "asignado"
        if not accepted:
            nombre = asg.tecnico.get_full_name() or asg.tecnico.username
            pendientes_aceptar.append(nombre)

    can_finish = (a.estado == "en_proceso" and not faltantes_global and not pendientes_aceptar)

    evidencias = _order_evidencias(a.evidencias.select_related("requisito"))

    return render(request, "operaciones/fotos_upload.html", {
        "a": a,
        "servicio": servicio,
        "requisitos": requisitos,
        "evidencias": evidencias,
        "locked_ids": locked_ids,
        "faltantes_global": faltantes_global,
        "pendientes_aceptar": pendientes_aceptar,
        "can_finish": can_finish,
        "is_proyecto_especial": s.proyecto_especial,
    })

# ============================
# SUPERVISOR — Revisión + generación Excel y Acta PDF
# ============================


def _order_evidencias(qs):
    """
    Ordena: primero con requisito (por orden del requisito), luego las extras (sin requisito),
    y dentro de cada grupo por fecha e id.
    """
    return (qs
            .annotate(
                is_extra=Case(
                    When(requisito__isnull=True, then=Value(1)),
                    default=Value(0),
                    output_field=IntegerField(),
                )
            )
            .order_by('is_extra', 'requisito__orden', 'tomada_en', 'id')
            )


@login_required
@rol_requerido('usuario')
@require_POST
def upload_evidencias_ajax(request, pk):
    a = get_object_or_404(SesionFotoTecnico, pk=pk, tecnico=request.user)
    s = a.sesion

    puede_subir = (a.estado == "en_proceso") or (
        a.estado == "rechazado_supervisor" and a.reintento_habilitado
    )
    if not puede_subir:
        return JsonResponse({"ok": False, "error": "Asignación no abierta para subir fotos."}, status=400)

    # ===== Helpers locales para límite de "extras" =====
    # Pon EXTRA_MAX = None para ILIMITADO, o un entero (p.ej., 400) para tope.
    EXTRA_MAX = None
    # EXTRA_MAX = 400

    def _extra_count(sesion):
        return EvidenciaFoto.objects.filter(
            tecnico_sesion__sesion=sesion, requisito__isnull=True
        ).count()

    def _limit_reached(sesion, adding: int = 1) -> bool:
        if EXTRA_MAX is None:
            return False
        return _extra_count(sesion) + adding > EXTRA_MAX

    def _extras_left(sesion):
        if EXTRA_MAX is None:
            return None
        return max(0, EXTRA_MAX - _extra_count(sesion))
    # ===================================================

    # Meta recibida
    req_id_raw = request.POST.get("req_id") or None
    nota = (request.POST.get("nota") or "").strip()
    lat = request.POST.get("lat") or None
    lng = request.POST.get("lng") or None
    acc = request.POST.get("acc") or None
    taken = request.POST.get("client_taken_at")
    taken_dt = parse_datetime(taken) if taken else None
    titulo_manual = (request.POST.get("titulo_manual") or "").strip()
    direccion_manual = (request.POST.get("direccion_manual") or "").strip()

    # Validar req_id: si viene, debe ser de ESTA asignación y estar ACTIVO
    req_id = None
    if req_id_raw:
        try:
            req_id_int = int(req_id_raw)
        except Exception:
            return JsonResponse({"ok": False, "error": "Requisito inválido."}, status=400)

        req_ok = RequisitoFoto.objects.filter(
            id=req_id_int,
            tecnico_sesion=a,
            activo=True,               # 👈 solo requisitos activos
        ).exists()

        if not req_ok:
            return JsonResponse({"ok": False, "error": "El requisito no existe o fue desactivado."}, status=400)

        req_id = req_id_int  # validado

    # Proyecto especial: si NO hay req_id, exigir título
    if s.proyecto_especial and not req_id and not titulo_manual:
        return JsonResponse({"ok": False, "error": "Ingresa un Título (proyecto especial)."}, status=400)

    # Límite GLOBAL de extras (solo aplica cuando NO hay req_id)
    if not req_id and _limit_reached(s, adding=1):
        lim = EXTRA_MAX
        return JsonResponse({"ok": False, "error": f"Límite alcanzado: máximo {lim} fotos extra por proyecto."}, status=400)

    file = request.FILES.get("imagen")
    if not file:
        return JsonResponse({"ok": False, "error": "No llegó la imagen."}, status=400)

    f_conv = _to_jpeg_if_needed(file)

    try:
        f_conv.seek(0)
        im = Image.open(f_conv)
        exif_lat, exif_lng, exif_dt = _exif_to_latlng_taken_at(im)
    except Exception:
        exif_lat = exif_lng = exif_dt = None
    finally:
        f_conv.seek(0)

    use_lat = lat or exif_lat
    use_lng = lng or exif_lng
    use_taken = taken_dt or exif_dt

    ev = a.evidencias.create(
        requisito_id=req_id,                 # 👈 solo llega si fue validado y activo
        imagen=f_conv,
        nota=nota,
        lat=use_lat, lng=use_lng, gps_accuracy_m=acc,
        client_taken_at=use_taken,
        titulo_manual=titulo_manual,
        direccion_manual=direccion_manual or "",
    )

    extras_left = _extras_left(s)

    # Título seguro
    if ev.requisito_id and ev.requisito:
        titulo = ev.requisito.titulo
    else:
        titulo = ev.titulo_manual or "Extra"

    fecha_txt = (ev.client_taken_at or ev.tomada_en).strftime("%Y-%m-%d %H:%M")

    return JsonResponse({
        "ok": True,
        "evidencia": {
            "id": ev.id,
            "url": ev.imagen.url,
            "titulo": titulo,
            "fecha": fecha_txt,
            "lat": ev.lat, "lng": ev.lng, "acc": ev.gps_accuracy_m,
            "req_id": ev.requisito_id if ev.requisito_id else None,
        },
        "extras_left": extras_left,
        "max_extra": EXTRA_MAX,
    })


# views.py

import re
import unicodedata

from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.http import JsonResponse

from .models import EvidenciaFoto, RequisitoFoto, SesionFotoTecnico


def _norm_title(s: str) -> str:
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"\s+", " ", s)
    return s

@login_required
def fotos_status_json(request, asignacion_id=None, asig_id=None, *args, **kwargs):
    """
    Acepta 'asignacion_id' o 'asig_id' según cómo esté definido el path().
    """
    asignacion_id = asignacion_id or asig_id

    a = (SesionFotoTecnico.objects
         .select_related("sesion")
         .filter(pk=asignacion_id, tecnico__is_active=True)  # opcional
         .first())
    if not a:
        return JsonResponse({"ok": False, "error": "Asignación no encontrada."}, status=404)
    s = a.sesion

    # ====== Límite global de EXTRAS ======
    EXTRA_MAX = None  # o un entero (p.ej. 400)

    def _extra_count(sesion):
        return EvidenciaFoto.objects.filter(
            tecnico_sesion__sesion=sesion, requisito__isnull=True
        ).count()

    def _extras_left(sesion):
        if EXTRA_MAX is None:
            return None
        return max(0, EXTRA_MAX - _extra_count(sesion))
    # =====================================

    # Requisitos activos de la sesión, consolidados por norma de título
    reqs_session_qs = (
        RequisitoFoto.objects
        .filter(tecnico_sesion__sesion=s, activo=True)
        .values("id", "titulo", "obligatorio", "orden")
    )

    canon_by_norm = {}  # norm -> {"id","titulo","obligatorio","orden","ids":set()}
    for r in reqs_session_qs:
        norm = _norm_title(r["titulo"])
        b = canon_by_norm.get(norm)
        if not b:
            canon_by_norm[norm] = {
                "id": r["id"], "titulo": r["titulo"], "obligatorio": r["obligatorio"],
                "orden": r["orden"], "ids": {r["id"]}
            }
        else:
            b["ids"].add(r["id"])
            if (r["orden"], r["id"]) < (b["orden"], b["id"]):
                b["id"] = r["id"]; b["titulo"] = r["titulo"]
                b["obligatorio"] = r["obligatorio"]; b["orden"] = r["orden"]

    global_done_by_norm = {norm: False for norm in canon_by_norm.keys()}
    if canon_by_norm:
        all_ids = [rid for v in canon_by_norm.values() for rid in v["ids"]]
        ids_with_ev = set(
            EvidenciaFoto.objects
            .filter(requisito_id__in=all_ids)
            .values_list("requisito_id", flat=True)
        )
        for norm, b in canon_by_norm.items():
            if any(rid in ids_with_ev for rid in b["ids"]):
                global_done_by_norm[norm] = True

    # Requisitos activos del técnico actual (para pintar badges por ID)
    reqs_tech = list(
        RequisitoFoto.objects
        .filter(tecnico_sesion=a, activo=True)
        .values("id", "titulo")
        .order_by("orden", "id")
    )
    requisitos_resp = []
    for r in reqs_tech:
        requisitos_resp.append({
            "id": r["id"],
            "global_done": bool(global_done_by_norm.get(_norm_title(r["titulo"]), False)),
        })

    # Faltantes globales (solo obligatorios activos por norma)
    faltantes_global = []
    for norm, b in sorted(canon_by_norm.items(), key=lambda x: (x[1]["orden"], x[1]["id"])):
        if b["obligatorio"] and not global_done_by_norm.get(norm, False):
            faltantes_global.append(b["titulo"])

    # Técnicos pendientes de aceptar (ajusta el estado según tu flujo)
    pendientes_aceptar = s.asignaciones.filter(estado="asignado").exists()

    can_finish = (len(faltantes_global) == 0) and (not pendientes_aceptar)

    return JsonResponse({
        "ok": True,
        "requisitos": requisitos_resp,
        "faltantes_global": faltantes_global,
        "can_finish": can_finish,
        "extras_left": _extras_left(s),
        "max_extra": EXTRA_MAX,
    })


@login_required
@rol_requerido('usuario')
def borrar_evidencia_foto(request, ev_id: int):
    ev = get_object_or_404(
        EvidenciaFoto,
        pk=ev_id,
        tecnico_sesion__tecnico=request.user
    )
    a = ev.tecnico_sesion
    permitido = (a.estado == 'en_proceso') or (
        a.estado == 'rechazado_supervisor' and a.reintento_habilitado
    )

    if not permitido:
        messages.warning(
            request, "No puedes eliminar evidencias en el estado actual.")
        return redirect('operaciones:fotos_upload', pk=a.pk)

    if request.method == "POST":
        # 🔒 Borrado seguro del archivo físico (si nadie más lo referencia)
        try:
            if ev.imagen and ev.imagen.name:
                same = EvidenciaFoto.objects.filter(imagen=ev.imagen.name)\
                                            .exclude(pk=ev.pk).exists()
                if not same:
                    ev.imagen.delete(save=False)
        except Exception:
            pass

        ev.delete()
        messages.success(request, "Foto eliminada correctamente.")
    else:
        messages.error(request, "Método no permitido.")

    return redirect('operaciones:fotos_upload', pk=a.pk)


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def borrar_evidencia_supervisor(request, ev_id: int):
    ev = get_object_or_404(EvidenciaFoto, pk=ev_id)
    servicio = ev.tecnico_sesion.sesion.servicio

    ESTADOS_PERMITIDOS = {'en_revision_supervisor',
                          'en_progreso', 'rechazado_supervisor'}
    if servicio.estado not in ESTADOS_PERMITIDOS:
        messages.warning(
            request, "No se puede eliminar evidencias en el estado actual del proyecto.")
        return redirect('operaciones:fotos_revisar_sesion', servicio_id=servicio.id)

    if request.method == "POST":
        # 🔒 Borrado seguro del archivo físico (si nadie más lo referencia)
        try:
            if ev.imagen and ev.imagen.name:
                same = EvidenciaFoto.objects.filter(imagen=ev.imagen.name)\
                                            .exclude(pk=ev.pk).exists()
                if not same:
                    ev.imagen.delete(save=False)
        except Exception:
            pass

        ev.delete()
        messages.success(request, "Evidencia eliminada correctamente.")
    else:
        messages.error(request, "Método no permitido.")

    return redirect('operaciones:fotos_revisar_sesion', servicio_id=servicio.id)


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def revisar_sesion_fotos(request, servicio_id):
    servicio = get_object_or_404(ServicioCotizado, pk=servicio_id)
    sesion = _get_or_create_sesion(servicio)
    asignaciones = (
        sesion.asignaciones
        .select_related("tecnico")
        .prefetch_related("evidencias__requisito")
        .all()
    )
    # Solo se puede aprobar/rechazar cuando la sesión está en revisión
    can_review = sesion.estado == "en_revision_supervisor"

    if request.method == "POST":
        accion = request.POST.get("accion")
        comentario = (request.POST.get("comentario") or "").strip()

        if not can_review:
            messages.error(
                request, "Este proyecto no está listo para revisión.")
            return redirect("operaciones:fotos_revisar_sesion", servicio_id=servicio.id)

        if accion == "aprobar":
            # ✅ Siempre generamos/reemplazamos al aprobar (también después de un rechazo)
            with transaction.atomic():
                # 1) Generar XLSX e ACTA (bytes)
                try:
                    xlsx_path = _xlsx_path_reporte_fotografico(servicio)
                except Exception as e:
                    messages.error(
                        request, f"No se pudo generar el informe: {e}")
                    return redirect("operaciones:fotos_revisar_sesion", servicio_id=servicio.id)

                try:
                    bytes_pdf = _bytes_acta_aceptacion(servicio)
                except Exception as e:
                    messages.error(request, f"No se pudo generar el acta: {e}")
                    return redirect("operaciones:fotos_revisar_sesion", servicio_id=servicio.id)

                from .models import _excel_filename, _pdf_filename

                # 2) Si ya existían archivos, eliminarlos para reescribir
                try:
                    if getattr(servicio.reporte_fotografico, "name", ""):
                        servicio.reporte_fotografico.delete(save=False)
                except Exception:
                    pass
                try:
                    if getattr(servicio.acta_aceptacion_pdf, "name", ""):
                        servicio.acta_aceptacion_pdf.delete(save=False)
                except Exception:
                    pass

                # 3) Adjuntar nuevos archivos
                excel_name = _excel_filename(servicio)
                pdf_name = _pdf_filename(
                    servicio, servicio.documento_compra or "DOC")

                try:
                    with open(xlsx_path, "rb") as f:
                        servicio.reporte_fotografico.save(
                            excel_name, File(f), save=False)
                finally:
                    try:
                        if xlsx_path and os.path.exists(xlsx_path):
                            os.unlink(xlsx_path)
                    except Exception:
                        pass

                servicio.acta_aceptacion_pdf.save(
                    pdf_name, ContentFile(bytes_pdf), save=False)

                # 4) Actualizar estados de servicio/sesión/asignaciones
                servicio.estado = "aprobado_supervisor"
                servicio.supervisor_aprobo = request.user
                servicio.fecha_aprobacion_supervisor = timezone.now()
                servicio._skip_report_signal = True  # evitar doble generación por signals

                servicio.save(update_fields=[
                    "reporte_fotografico", "acta_aceptacion_pdf",
                    "estado", "supervisor_aprobo", "fecha_aprobacion_supervisor"
                ])

                sesion.estado = "aprobado_supervisor"
                sesion.save(update_fields=["estado"])
                sesion.asignaciones.update(estado="aprobado_supervisor")

                messages.success(
                    request, "Proyecto aprobado. Informe y acta generados.")
                return redirect("operaciones:fotos_revisar_sesion", servicio_id=servicio.id)

        elif accion == "rechazar":
            # Rechazo: habilita reintento a técnicos
            sesion.estado = "rechazado_supervisor"
            sesion.save(update_fields=["estado"])
            sesion.asignaciones.update(
                estado="rechazado_supervisor",
                reintento_habilitado=True
            )
            servicio.estado = "rechazado_supervisor"
            servicio.supervisor_rechazo = request.user
            servicio.motivo_rechazo = comentario or "Rechazado por supervisor."
            servicio.save(
                update_fields=["estado", "supervisor_rechazo", "motivo_rechazo"])

            messages.warning(
                request, "Proyecto rechazado. Técnicos habilitados para reintento.")
            return redirect("operaciones:fotos_revisar_sesion", servicio_id=servicio.id)

        messages.error(request, "Acción no válida.")
        return redirect("operaciones:fotos_revisar_sesion", servicio_id=servicio.id)

    # GET — armar contexto
    evidencias_por_asig = []
    for a in asignaciones:
        evs = _order_evidencias(
            a.evidencias.select_related("requisito")
        )
        evidencias_por_asig.append((a, evs))

    return render(request, "operaciones/fotos_revisar_sesion.html", {
        "servicio": servicio,
        "sesion": sesion,
        "evidencias_por_asig": evidencias_por_asig,
        "can_review": can_review,
        "reporte_listo": bool(servicio.reporte_fotografico),
        "reporte_url": servicio.reporte_fotografico.url if servicio.reporte_fotografico else "",
        "acta_url": servicio.acta_aceptacion_pdf.url if servicio.acta_aceptacion_pdf else "",
    })


@login_required
# si no quieres incluir PM, deja solo 'supervisor', 'admin'
@rol_requerido('supervisor', 'admin', 'pm')
def borrar_evidencia_supervisor(request, ev_id: int):
    ev = get_object_or_404(EvidenciaFoto, pk=ev_id)
    servicio = ev.tecnico_sesion.sesion.servicio

    ESTADOS_PERMITIDOS = {'en_revision_supervisor',
                          'en_progreso', 'rechazado_supervisor'}
    if servicio.estado not in ESTADOS_PERMITIDOS:
        messages.warning(
            request, "No se puede eliminar evidencias en el estado actual del proyecto.")
        return redirect('operaciones:fotos_revisar_sesion', servicio_id=servicio.id)

    if request.method == "POST":
        # 🔒 Borrado seguro: elimina el archivo físico solo si no hay otra evidencia que lo referencie
        try:
            if ev.imagen and ev.imagen.name:
                same = EvidenciaFoto.objects.filter(
                    imagen=ev.imagen.name).exclude(pk=ev.pk).exists()
                if not same:
                    ev.imagen.delete(save=False)
        except Exception:
            pass

        ev.delete()
        messages.success(request, "Evidencia eliminada correctamente.")
    else:
        messages.error(request, "Método no permitido.")

    return redirect('operaciones:fotos_revisar_sesion', servicio_id=servicio.id)

# ---- Nombre de archivos ----


def _excel_filename(servicio) -> str:
    from .models import _site_name_for
    sitio = _site_name_for(servicio) or "Sitio"
    idc = (servicio.id_claro or f"DU{servicio.du}" or "Proyecto").replace(
        "/", "-")
    base = f"{idc}_{sitio} - Mantencion Correctiva".strip()
    return f"{base}.xlsx"


def _excel_filename_parcial(servicio) -> str:
    from .models import _site_name_for
    sitio = _site_name_for(servicio) or "Sitio"
    idc = (servicio.id_claro or f"DU{servicio.du}" or "Proyecto").replace(
        "/", "-")
    base = f"{idc}_{sitio} - Mantencion Correctiva (PARCIAL)".strip()
    return f"{base}.xlsx"

# ---- JPEG temporal reescalado ----


def _tmp_jpeg_from_filefield(filefield, max_px=1600, quality=90):
    filefield.open("rb")
    raw = filefield.read()

    im = Image.open(io.BytesIO(raw))
    im = ImageOps.exif_transpose(im)

    if im.mode in ("RGBA", "LA"):
        bg = Image.new("RGB", im.size, (255, 255, 255))
        alpha = im.split()[-1]
        bg.paste(im, mask=alpha)
        im = bg
    elif im.mode == "P":
        im = im.convert("RGBA")
        bg = Image.new("RGB", im.size, (255, 255, 255))
        bg.paste(im, mask=im.split()[-1])
        im = bg
    else:
        im = im.convert("RGB")

    w, h = im.size
    if max(w, h) > max_px:
        scale = max_px / float(max(w, h))
        im = im.resize((int(w * scale), int(h * scale)), Image.LANCZOS)
        w, h = im.size

    tmp = NamedTemporaryFile(delete=False, suffix=".jpg")
    im.save(tmp.name, format="JPEG", quality=quality, optimize=True)
    tmp.close()
    return tmp.name, w, h

# ---- Cargar workbook desde template ----


def _wb_from_template():
    from openpyxl import Workbook, load_workbook
    tpl_path = getattr(settings, "REPORTE_FOTOS_TEMPLATE_XLSX", "")
    if tpl_path and os.path.exists(tpl_path):
        try:
            # Cargamos la plantilla tal cual
            return load_workbook(tpl_path, data_only=False)
        except Exception:
            pass
    # Fallback sin plantilla
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"
    ws["A1"] = "Reporte Fotográfico"
    return wb


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def generar_reporte_parcial_proyecto(request, servicio_id: int):
    servicio = get_object_or_404(ServicioCotizado, pk=servicio_id)
    estados_permitidos = {'en_progreso',
                          'en_revision_supervisor', 'rechazado_supervisor'}
    if servicio.estado not in estados_permitidos:
        messages.warning(
            request,
            "El reporte parcial solo está disponible mientras el proyecto está en proceso o en revisión."
        )
        return redirect('operaciones:fotos_revisar_sesion', servicio_id=servicio.id)

    try:
        xlsx_path = _xlsx_path_reporte_fotografico(servicio)
    except Exception as e:
        messages.error(request, f"No se pudo generar el reporte parcial: {e}")
        return redirect('operaciones:fotos_revisar_sesion', servicio_id=servicio.id)

    filename = _excel_filename_parcial(servicio)
    resp = FileResponse(open(xlsx_path, "rb"),
                        as_attachment=True, filename=filename)
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    return resp

# ============================
# GENERADORES: Excel + Acta PDF
# ============================


def _outline(ws, r1, c1, r2, c2, color="FF0000", style="medium"):
    red = Side(style=style, color=color)
    for c in range(c1, c2 + 1):
        ws.cell(r1, c).border = Border(top=red)
    for c in range(c1, c2 + 1):
        ws.cell(r2, c).border = Border(bottom=red)
    for r in range(r1, r2 + 1):
        ws.cell(r, c1).border = Border(left=red)
    for r in range(r1, r2 + 1):
        ws.cell(r, c2).border = Border(right=red)


def _set_compact_grid(ws, last_col=30, default_col_w=4.5, body_row_h=18):
    for c in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = default_col_w
    for r in range(1, 200):
        if ws.row_dimensions[r].height is None:
            ws.row_dimensions[r].height = body_row_h


# =========================================
# HOJA 2: REPORTE (fotos centradas, 2 por fila)
# =========================================


def _xlsx_path_from_evqs(servicio, ev_qs) -> str:
    """
    Hoja 'Fotografias' (2 por fila, **1 columna por bloque**):
      - Fill height (llena el alto).
      - Ensanchamiento horizontal por REPORT_IMG_WIDEN_X.
      - **Centrado con CANVAS** del tamaño exacto de la caja.
      - Nunca desborda.
    """
    import os
    from tempfile import NamedTemporaryFile

    from django.conf import settings
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Alignment, Border, Font, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.units import pixels_to_EMU
    from PIL import Image
    from PIL import Image as PILImage
    from PIL import ImageOps

    from .models import _site_name_for

    wb = _wb_from_template()
    _prefill_hoja_datos_general(wb, servicio)

    if "REPORTE" in wb.sheetnames:
        try:
            wb.remove(wb["REPORTE"])
        except Exception:
            pass
    ws = wb.create_sheet(title="Fotografias", index=1)
    ws.sheet_view.showGridLines = False
    wb.active = 0

    thick = Side(style="thick", color="000000")
    border_all = Border(left=thick, right=thick, top=thick, bottom=thick)

    # ===== Layout: 1 columna por bloque =====
    BLOCK_COLS, SEP_COLS = 1, 1
    LEFT_COL = 1                                   # A
    RIGHT_COL = LEFT_COL + BLOCK_COLS + SEP_COLS   # C  (A | B sep | C)

    # ---------- Helpers ----------
    def _colwidth_to_px(width_chars: float) -> int:
        return int(((256 * width_chars + 128) // 256) * 7)

    def _pts_to_px(pts: float) -> int:
        return int(round(pts * 96 / 72))

    def _box_size_px(start_col_idx: int, img_top_row: int) -> tuple[int, int]:
        w_px = 0
        for c in range(start_col_idx, start_col_idx + BLOCK_COLS):
            w_chars = ws.column_dimensions[get_column_letter(c)].width
            w_px += _colwidth_to_px(w_chars if w_chars is not None else 8.43)
        h_px = 0
        for r in range(img_top_row, img_top_row + ROWS_IMG):
            h_pts = ws.row_dimensions[r].height
            h_px += _pts_to_px(h_pts if h_pts is not None else 15)
        return w_px, h_px

    def _merge_with_border(r0, c0, r1, c1, text, align="center"):
        ws.merge_cells(start_row=r0, start_column=c0,
                       end_row=r1, end_column=c1)
        cell = ws.cell(row=r0, column=c0, value=text)
        cell.alignment = Alignment(
            horizontal=align, vertical="center", wrap_text=True)
        for rr in range(r0, r1 + 1):
            for cc in range(c0, c1 + 1):
                ws.cell(row=rr, column=cc).border = border_all

    # === Anchos (bloque ~59 chars) ===
    BLOCK_WIDTH_CH = int(getattr(settings, "REPORT_BLOCK_COL_WIDTH_CH", 59))
    SEP_WIDTH_CH = int(getattr(settings, "REPORT_SEP_COL_WIDTH_CH", 2))

    def _set_block_cols(col_start_letter: str):
        idx = ws[col_start_letter][0].column
        for off in range(BLOCK_COLS):
            ws.column_dimensions[get_column_letter(
                idx + off)].width = BLOCK_WIDTH_CH

    _set_block_cols("A")  # bloque izq
    ws.column_dimensions[get_column_letter(
        LEFT_COL + BLOCK_COLS)].width = SEP_WIDTH_CH  # B (sep)
    _set_block_cols(get_column_letter(RIGHT_COL))  # bloque der (C)

    # Altura de la caja
    HEAD_ROWS = 1
    ROWS_IMG = int(getattr(settings, "REPORT_ROWS_IMG", 16))  # antes: 12
    ROW_INFO = 1
    ROW_SPACE = 1
    ROW_HEIGHT_PTS = float(
        getattr(settings, "REPORT_ROW_HEIGHT_PTS", 20.0))  # antes: 18

    def _set_rows(r0: int, count: int, height_pts: float):
        for r in range(r0, r0 + count):
            ws.row_dimensions[r].height = height_pts

    # Título principal
    site_name = _site_name_for(servicio)
    title = f"ID CLARO: {servicio.id_claro or ''} — SITIO: {site_name or ''}"
    ws.merge_cells(start_row=1, start_column=LEFT_COL,
                   end_row=1, end_column=LEFT_COL + (BLOCK_COLS * 2 + SEP_COLS) - 1)
    c_title = ws.cell(row=1, column=LEFT_COL, value=title)
    c_title.alignment = Alignment(horizontal="center", vertical="center")
    c_title.font = Font(bold=True)
    ws.row_dimensions[1].height = 24

    # ---- temporales
    tmp_files_to_delete: list[str] = []

    def _track_tmp(path: str):
        if path:
            tmp_files_to_delete.append(path)

    # ---------- Dibuja un bloque ----------
    def _draw_block(top_row: int, left_col_idx: int, ev) -> None:
        # Encabezado
        if getattr(servicio.sesion_fotos, "proyecto_especial", False) and ev.requisito_id is None:
            titulo_req = (ev.titulo_manual or "").strip() or "Extra"
        else:
            titulo_req = ((getattr(ev.requisito, "titulo", "")
                          or "").strip() or "Extra")

        ws.merge_cells(start_row=top_row, start_column=left_col_idx,
                       end_row=top_row, end_column=left_col_idx + BLOCK_COLS - 1)
        c_head = ws.cell(row=top_row, column=left_col_idx, value=titulo_req)
        c_head.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True)
        c_head.font = Font(bold=True)
        _set_rows(top_row, 1, 20)
        for cc in range(left_col_idx, left_col_idx + BLOCK_COLS):
            ws.cell(row=top_row, column=cc).border = border_all

        # Caja imagen
        img_top = top_row + HEAD_ROWS
        img_bottom = img_top + ROWS_IMG - 1
        ws.merge_cells(start_row=img_top, start_column=left_col_idx,
                       end_row=img_bottom, end_column=left_col_idx + BLOCK_COLS - 1)
        _set_rows(img_top, ROWS_IMG, ROW_HEIGHT_PTS)
        for rr in range(img_top, img_bottom + 1):
            for cc in range(left_col_idx, left_col_idx + BLOCK_COLS):
                ws.cell(row=rr, column=cc).border = border_all

        # Medidas de la caja
        box_w_px, box_h_px = _box_size_px(left_col_idx, img_top)

        # --- Fudge para evitar solape con la nota ---
        fudge_px = int(
            getattr(settings, "REPORT_IMG_BOX_FUDGE_PX", 14))  # ↑ antes 3
        box_h_eff = max(1, box_h_px - fudge_px)

        # Pads y ensanchamiento
        side_pad = int(getattr(settings, "REPORT_IMG_SIDE_PAD_PX", 9))
        top_pad = int(getattr(settings, "REPORT_IMG_TOP_PAD_PX", 0))
        widen_x = float(getattr(settings, "REPORT_IMG_WIDEN_X", 1.30))
        widen_x = max(1.0, min(widen_x, 1.40))

        tmp_img_path = ""
        try:
            # Base JPG con mayor fuente y menos compresión
            tmp_img_path, w, h = _tmp_jpeg_from_filefield(
                ev.imagen,
                max_px=int(getattr(settings, "REPORT_IMG_MAX_PX", 2400)),
                quality=int(
                    getattr(settings, "REPORT_IMG_JPG_QUALITY_SRC", 93)),
            )
            _track_tmp(tmp_img_path)

            # Fill height + widen, con tope de ancho (para que nunca desborde)
            target_h = max(1, box_h_eff - 2 * top_pad)
            scale_h = target_h / float(h)
            base_w = max(1, int(round(w * scale_h)))
            target_w = max(1, int(round(base_w * widen_x)))
            max_w = max(1, box_w_px - 2 * side_pad)
            if target_w > max_w:
                target_w = max_w

            # === CANVAS del tamaño exacto de la caja efectiva ===
            canvas = Image.new("RGB", (box_w_px, box_h_eff), (255, 255, 255))

            with PILImage.open(tmp_img_path) as im:
                im = ImageOps.exif_transpose(im).convert("RGB")
                im = im.resize((target_w, target_h), PILImage.LANCZOS)

                # Centramos considerando pads
                off_x = side_pad + (box_w_px - 2 * side_pad - target_w) // 2
                off_y = top_pad + (box_h_eff - 2 * top_pad - target_h) // 2
                off_x = max(0, off_x)
                off_y = max(0, off_y)

                canvas.paste(im, (off_x, off_y))

            # Guardar canvas como JPG con mejor calidad y sin submuestreo
            tmp_canvas = NamedTemporaryFile(delete=False, suffix=".jpg")
            canvas.save(
                tmp_canvas.name,
                format="JPEG",
                quality=int(getattr(settings, "REPORT_IMG_JPG_QUALITY", 93)),
                subsampling=0,       # 4:4:4
                optimize=True,
                progressive=True,
            )
            tmp_canvas.close()
            canvas_path = tmp_canvas.name
            _track_tmp(canvas_path)

            xl_img = XLImage(canvas_path)

            # Anclar al inicio de la caja y forzar tamaño exacto de la caja efectiva
            try:
                from openpyxl.drawing.spreadsheet_drawing import (
                    AnchorMarker, Extent, OneCellAnchor)
                col0 = left_col_idx - 1
                row0 = img_top - 1
                anchor = AnchorMarker(col=col0, row=row0, colOff=0, rowOff=0)
                # -1 px extra en altura para vencer redondeo de Excel
                extent = Extent(cx=pixels_to_EMU(box_w_px),
                                cy=pixels_to_EMU(max(1, box_h_eff - 1)))
                one = OneCellAnchor(_from=anchor, ext=extent, editAs="oneCell")
                xl_img.anchor = one
                ws.add_image(xl_img)
            except Exception:
                xl_img.width = box_w_px
                xl_img.height = max(1, box_h_eff - 1)
                ws.add_image(
                    xl_img, f"{get_column_letter(left_col_idx)}{img_top}")

        except Exception:
            pass

        # Nota
        info_row = img_bottom + 1
        _set_rows(info_row, 1, 30)
        nota_txt = (ev.nota or "").strip()
        _merge_with_border(info_row, left_col_idx, info_row,
                           left_col_idx + BLOCK_COLS - 1, nota_txt, align="center")

    # Iteración
    cur_row = 3
    ev_iter = ev_qs.iterator() if callable(
        getattr(ev_qs, "iterator", None)) else ev_qs

    idx = 0
    try:
        for ev in ev_iter:
            if idx % 2 == 0:
                _draw_block(cur_row, LEFT_COL, ev)
            else:
                _draw_block(cur_row, RIGHT_COL, ev)
                cur_row += (HEAD_ROWS + ROWS_IMG + ROW_INFO + ROW_SPACE)
            idx += 1

        if idx % 2 == 1:
            cur_row += (HEAD_ROWS + ROWS_IMG + ROW_INFO + ROW_SPACE)

        tmp_xlsx = NamedTemporaryFile(delete=False, suffix=".xlsx")
        tmp_xlsx.close()
        wb.save(tmp_xlsx.name)
        return tmp_xlsx.name
    finally:
        for p in tmp_files_to_delete:
            try:
                os.unlink(p)
            except Exception:
                pass


def _prefill_hoja_datos_general(wb: Workbook, servicio: ServicioCotizado) -> None:
    """
    Completa la hoja 'Datos General' del template con:
      - CRQ/ID          -> ID Claro
      - CODIGO SITES    -> ID NEW
      - ALIAS           -> Nombre del sitio
      - DIRECCION       -> Dirección
      - COMUNA          -> Comuna
      - NOMBRE EJECUTANTE -> 'Grupo GZ Services'
      - FECHA MANTENCION -> fecha en que el técnico finaliza (máximo finalizado_en)
    El template usa celdas/merges:
      AB2:AH3 -> valor de CRQ/ID (escribe en AB2)
      D7:G7   -> CODIGO SITES
      J7:U7   -> ALIAS
      D8:U8   -> DIRECCION
      Z8:AH8  -> COMUNA
      D9:U9   -> NOMBRE EJECUTANTE
      Z9:AH9  -> FECHA MANTENCION (dd-mm-YYYY)
    """
    if "Datos General" not in wb.sheetnames:
        return

    ws = wb["Datos General"]

    # --- Datos base del sitio ---
    id_claro = (servicio.id_claro or "").strip()
    id_new = (servicio.id_new or "").strip()

    sm = None
    if id_claro:
        sm = SitioMovil.objects.filter(id_claro=id_claro).first()

    alias = (sm.nombre or _site_name_for(servicio) or "").strip()
    direccion = (sm.direccion or "").strip()
    comuna = (sm.comuna or "").strip()

    # Fecha mantención: última finalización de las asignaciones del servicio
    fecha_mant = None
    try:
        sesion = servicio.sesion_fotos
        agg = sesion.asignaciones.aggregate(ult=Max("finalizado_en"))
        fecha_mant = agg["ult"]
    except Exception:
        pass
    if not fecha_mant:
        # Fallback: ahora (o puedes usar servicio.fecha_aprobacion_supervisor, etc.)
        fecha_mant = timezone.localtime()
    fecha_mant_txt = fecha_mant.strftime("%d-%m-%Y")

    # --- Escribir en el template (si hay merges, se escribe en la celda superior-izquierda) ---
    # CRQ/ID
    ws["AB2"].value = id_claro

    # Fila de “Datos general del sitio”
    ws["D7"].value = id_new
    ws["J7"].value = alias
    ws["D8"].value = direccion
    ws["Z8"].value = comuna
    ws["D9"].value = "Grupo GZ Services"
    ws["Z9"].value = fecha_mant_txt


def _xlsx_path_reporte_fotografico(servicio) -> str:
    from openpyxl import Workbook
    wb = _wb_from_template()
    _prefill_hoja_datos_general(wb, servicio)
    ev_qs = _order_evidencias(
        EvidenciaFoto.objects
        .filter(tecnico_sesion__sesion=servicio.sesion_fotos)
        .select_related("requisito")
    )
    return _xlsx_path_from_evqs(servicio, ev_qs)


def _colwidth_to_px(width_chars: float) -> int:
    return int(((256 * width_chars + 128) // 256) * 7)


def _pts_to_px(pts: float) -> int:
    return int(round(pts * 96 / 72))


def _sheet_xy_px(ws, col_idx_1based: int, row_idx_1based: int) -> tuple[int, int]:
    x = 0
    for c in range(1, col_idx_1based):
        w = ws.column_dimensions[get_column_letter(c)].width
        x += _colwidth_to_px(w if w is not None else 8.43)
    y = 0
    for r in range(1, row_idx_1based):
        h = ws.row_dimensions[r].height
        y += _pts_to_px(h if h is not None else 15)
    return x, y


def _col_width_px(ws, col_idx_1based: int) -> int:
    w_chars = ws.column_dimensions[get_column_letter(col_idx_1based)].width
    return _colwidth_to_px(w_chars if w_chars is not None else 8.43)


def _row_height_px(ws, row_idx_1based: int) -> int:
    h_pts = ws.row_dimensions[row_idx_1based].height
    return _pts_to_px(h_pts if h_pts is not None else 15)


def _px_to_anchor(ws, start_col_1b: int, start_row_1b: int, off_x_px: int, off_y_px: int):
    col = start_col_1b
    x_left = 0
    while True:
        w = _col_width_px(ws, col)
        if off_x_px < x_left + w:
            col_idx0 = col - 1
            colOffEMU = pixels_to_EMU(off_x_px - x_left)
            break
        x_left += w
        col += 1

    row = start_row_1b
    y_top = 0
    while True:
        h = _row_height_px(ws, row)
        if off_y_px < y_top + h:
            row_idx0 = row - 1
            rowOffEMU = pixels_to_EMU(off_y_px - y_top)
            break
        y_top += h
        row += 1

    return col_idx0, row_idx0, colOffEMU, rowOffEMU


def _absolute_anchor_compat(x_px, y_px, w_px, h_px):
    from openpyxl.utils.units import pixels_to_EMU
    try:
        from openpyxl.drawing.geometry import Point2D
        from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor, Extent
        pos = Point2D(pixels_to_EMU(x_px), pixels_to_EMU(y_px))
        ext = Extent(cx=pixels_to_EMU(w_px), cy=pixels_to_EMU(h_px))
        return AbsoluteAnchor(pos=pos, ext=ext)
    except Exception:
        from openpyxl.drawing.spreadsheet_drawing import AbsoluteAnchor
        try:
            from openpyxl.drawing.geometry import XDRPoint2D, XDRPositiveSize2D
        except Exception:
            return None
        pos = XDRPoint2D(pixels_to_EMU(x_px), pixels_to_EMU(y_px))
        ext = XDRPositiveSize2D(pixels_to_EMU(w_px), pixels_to_EMU(h_px))
        return AbsoluteAnchor(pos=pos, ext=ext)


def _onecell_anchor_compat(col_idx0, row_idx0, off_x_px, off_y_px, w_px, h_px):
    from openpyxl.utils.units import pixels_to_EMU
    try:
        from openpyxl.drawing.spreadsheet_drawing import (AnchorMarker, Extent,
                                                          OneCellAnchor)
        anchor = AnchorMarker(
            col=col_idx0, row=row_idx0,
            colOff=pixels_to_EMU(off_x_px), rowOff=pixels_to_EMU(off_y_px)
        )
        extent = Extent(cx=pixels_to_EMU(w_px), cy=pixels_to_EMU(h_px))
        return OneCellAnchor(_from=anchor, ext=extent, editAs="oneCell")
    except Exception:
        from openpyxl.drawing.spreadsheet_drawing import (AnchorMarker,
                                                          OneCellAnchor)
        try:
            from openpyxl.drawing.geometry import XDRPositiveSize2D
        except Exception:
            return None
        anchor = AnchorMarker(
            col=col_idx0, row=row_idx0,
            colOff=pixels_to_EMU(off_x_px), rowOff=pixels_to_EMU(off_y_px)
        )
        extent = XDRPositiveSize2D(pixels_to_EMU(w_px), pixels_to_EMU(h_px))
        return OneCellAnchor(_from=anchor, ext=extent, editAs="oneCell")


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def generar_acta_preview(request, servicio_id: int):
    """
    Genera el ACTA en PDF.

    Modos:
      - Preview (por defecto): muestra inline sin guardar.
      - Guardar/Reemplazar (si ?save=1 o si hay OC y el servicio está aprobado):
        guarda el PDF nuevo en `acta_aceptacion_pdf`, borrando el archivo previo si existía.

    Notas:
      - Si hay OC asociada, se fuerza la regeneración para que el PDF incluya sus datos.
      - IMPORTANTE: no agregar query params a URLs prefirmadas (S3/Wasabi), rompe la firma.
    """
    import io

    from django.contrib import messages
    from django.core.files.base import ContentFile
    from django.http import FileResponse
    from django.shortcuts import get_object_or_404, redirect

    servicio = get_object_or_404(ServicioCotizado, pk=servicio_id)

    # Estados permitidos para ver/generar el acta (preview/guardar)
    estados_permitidos = {
        'en_progreso',
        'en_revision_supervisor',
        'rechazado_supervisor',
        'aprobado_supervisor',
        'aprobado_pm',
    }
    if servicio.estado not in estados_permitidos:
        messages.error(request, "El acta no está disponible para este estado.")
        return redirect('operaciones:fotos_revisar_sesion', servicio_id=servicio.id)

    # Flags de query
    save_flag = str(request.GET.get("save", "")).lower() in {"1", "true", "on", "yes", "si", "sí"}
    force = request.GET.get("force")

    # Detectar si existe OC (du puede ser FK o texto)
    oc_exists = False
    try:
        from facturacion.models import OrdenCompraFacturacion
        try:
            # Caso 1: du es FK a ServicioCotizado
            oc_exists = OrdenCompraFacturacion.objects.filter(du=servicio).exists()
        except Exception:
            # Caso 2: du es texto/código (ej. "DU1234")
            du_val = getattr(servicio, 'du', None)
            if du_val:
                oc_exists = OrdenCompraFacturacion.objects.filter(du=du_val).exists()
    except Exception:
        pass

    # Si vamos a guardar, forzamos regeneración
    if save_flag:
        force = "1"

    # Si no se pidió force y existe OC, forzamos regeneración para incluirla
    if not force and oc_exists:
        force = "1"

    # Auto-guardar cuando hay OC y el servicio ya está aprobado
    if oc_exists and servicio.estado in {"aprobado_supervisor", "aprobado_pm"}:
        save_flag = True
        force = "1"

    # Si hay PDF previo y NO se fuerza y NO vamos a guardar, servimos el ya existente
    if getattr(servicio.acta_aceptacion_pdf, "name", "") and not force and not save_flag:
        return redirect(servicio.acta_aceptacion_pdf.url)

    # Generar bytes del acta (regeneramos en force/save_flag)
    try:
        pdf_bytes = _bytes_acta_aceptacion(servicio)
    except Exception as e:
        messages.error(request, f"No se pudo generar el acta: {e}")
        return redirect('operaciones:fotos_revisar_sesion', servicio_id=servicio.id)

    # Determinar nombre de archivo preferentemente con la OC
    from .models import _pdf_filename
    doc_compra_for_name = getattr(servicio, "documento_compra", None)
    if not doc_compra_for_name:
        try:
            from facturacion.models import OrdenCompraFacturacion
            oc = None
            try:
                oc = (OrdenCompraFacturacion.objects
                      .filter(du=servicio).order_by("-creado").first())
            except Exception:
                du_val = getattr(servicio, 'du', None)
                if du_val:
                    oc = (OrdenCompraFacturacion.objects
                          .filter(du=du_val).order_by("-creado").first())
            if oc and getattr(oc, "orden_compra", None):
                doc_compra_for_name = oc.orden_compra
        except Exception:
            pass
    filename = _pdf_filename(servicio, doc_compra_for_name or "DOC")

    # Guardar (reemplazar) o mostrar inline
    if save_flag:
        old_name = getattr(servicio.acta_aceptacion_pdf, "name", "") or None

        # Borramos el archivo anterior si existe (por si el storage no sobreescribe)
        if old_name:
            try:
                servicio.acta_aceptacion_pdf.storage.delete(old_name)
            except Exception:
                pass

        # Guardar bytes en el FileField (el storage decide si sobreescribe o versiona)
        servicio.acta_aceptacion_pdf.save(filename, ContentFile(pdf_bytes), save=True)

        messages.success(request, "Acta actualizada y guardada correctamente.")
        # ⚠️ No agregar parámetros a la URL prefirmada
        return redirect(servicio.acta_aceptacion_pdf.url)

    # Preview inline (sin guardar en el modelo)
    resp = FileResponse(io.BytesIO(pdf_bytes), content_type="application/pdf")
    resp["Content-Disposition"] = f'inline; filename="{filename}"'
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    return resp

def _bytes_acta_aceptacion(servicio) -> bytes:
    """
    Genera el ACTA PDF:
    - Título subrayado
    - 'Grupo GZS Services' subrayado en el texto inicial
    - Tabla compacta con wrap en Conformidad
    - Firmas alineadas y firma de Edgardo embebida
    """
    import io
    import os

    from django.conf import settings
    from django.utils import timezone
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer,
                                    Table, TableStyle)

    from .models import _site_name_for

    # -------- Datos base --------
    fecha_txt = timezone.localtime().strftime("%d-%m-%Y")
    id_claro = servicio.id_claro or ""
    nombre_sitio = _site_name_for(servicio) or ""
    detalle_serv = "Mantencion Correctiva"

    try:
        from facturacion.models import OrdenCompraFacturacion
        oc = (OrdenCompraFacturacion.objects
              .filter(du=servicio).order_by("-creado").first())
    except Exception:
        oc = None

    doc_compra = (getattr(servicio, "documento_compra", None)
                  or (oc.orden_compra if oc else "")
                  or "—")
    pos_oc = oc.pos if oc and oc.pos else "0"
    precio_unitario = float(
        oc.precio_unitario) if oc and oc.precio_unitario is not None else 0.0
    monto = float(oc.monto) if oc and oc.monto is not None else 0.0
    monto_pendiente = 0.0

    def fmt_uf(x: float) -> str:
        return f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    # -------- Estilos --------
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "title", parent=styles["Heading2"], alignment=1, spaceAfter=12)
    normal = ParagraphStyle(
        "normal", parent=styles["Normal"], fontSize=10, leading=13, spaceAfter=4)
    normal_j = ParagraphStyle("normal_j", parent=normal, alignment=TA_JUSTIFY)
    table_head = ParagraphStyle("table_head", parent=styles["Normal"], fontSize=8, leading=10,
                                textColor=colors.white, alignment=1)
    table_cell = ParagraphStyle(
        "table_cell", parent=styles["Normal"], fontSize=8, leading=10, alignment=1)
    table_cell_left = ParagraphStyle(
        "table_cell_left", parent=table_cell, alignment=TA_LEFT)

    # -------- Documento --------
    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm
    )
    frame_width = A4[0] - doc.leftMargin - doc.rightMargin
    elems = []

    # Logo
    logo_path = getattr(settings, "ACTA_SITES_LOGO_PATH", "")
    if not (logo_path and os.path.exists(logo_path)):
        logo_path = os.path.join(
            settings.BASE_DIR, "static", "images", "sites_logo.png")
    if os.path.exists(logo_path):
        img = RLImage(logo_path, width=2.8*cm, height=2.8*cm)
        t_logo = Table([[img]], colWidths=[frame_width], rowHeights=[3.2*cm])
        t_logo.setStyle(TableStyle(
            [("ALIGN", (0, 0), (-1, -1), "CENTER"), ("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
        elems += [t_logo, Spacer(1, 4)]

    # Título (subrayado)
    elems.append(Paragraph(
        "<u><b>ACTA DE ACEPTACION DE SERVICIOS Y/O MATERIALES</b></u>", title))

    # Intro (subrayando solo “Grupo GZS Services”)
    intro = (
        f"En Santiago, con fecha <b>{fecha_txt}</b> la empresa <u><b>Grupo GZS Services</b></u> "
        f"representada por el señor Edgardo Zapata, hace entrega a <b>SITES CHILE S.A.</b>, "
        f"representada por el señor Jean Paul Jofre Diaz, otorga la recepción de los siguientes servicios:"
    )
    elems += [Paragraph(intro, normal_j), Spacer(1, 6)]

    # Bloque label : valor
    def row(lbl, val): return [Paragraph(
        lbl, table_cell_left), Paragraph(f":  {val}", table_cell_left)]
    info = [
        row("Documento de Compras", doc_compra),
        row("Monto OC (UF NETO)", fmt_uf(precio_unitario)),
        row("Detalle de los Servicios", detalle_serv),
        row("Nombre del sitio", nombre_sitio),
        row("ID Sitio", id_claro),
    ]
    t_info = Table(info, colWidths=[5.0*cm, frame_width-5.0*cm])
    t_info.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elems += [t_info, Spacer(1, 8)]

    # Tabla principal (compacta)
    headers = ["Req.", "Id Sitio", "Monto de OC", "Monto A Pagar",
               "Monto pendiente", "POS OC", "Conformidad"]
    conformidad_txt = f"{id_claro}_{nombre_sitio} - {detalle_serv}"
    data = [
        [Paragraph(h, table_head) for h in headers],
        [
            Paragraph("MTTO", table_cell),
            Paragraph(id_claro or "—", table_cell),
            Paragraph(f"{fmt_uf(precio_unitario)} UF", table_cell),
            Paragraph(f"{fmt_uf(monto)} UF", table_cell),
            Paragraph(f"{fmt_uf(monto_pendiente)} UF", table_cell),
            Paragraph(str(pos_oc), table_cell),
            Paragraph(conformidad_txt, table_cell_left),
        ],
    ]
    col_widths = [16*mm, 22*mm, 26*mm, 26*mm, 26*mm, 14*mm, 30*mm]
    tabla = Table(data, colWidths=col_widths, hAlign="LEFT")

    LIGHT_TEAL = colors.HexColor("#6CB3C4")
    DARK_TEAL = colors.HexColor("#2F6C86")
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (2, 0), LIGHT_TEAL),
        ("BACKGROUND", (3, 0), (6, 0), DARK_TEAL),
        ("TEXTCOLOR", (0, 0), (6, 0), colors.white),
        ("FONTNAME", (0, 0), (6, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (6, 0), 8),
        ("ALIGN", (0, 0), (6, 0), "CENTER"),
        ("VALIGN", (0, 0), (6, 0), "MIDDLE"),
        ("FONTSIZE", (0, 1), (6, 1), 8),
        ("ALIGN", (0, 1), (5, 1), "CENTER"),
        ("ALIGN", (6, 1), (6, 1), "LEFT"),
        ("VALIGN", (0, 1), (6, 1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("BOX", (0, 0), (-1, -1), 0.8, colors.black),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elems += [tabla, Spacer(1, 8)]

    # Texto legal
    legal = (
        "Sin perjuicio de lo anterior, la presente recepción de trabajos podrá ser impugnada "
        "en cualquier tiempo por SITES, si la información aportada fuere falsa, inexacta y/o "
        "incompleta, quedando desde ya SITES facultado para reclamar cualquier perjuicio que "
        "se origine y a su vez, solicitar la restitución de los montos que hubieren sido pagados "
        "con ocasión de los trabajos encomendados."
    )
    elems += [Paragraph(legal, normal_j), Spacer(1, 22)]

    # Firmas
    firma_path = getattr(settings, "ACTA_FIRMA_EDGARDO_PATH", "")
    if not (firma_path and os.path.exists(firma_path)):
        firma_path = os.path.join(
            settings.BASE_DIR, "static", "images", "edgardo_zapata.png")
    firma_img = RLImage(firma_path, width=4.0*cm, height=2.0 *
                        cm) if os.path.exists(firma_path) else ""
    firma_txt_left = Paragraph(
        "NOMBRE Y FIRMA<br/>SITES CHILE S.A.", table_cell)
    firma_txt_right = Paragraph("EDGARDO ZAPATA<br/>CONTRATISTA", table_cell)
    firmas_tbl = Table([["", firma_img], [firma_txt_left, firma_txt_right]],
                       colWidths=[frame_width/2.0, frame_width/2.0],
                       rowHeights=[2.2*cm, None], hAlign="CENTER")
    firmas_tbl.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elems.append(firmas_tbl)

    doc.build(elems)
    bio.seek(0)
    return bio.getvalue()
