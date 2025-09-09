# operaciones/services/reportes.py
from io import BytesIO
from decimal import Decimal
from pathlib import Path

from django.utils import timezone
from django.conf import settings

from openpyxl import Workbook, load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm

from ..models import (
    ServicioCotizado, SitioMovil,
    _site_name_for, _excel_filename, _pdf_filename,
    upload_to_reporte, upload_to_acta
)
from ..storage_backends import GZWasabiStorage

_gz = GZWasabiStorage()


def _get_doc_compra(servicio: ServicioCotizado) -> str:
    return (getattr(servicio, "documento_compra", None) or servicio.id_new or "N/A")


def _get_monto_uf(servicio: ServicioCotizado) -> Decimal:
    if getattr(servicio, "monto_uf", None) is not None:
        return servicio.monto_uf
    return Decimal("0.00")


def _get_template_path() -> Path | None:
    # Busca plantilla en operaciones/static/reportes/plantilla_reporte.xlsx
    base = Path(settings.BASE_DIR)
    candidate = base / "operaciones" / "static" / \
        "reportes" / "plantilla_reporte.xlsx"
    return candidate if candidate.exists() else None


def generar_excel(servicio: ServicioCotizado) -> str:
    """
    Genera el Excel (usando plantilla si existe) y lo sube a Wasabi.
    Nombre: "<id_claro> - <nombre_sitio>.xlsx"
    """
    site_name = _site_name_for(servicio)
    fecha = timezone.localdate().strftime('%d-%m-%Y')

    tpl = _get_template_path()
    if tpl:
        wb = load_workbook(tpl)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen"

    # Relleno mínimo común (ajusta celdas según tu plantilla real)
    try:
        ws["A1"] = "Reporte Fotográfico"
        ws["A2"] = f"ID Claro: {servicio.id_claro or ''}"
        ws["A3"] = f"Sitio: {site_name}"
        ws["A4"] = f"DU: {servicio.du or ''}"
        ws["A5"] = f"Detalle: {(servicio.detalle_tarea or '')[:200]}"
        ws["A6"] = f"Fecha: {fecha}"
    except Exception:
        # Si tu plantilla usa nombres de rango, aquí podrías buscarlos
        pass

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    key = upload_to_reporte(servicio, _excel_filename(servicio))
    _gz.save(key, bio)
    return key


def _wrap_text(text, max_chars=100):
    words, line, out = text.split(), [], []
    for w in words:
        if sum(len(x) for x in line) + len(line) + len(w) <= max_chars:
            line.append(w)
        else:
            out.append(" ".join(line))
            line = [w]
    if line:
        out.append(" ".join(line))
    return out


def generar_acta_pdf(servicio: ServicioCotizado) -> str:
    """
    Genera el PDF 'Acta de Aceptacion (...)_(...) Grupo GZS Services.pdf' y lo sube a Wasabi.
    """
    doc_compra = _get_doc_compra(servicio)
    monto_uf = _get_monto_uf(servicio)
    site_name = _site_name_for(servicio)
    fecha_str = timezone.localdate().strftime('%d-%m-%Y')

    key = upload_to_acta(servicio, doc_compra)
    buffer = BytesIO()

    c = canvas.Canvas(buffer, pagesize=A4)
    w, h = A4

    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(
        w/2, h - 2*cm, "ACTA DE ACEPTACION DE SERVICIOS Y/O MATERIALES")

    c.setFont("Helvetica", 11)
    y = h - 3.2*cm
    c.drawString(
        2*cm, y, f"En Santiago, con fecha {fecha_str} la empresa Grupo GZS Services representada por el")
    y -= 0.6*cm
    c.drawString(
        2*cm, y, "señor Edgardo Zapata, hace entrega a SITES CHILE S.A., representada por el señor Jean")
    y -= 0.6*cm
    c.drawString(
        2*cm, y, "Paul Jofre Diaz, otorga la recepción de los siguientes servicios:")

    y -= 1.0*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, "Documento de Compras   :")
    c.setFont("Helvetica", 11)
    c.drawString(8.6*cm, y, f"{doc_compra}")

    y -= 0.6*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, "Monto OC (UF NETO)     :")
    c.setFont("Helvetica", 11)
    c.drawString(8.6*cm, y, f"{monto_uf}")

    y -= 0.6*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, "Detalle de los Servicios:")
    c.setFont("Helvetica", 11)
    c.drawString(8.6*cm, y, f"{(servicio.detalle_tarea or '')[:60]}")

    y -= 0.6*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, "Nombre del sitio       :")
    c.setFont("Helvetica", 11)
    c.drawString(8.6*cm, y, f"{site_name}")

    y -= 0.6*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, "ID Sitio               :")
    c.setFont("Helvetica", 11)
    c.drawString(8.6*cm, y, f"{servicio.id_claro or ''}")

    y -= 1.2*cm
    c.setFont("Helvetica", 9)
    texto = (
        "Sin perjuicio de lo anterior, la presente recepción de trabajos podrá ser impugnada en cualquier "
        "tiempo por SITES si la información aportada fuere falsa, inexacta y/o incompleta, quedando desde ya "
        "SITES facultado para reclamar cualquier perjuicio que se origine y a su vez, solicitar la restitución "
        "de los montos que hubieren sido pagados con ocasión de los trabajos encomendados."
    )
    for line in _wrap_text(texto, max_chars=110):
        c.drawString(2*cm, y, line)
        y -= 0.5*cm

    # Zona de firma simple
    y -= 1.2*cm
    c.setFont("Helvetica-Bold", 11)
    c.drawString(2*cm, y, "NOMBRE Y FIRMA")
    y -= 1.4*cm
    c.line(2*cm, y, 9*cm, y)
    c.setFont("Helvetica", 11)
    c.drawString(2*cm, y - 0.5*cm, "SITES CHILE S.A.")

    c.showPage()
    c.save()
    buffer.seek(0)

    _gz.save(key, buffer)
    return key


def generar_reporte_y_acta(servicio_id: int) -> tuple[str, str]:
    servicio = ServicioCotizado.objects.get(pk=servicio_id)
    key_xlsx = generar_excel(servicio)
    key_pdf = generar_acta_pdf(servicio)
    return key_xlsx, key_pdf
