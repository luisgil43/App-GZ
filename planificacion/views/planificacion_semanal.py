from io import BytesIO

from django.conf import settings
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist
from django.db import transaction
from django.db.models import Q
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from planificacion.forms.planificacion_semanal import CrearBatchSemanalForm
from planificacion.modelos import (SalidaPlanificacionDiaria,
                                   SitioSalidaPlanificacionDiaria)
from planificacion.models import (BatchPlanificacionSemanal,
                                  PlanificacionMensual, SitioBatchSemanal,
                                  SitioPlanificado)
from planificacion.services.planificacion_diaria import \
    obtener_estado_operacional_sitio
from planificacion.services.planificacion_semanal import (
    actualizar_permiso_desde_batch, agregar_sitios_al_batch,
    cerrar_propuesta_batch, confirmar_sitios_para_planificacion,
    crear_batch_semanal, marcar_gestion_permisos_enviada,
    obtener_candidatos_batch, obtener_resumen_batch,
    obtener_resumen_planificacion_mensual, quitar_sitio_del_batch)
from usuarios.decoradores import rol_requerido

# ============================================================
# ROLES
# ============================================================

ROLES_PLANIFICACION = [
    "admin",
    "pm",
    "supervisor",
]


# ============================================================
# CONFIGURACIÓN DEL OBJETIVO SEMANAL
# ============================================================


def _configurar_objetivo_formulario(
    *,
    form,
    disponibles,
    es_post=False,
):
    """
    Configura el objetivo operacional de una semana.

    IMPORTANTE
    ==========================================================

    `objetivo_sitios` representa la cantidad TOTAL de sitios
    que operacionalmente queremos trabajar durante la semana.

    NO representa la cantidad de sitios que quedan disponibles
    dentro de la planificación mensual actual.

    Ejemplo:

        objetivo semanal:      40
        disponibles agosto:    15

    El batch puede crearse con objetivo 40.

    Los 15 disponibles pertenecientes a agosto podrán formar
    parte inicialmente del batch y posteriormente el déficit
    podrá completarse utilizando sitios provenientes de otra
    planificación mensual, por ejemplo septiembre.
    """

    disponibles = max(
        int(disponibles or 0),
        0,
    )

    campo = form.fields["objetivo_sitios"]

    # ========================================================
    # EL OBJETIVO NO ESTÁ LIMITADO POR EL INVENTARIO DEL MES
    # ========================================================

    campo.widget.attrs.pop(
        "max",
        None,
    )

    # Conservamos el dato únicamente para información visual
    # o futuras validaciones de interfaz.
    campo.widget.attrs["data-disponibles-mes"] = str(disponibles)

    # Eliminamos el atributo antiguo por seguridad si quedó
    # definido desde una carga anterior.
    campo.widget.attrs.pop(
        "data-max-nuevo-batch",
        None,
    )

    # ========================================================
    # OBJETIVO INICIAL
    # ========================================================

    if not es_post:

        campo.initial = 40


# ============================================================
# UTILIDADES EXCEL
# ============================================================


def _valor_sitio(
    sitio,
    nombre_campo,
    default="",
):
    """
    Obtiene un campo del sitio sin romper el Excel si una
    instalación antigua todavía no posee alguno de los campos.
    """

    valor = getattr(
        sitio,
        nombre_campo,
        default,
    )

    if valor is None:
        return default

    return valor


def _ajustar_ancho_columnas_excel(
    worksheet,
):
    for columna in worksheet.columns:

        maximo = 0

        letra = get_column_letter(columna[0].column)

        for celda in columna:

            valor = celda.value

            if valor is None:
                continue

            maximo = max(
                maximo,
                len(str(valor)),
            )

        worksheet.column_dimensions[letra].width = min(
            max(
                maximo + 2,
                12,
            ),
            45,
        )


# ============================================================
# LISTA DE BATCHES SEMANALES DEL MES
# ============================================================


@rol_requerido(*ROLES_PLANIFICACION)
def lista_planificacion_semanal(
    request,
    mensual_id,
):
    """
    Lista todas las semanas vinculadas al mes.

    Incluye tanto:

        planificaciones_origen

    como compatibilidad con batches legacy.
    """

    mensual = get_object_or_404(
        PlanificacionMensual,
        pk=mensual_id,
    )

    batches = (
        BatchPlanificacionSemanal.objects.filter(
            Q(
                planificaciones_origen=mensual,
            )
            | Q(
                planificacion=mensual,
            )
        )
        .distinct()
        .select_related(
            "configuracion_semana",
            "planificacion",
        )
        .prefetch_related(
            "planificaciones_origen",
            "sitios",
            (
                "configuracion_semana__"
                "disponibilidades_cuadrillas__"
                "cuadrilla_operativa"
            ),
        )
        .order_by(
            "fecha_inicio",
            "id",
        )
    )

    filas = []

    for batch in batches:

        filas.append(
            {
                "batch": batch,
                "resumen": obtener_resumen_batch(batch),
            }
        )

    resumen_mensual = obtener_resumen_planificacion_mensual(mensual)

    return render(
        request,
        "planificacion/semanal/lista.html",
        {
            "mensual": mensual,
            "filas": filas,
            "resumen_mensual": resumen_mensual,
        },
    )


# ============================================================
# ELIMINAR BATCH
# ============================================================


@require_POST
@rol_requerido(
    "admin",
)
@transaction.atomic
def eliminar_planificacion_semanal(
    request,
    batch_id,
):
    """
    Elimina administrativamente un batch semanal.

    REGLA GENERAL
    ==========================================================

    Esta acción solamente puede ser ejecutada por ADMIN.

    Elimina la estructura de PLANIFICACIÓN asociada a la
    semana:

    - prioridades diarias asociadas al batch;
    - participaciones de sitios en planificación diaria;
    - salidas de planificación diaria;
    - SitioBatchSemanal;
    - BatchPlanificacionSemanal;
    - ConfiguracionSemana, si queda huérfana.

    IMPORTANTE
    ==========================================================

    Esta acción NO elimina:

    - SitioPlanificado;
    - SitioMovil;
    - ServicioCotizado;
    - técnicos;
    - sesiones;
    - evidencias;
    - información de Operaciones.

    Antes de eliminar la planificación se consulta el estado
    operacional REAL de cada sitio.

    Después de eliminar el batch:

    - finalizado en Operaciones:
        SitioPlanificado -> completado

    - en ejecución:
        SitioPlanificado -> en_ejecucion

    - revisión:
        SitioPlanificado -> en_ejecucion

    - asignado:
        SitioPlanificado -> planificado

    - sin compromiso operacional:
        aprobado/no requiere -> listo_planificar
        resto               -> pendiente

    De esta manera borrar una planificación NO hace que un
    sitio ejecutado vuelva accidentalmente al pool disponible.
    """

    # ========================================================
    # OBTENER BATCH BLOQUEADO
    # ========================================================

    batch = get_object_or_404(
        BatchPlanificacionSemanal.objects.select_for_update().select_related(
            "planificacion",
            "configuracion_semana",
        ),
        pk=batch_id,
    )

    mensual = batch.planificacion

    configuracion = (
        batch.configuracion_semana if batch.configuracion_semana_id else None
    )

    nombre_batch = batch.nombre or f"Semana {batch.fecha_inicio:%d/%m/%Y}"

    # ========================================================
    # SITIOS DEL BATCH
    # ========================================================
    #
    # Conservamos los objetos SitioPlanificado porque el batch
    # y sus SitioBatchSemanal serán eliminados posteriormente.
    # ========================================================

    items_batch = list(
        SitioBatchSemanal.objects.select_for_update()
        .filter(
            batch=batch,
        )
        .select_related(
            "sitio_planificado",
            "sitio_planificado__sitio",
        )
        .order_by(
            "id",
        )
    )

    # ========================================================
    # ESTADO OPERACIONAL REAL
    # ========================================================
    #
    # Tomamos una fotografía lógica de la situación operacional
    # ANTES de desmontar la planificación.
    # ========================================================

    estados_sitios = []

    for item in items_batch:

        sitio_planificado = item.sitio_planificado

        estado_operacional = obtener_estado_operacional_sitio(
            sitio_planificado,
        )

        estados_sitios.append(
            {
                "sitio_planificado": sitio_planificado,
                "estado_operacional": estado_operacional,
            }
        )

    # ========================================================
    # CONTADORES PARA MENSAJE
    # ========================================================

    cantidad_items_batch = len(
        items_batch,
    )

    cantidad_salidas = SalidaPlanificacionDiaria.objects.filter(
        batch=batch,
    ).count()

    cantidad_participaciones = SitioSalidaPlanificacionDiaria.objects.filter(
        salida__batch=batch,
    ).count()

    cantidad_prioridades = 0

    # ========================================================
    # ELIMINACIÓN PROTEGIDA POR SAVEPOINT
    # ========================================================
    #
    # Si aparece cualquier otra relación PROTECT que todavía
    # no conocemos, TODO este bloque se revierte.
    #
    # No dejamos una eliminación parcial.
    # ========================================================

    try:

        with transaction.atomic():

            # =================================================
            # 1. ELIMINAR PRIORIDADES DIARIAS
            # =================================================
            #
            # PrioridadPlanificacionDiaria está asociada al
            # SitioBatchSemanal mediante prioridad_diaria.
            #
            # La eliminamos explícitamente antes de eliminar
            # los SitioBatchSemanal.
            # =================================================

            for item in items_batch:

                try:

                    prioridad = item.prioridad_diaria

                except ObjectDoesNotExist:

                    prioridad = None

                if prioridad is not None:

                    prioridad.delete()

                    cantidad_prioridades += 1

            # =================================================
            # 2. ELIMINAR PARTICIPACIONES DIARIAS
            # =================================================
            #
            # Este paso es crítico.
            #
            # SitioSalidaPlanificacionDiaria.sitio_batch usa
            # on_delete=PROTECT.
            #
            # Por eso debemos eliminar estas participaciones
            # ANTES de intentar eliminar el batch.
            # =================================================

            SitioSalidaPlanificacionDiaria.objects.filter(
                salida__batch=batch,
            ).delete()

            # =================================================
            # 3. ELIMINAR SALIDAS DIARIAS
            # =================================================

            SalidaPlanificacionDiaria.objects.filter(
                batch=batch,
            ).delete()

            # =================================================
            # 4. ELIMINAR BATCH
            # =================================================
            #
            # Ahora los SitioBatchSemanal pueden desaparecer
            # mediante CASCADE desde BatchPlanificacionSemanal.
            # =================================================

            batch.delete()

            # =================================================
            # 5. ELIMINAR CONFIGURACIÓN SEMANAL HUÉRFANA
            # =================================================

            if configuracion is not None:

                existe_otro_batch = BatchPlanificacionSemanal.objects.filter(
                    configuracion_semana=configuracion,
                ).exists()

                if not existe_otro_batch:

                    configuracion.delete()

    except ProtectedError as exc:

        messages.error(
            request,
            (
                f'No fue posible eliminar "{nombre_batch}" porque '
                "todavía existe información protegida asociada "
                "a esta planificación. "
                "No se eliminó ningún dato."
            ),
        )

        return redirect(
            "planificacion:lista_planificacion_semanal",
            mensual_id=mensual.pk,
        )

    # ========================================================
    # RESTAURAR ESTADO DE SITIOPLANIFICADO
    # ========================================================
    #
    # IMPORTANTE:
    #
    # Ya desapareció la estructura semanal/diaria, pero
    # SitioPlanificado sigue existiendo.
    #
    # Ahora reflejamos la situación REAL de Operaciones.
    # ========================================================

    cantidad_disponibles = 0

    cantidad_planificados = 0

    cantidad_ejecucion = 0

    cantidad_completados = 0

    for registro in estados_sitios:

        sitio_planificado = registro["sitio_planificado"]

        estado_operacional = registro["estado_operacional"]

        # ====================================================
        # ESTADO NORMALIZADO DESDE OPERACIONES
        # ====================================================

        estado_planificacion_operacional = estado_operacional.get(
            "estado_planificacion",
        )

        # ====================================================
        # FINALIZADO
        # ====================================================

        if estado_planificacion_operacional == "finalizado":

            nuevo_estado = "completado"

            cantidad_completados += 1

        # ====================================================
        # EN EJECUCIÓN
        # ====================================================

        elif estado_planificacion_operacional == "en_ejecucion":

            nuevo_estado = "en_ejecucion"

            cantidad_ejecucion += 1

        # ====================================================
        # REVISIÓN
        # ====================================================
        #
        # SitioPlanificado no posee un estado "revision".
        #
        # Sigue operacionalmente comprometido, por eso lo
        # conservamos como en_ejecucion y NO vuelve al pool.
        # ====================================================

        elif estado_planificacion_operacional == "revision":

            nuevo_estado = "en_ejecucion"

            cantidad_ejecucion += 1

        # ====================================================
        # ASIGNADO
        # ====================================================

        elif estado_planificacion_operacional == "asignado":

            nuevo_estado = "planificado"

            cantidad_planificados += 1

        # ====================================================
        # SIN COMPROMISO OPERACIONAL
        # ====================================================

        else:

            if sitio_planificado.estado_permiso in [
                "aprobado",
                "no_requiere",
            ]:

                nuevo_estado = "listo_planificar"

            else:

                nuevo_estado = "pendiente"

            cantidad_disponibles += 1

        # ====================================================
        # LIMPIAR DATOS EXCLUSIVAMENTE DE PLANIFICACIÓN
        # ====================================================

        sitio_planificado.estado = nuevo_estado

        # La fecha anterior pertenecía al batch eliminado.
        sitio_planificado.fecha_planificada = None

        sitio_planificado.orden_dia = 0

        sitio_planificado.bloqueado_motor = False

        sitio_planificado.planificado_manualmente = False

        sitio_planificado.motivo_bloqueo = ""

        sitio_planificado.actualizado_por = request.user

        sitio_planificado.save(
            update_fields=[
                "estado",
                "fecha_planificada",
                "orden_dia",
                "bloqueado_motor",
                "planificado_manualmente",
                "motivo_bloqueo",
                "actualizado_por",
                "actualizado_en",
            ]
        )

    # ========================================================
    # MENSAJE FINAL
    # ========================================================

    messages.success(
        request,
        (
            f'El batch "{nombre_batch}" fue eliminado correctamente. '
            f"Se retiraron {cantidad_salidas} salida(s) diaria(s), "
            f"{cantidad_participaciones} participación(es), "
            f"{cantidad_prioridades} prioridad(es) y "
            f"{cantidad_items_batch} sitio(s) del batch. "
            "Operaciones no fue modificada. "
            f"{cantidad_disponibles} sitio(s) quedaron nuevamente "
            "disponibles para planificación, "
            f"{cantidad_planificados} permanecen comprometidos, "
            f"{cantidad_ejecucion} continúan en ejecución/revisión y "
            f"{cantidad_completados} permanecen completados."
        ),
    )

    return redirect(
        "planificacion:lista_planificacion_semanal",
        mensual_id=mensual.pk,
    )


# ============================================================
# EXCEL DEL BATCH
# ============================================================


@require_GET
@rol_requerido(*ROLES_PLANIFICACION)
def descargar_excel_batch(
    request,
    batch_id,
):
    """
    Genera el archivo Excel que puede enviarse al cliente
    con los sitios correspondientes a una semana.

    Incluye principales y reservas, claramente identificados.
    """

    batch = get_object_or_404(
        BatchPlanificacionSemanal.objects.select_related(
            "planificacion",
        ),
        pk=batch_id,
    )

    items = (
        SitioBatchSemanal.objects.filter(
            batch=batch,
        )
        .exclude(
            estado__in=[
                "excluido",
                "reemplazado",
            ],
        )
        .select_related(
            "sitio_planificado",
            "sitio_planificado__sitio",
        )
        .order_by(
            "es_reserva",
            "cluster_codigo",
            "id",
        )
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Sitios semana"

    # --------------------------------------------------------
    # TÍTULO
    # --------------------------------------------------------

    worksheet.merge_cells("A1:Q1")

    titulo = worksheet["A1"]

    titulo.value = f"Planificación semanal - " f"{batch.fecha_inicio:%d/%m/%Y}"

    titulo.font = Font(
        bold=True,
        size=15,
    )

    titulo.alignment = Alignment(
        vertical="center",
    )

    worksheet.row_dimensions[1].height = 24

    # --------------------------------------------------------
    # DATOS DEL BATCH
    # --------------------------------------------------------

    worksheet["A2"] = "Batch"
    worksheet["B2"] = batch.nombre or f"Semana {batch.fecha_inicio:%d/%m/%Y}"

    worksheet["D2"] = "Objetivo"
    worksheet["E2"] = batch.objetivo_sitios

    worksheet["G2"] = "Estado"
    worksheet["H2"] = batch.get_estado_display()

    # --------------------------------------------------------
    # ENCABEZADOS
    # --------------------------------------------------------

    encabezados = [
        "Tipo",
        "ID Claro",
        "ID Sites",
        "ID Sites NEW",
        "Nombre",
        "Región",
        "Comuna",
        "Dirección",
        "Tipo de Zona",
        "Latitud",
        "Longitud",
        "Condiciones de acceso",
        "Estado permiso",
        "Prioridad",
        "Cluster",
        "Estado batch",
        "Observaciones",
    ]

    fila_encabezado = 4

    for columna, encabezado in enumerate(
        encabezados,
        start=1,
    ):

        celda = worksheet.cell(
            row=fila_encabezado,
            column=columna,
            value=encabezado,
        )

        celda.font = Font(
            bold=True,
        )

        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="E2E8F0",
        )

        celda.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # --------------------------------------------------------
    # FILAS
    # --------------------------------------------------------

    fila = fila_encabezado + 1

    for item in items:

        sitio_planificado = item.sitio_planificado

        sitio = sitio_planificado.sitio

        valores = [
            ("Reserva" if item.es_reserva else "Principal"),
            _valor_sitio(
                sitio,
                "id_claro",
            ),
            _valor_sitio(
                sitio,
                "id_sites",
            ),
            _valor_sitio(
                sitio,
                "id_sites_new",
            ),
            _valor_sitio(
                sitio,
                "nombre",
            ),
            _valor_sitio(
                sitio,
                "region",
            ),
            _valor_sitio(
                sitio,
                "comuna",
            ),
            _valor_sitio(
                sitio,
                "direccion",
            ),
            _valor_sitio(
                sitio,
                "tipo_zona",
            ),
            _valor_sitio(
                sitio,
                "latitud",
            ),
            _valor_sitio(
                sitio,
                "longitud",
            ),
            _valor_sitio(
                sitio,
                "condiciones_acceso",
            ),
            sitio_planificado.get_estado_permiso_display(),
            sitio_planificado.get_prioridad_display(),
            item.cluster_codigo,
            item.get_estado_display(),
            _valor_sitio(
                sitio,
                "observaciones_generales",
            ),
        ]

        for columna, valor in enumerate(
            valores,
            start=1,
        ):

            worksheet.cell(
                row=fila,
                column=columna,
                value=valor,
            )

        fila += 1

    worksheet.freeze_panes = "A5"

    worksheet.auto_filter.ref = f"A4:Q{max(fila - 1, 4)}"

    _ajustar_ancho_columnas_excel(worksheet)

    archivo = BytesIO()

    workbook.save(archivo)

    archivo.seek(0)

    nombre_seguro = batch.nombre or f"semana_{batch.fecha_inicio:%Y_%m_%d}"

    nombre_seguro = nombre_seguro.strip().replace(" ", "_").replace("/", "-")

    response = HttpResponse(
        archivo.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-" "officedocument.spreadsheetml.sheet"
        ),
    )

    response["Content-Disposition"] = f'attachment; filename="' f'{nombre_seguro}.xlsx"'

    return response


# ============================================================
# MAPA GENERAL DEL MES
# ============================================================


@require_GET
@rol_requerido(*ROLES_PLANIFICACION)
def mapa_batches_mensuales(
    request,
    mensual_id,
):
    """
    Muestra las semanas que participan en la planificación
    mensual, incluso si la semana nació desde otro mes.
    """

    mensual = get_object_or_404(
        PlanificacionMensual,
        pk=mensual_id,
    )

    batches = (
        BatchPlanificacionSemanal.objects.filter(
            Q(
                planificaciones_origen=mensual,
            )
            | Q(
                planificacion=mensual,
            )
        )
        .distinct()
        .prefetch_related(
            "planificaciones_origen",
        )
        .order_by(
            "fecha_inicio",
            "id",
        )
    )

    puntos = []

    resumen_batches = []

    for batch in batches:

        items = (
            SitioBatchSemanal.objects.filter(
                batch=batch,
            )
            .exclude(
                estado__in=[
                    "excluido",
                    "reemplazado",
                ],
            )
            .select_related(
                "sitio_planificado",
                "sitio_planificado__sitio",
                "sitio_planificado__planificacion",
            )
        )

        cantidad_puntos = 0

        for item in items:

            sitio_planificado = item.sitio_planificado

            sitio = sitio_planificado.sitio

            if sitio.latitud is None or sitio.longitud is None:
                continue

            try:

                latitud = float(sitio.latitud)

                longitud = float(sitio.longitud)

            except (
                TypeError,
                ValueError,
            ):
                continue

            puntos.append(
                {
                    "batch_id": batch.pk,
                    "batch_nombre": (
                        batch.nombre or ("Semana " f"{batch.fecha_inicio:%d/%m/%Y}")
                    ),
                    "fecha": (batch.fecha_inicio.strftime("%d/%m/%Y")),
                    "id_claro": (sitio.id_claro or ""),
                    "nombre": (sitio.nombre or ""),
                    "comuna": (sitio.comuna or ""),
                    "tipo_zona": (sitio.tipo_zona or ""),
                    "lat": latitud,
                    "lng": longitud,
                    "es_reserva": (item.es_reserva),
                    "cluster": (item.cluster_codigo or ""),
                    "planificacion_origen": str(sitio_planificado.planificacion),
                }
            )

            cantidad_puntos += 1

        resumen_batches.append(
            {
                "batch": batch,
                "cantidad_puntos": cantidad_puntos,
            }
        )

    return render(
        request,
        "planificacion/semanal/mapa_mes.html",
        {
            "mensual": mensual,
            "puntos_mapa": puntos,
            "resumen_batches": resumen_batches,
            "GOOGLE_MAPS_API_KEY": (settings.GOOGLE_MAPS_API_KEY),
        },
    )


# ============================================================
# CREAR / ABRIR BATCH SEMANAL GLOBAL
# ============================================================


@rol_requerido(*ROLES_PLANIFICACION)
def crear_planificacion_semanal(
    request,
    mensual_id,
):
    """
    Crea o abre una semana operacional GLOBAL.

    Una misma semana ISO puede ser utilizada por múltiples
    PlanificacionMensual, pero solo existe un batch.

    IMPORTANTE
    ==========================================================

    El objetivo del batch representa la cantidad operacional
    TOTAL que queremos trabajar durante la semana.

    Por lo tanto:

        objetivo_sitios

    NO está limitado por:

        disponibles_reales

    Ejemplo:

        W35 objetivo:          40
        disponibles agosto:    15

    El batch puede crearse con objetivo 40.

    Posteriormente podrá completarse con sitios provenientes
    de otra planificación mensual, como septiembre.
    """

    mensual = get_object_or_404(
        PlanificacionMensual,
        pk=mensual_id,
    )

    resumen_mensual = obtener_resumen_planificacion_mensual(
        mensual,
    )

    disponibles_reales = int(
        resumen_mensual.get(
            "disponibles_nuevo_batch",
            0,
        )
        or 0
    )

    # ========================================================
    # POST
    # ========================================================

    if request.method == "POST":

        form = CrearBatchSemanalForm(
            request.POST,
            mensual=mensual,
        )

        _configurar_objetivo_formulario(
            form=form,
            disponibles=disponibles_reales,
            es_post=True,
        )

        if form.is_valid():

            datos = form.cleaned_data

            fecha_inicio = datos["fecha_inicio"]

            nombre_batch = form.obtener_nombre_batch()

            objetivo_solicitado = int(datos["objetivo_sitios"])

            # =================================================
            # ¿YA EXISTE GLOBALMENTE?
            # =================================================

            batch_existente = form.obtener_batch_existente()

            # =================================================
            # ABRIR / VINCULAR EXISTENTE
            # =================================================

            if batch_existente is not None:

                with transaction.atomic():

                    batch = (
                        BatchPlanificacionSemanal.objects.select_for_update()
                        .select_related(
                            "planificacion",
                            "configuracion_semana",
                        )
                        .get(
                            pk=batch_existente.pk,
                        )
                    )

                    ya_vinculado = batch.planificaciones_origen.filter(
                        pk=mensual.pk,
                    ).exists()

                    if not ya_vinculado:

                        batch.planificaciones_origen.add(
                            mensual,
                        )

                    # =========================================
                    # COMPATIBILIDAD LEGACY
                    # =========================================

                    if batch.planificacion_id is None:

                        batch.planificacion = mensual

                        batch.actualizado_por = request.user

                        batch.save(
                            update_fields=[
                                "planificacion",
                                "actualizado_por",
                                "actualizado_en",
                            ]
                        )

                    # =========================================
                    # CONFIGURACIÓN LEGACY
                    # =========================================

                    if (
                        batch.configuracion_semana_id
                        and batch.configuracion_semana.planificacion_id is None
                    ):

                        configuracion = batch.configuracion_semana

                        configuracion.planificacion = mensual

                        configuracion.actualizado_por = request.user

                        configuracion.save(
                            update_fields=[
                                "planificacion",
                                "actualizado_por",
                                "actualizado_en",
                            ]
                        )

                if ya_vinculado:

                    messages.info(
                        request,
                        (
                            f"{nombre_batch} ya estaba "
                            "vinculada a esta planificación. "
                            "Se abrió la semana operacional."
                        ),
                    )

                else:

                    messages.success(
                        request,
                        (
                            f"{nombre_batch} ya existía. "
                            f"Se vinculó {mensual} a la "
                            "semana operacional."
                        ),
                    )

                return redirect(
                    "planificacion:" "detalle_planificacion_semanal",
                    batch_id=batch.pk,
                )

            # =================================================
            # NUEVA SEMANA
            # =================================================
            #
            # Ya NO comparamos:
            #
            #     objetivo_solicitado
            #         vs
            #     disponibles_reales
            #
            # porque son conceptos distintos.
            #
            # objetivo_solicitado:
            #     objetivo operacional semanal.
            #
            # disponibles_reales:
            #     sitios que este mes puede aportar ahora.
            # =================================================

            if disponibles_reales <= 0:

                form.add_error(
                    "objetivo_sitios",
                    (
                        "No quedan sitios disponibles "
                        "en esta planificación mensual "
                        "para iniciar una semana nueva."
                    ),
                )

            else:

                disponibilidades = form.obtener_disponibilidades()

                try:

                    with transaction.atomic():

                        # =====================================
                        # SEGUNDA COMPROBACIÓN GLOBAL
                        # =====================================

                        batch_concurrente = BatchPlanificacionSemanal.objects.filter(
                            fecha_inicio=fecha_inicio,
                        ).first()

                        if batch_concurrente is not None:

                            batch = BatchPlanificacionSemanal.objects.select_for_update().get(
                                pk=batch_concurrente.pk,
                            )

                            batch.planificaciones_origen.add(
                                mensual,
                            )

                            if batch.planificacion_id is None:

                                batch.planificacion = mensual

                                batch.actualizado_por = request.user

                                batch.save(
                                    update_fields=[
                                        "planificacion",
                                        "actualizado_por",
                                        "actualizado_en",
                                    ]
                                )

                            creada = False

                        else:

                            batch = crear_batch_semanal(
                                planificacion=mensual,
                                fecha_inicio=fecha_inicio,
                                objetivo_sitios=(objetivo_solicitado),
                                nombre=nombre_batch,
                                observaciones=(
                                    datos.get(
                                        "observaciones",
                                        "",
                                    )
                                ),
                                disponibilidades=(disponibilidades),
                                usuario=request.user,
                            )

                            # crear_batch_semanal ya agrega
                            # planificaciones_origen.
                            creada = True

                except ValueError as exc:

                    messages.error(
                        request,
                        str(exc),
                    )

                else:

                    if creada:

                        if objetivo_solicitado > disponibles_reales:

                            faltantes_objetivo = (
                                objetivo_solicitado - disponibles_reales
                            )

                            messages.success(
                                request,
                                (
                                    f"{nombre_batch} fue creada "
                                    "correctamente como semana "
                                    "operacional global. "
                                    f"Objetivo: "
                                    f"{objetivo_solicitado} sitio(s). "
                                    f"Esta planificación mensual "
                                    f"dispone actualmente de "
                                    f"{disponibles_reales} sitio(s), "
                                    f"por lo que quedan "
                                    f"{faltantes_objetivo} sitio(s) "
                                    "por completar desde otras "
                                    "planificaciones mensuales "
                                    "si fuese necesario."
                                ),
                            )

                        else:

                            messages.success(
                                request,
                                (
                                    f"{nombre_batch} fue creada "
                                    "correctamente como semana "
                                    "operacional global."
                                ),
                            )

                    else:

                        messages.info(
                            request,
                            (
                                f"{nombre_batch} ya había "
                                "sido creada. "
                                "Se vinculó esta planificación "
                                "mensual a la semana existente."
                            ),
                        )

                    return redirect(
                        "planificacion:" "detalle_planificacion_semanal",
                        batch_id=batch.pk,
                    )

    # ========================================================
    # GET
    # ========================================================

    else:

        form = CrearBatchSemanalForm(
            mensual=mensual,
        )

        _configurar_objetivo_formulario(
            form=form,
            disponibles=disponibles_reales,
            es_post=False,
        )

    return render(
        request,
        "planificacion/semanal/crear.html",
        {
            "mensual": mensual,
            "form": form,
            "resumen_mensual": (resumen_mensual),
            "disponibles_reales": (disponibles_reales),
            "semanas_disponibles": (form.semanas_disponibles),
            "semanas_ocupadas": (form.semanas_ocupadas),
        },
    )


# ============================================================
# DETALLE DEL BATCH SEMANAL
# ============================================================


@rol_requerido(*ROLES_PLANIFICACION)
def detalle_planificacion_semanal(
    request,
    batch_id,
):
    batch = get_object_or_404(
        BatchPlanificacionSemanal.objects.select_related(
            "planificacion",
            "configuracion_semana",
        ).prefetch_related(
            "planificaciones_origen",
        ),
        pk=batch_id,
    )

    # ========================================================
    # MESES PARTICIPANTES
    # ========================================================

    mensuales_origen = list(
        batch.planificaciones_origen.order_by(
            "anio",
            "mes",
        )
    )

    # ========================================================
    # CONTEXTO MENSUAL COMPATIBLE
    # ========================================================

    mensual = batch.planificacion

    if mensual is None and mensuales_origen:

        mensual = mensuales_origen[0]

    incluidos = (
        SitioBatchSemanal.objects.filter(
            batch=batch,
        )
        .select_related(
            "sitio_planificado",
            "sitio_planificado__sitio",
            "sitio_planificado__planificacion",
        )
        .order_by(
            "es_reserva",
            "-puntaje_motor",
            "id",
        )
    )

    candidatos_base = obtener_candidatos_batch(batch)

    # ========================================================
    # FILTROS
    # ========================================================

    comuna = request.GET.get(
        "comuna",
        "",
    ).strip()

    tipo_zona = request.GET.get(
        "tipo_zona",
        "",
    ).strip()

    permiso = request.GET.get(
        "permiso",
        "",
    ).strip()

    prioridad = request.GET.get(
        "prioridad",
        "",
    ).strip()

    busqueda = request.GET.get(
        "q",
        "",
    ).strip()

    candidatos = candidatos_base

    if comuna:

        candidatos = candidatos.filter(
            sitio__comuna__iexact=comuna,
        )

    if tipo_zona:

        candidatos = candidatos.filter(
            sitio__tipo_zona__iexact=tipo_zona,
        )

    if permiso:

        candidatos = candidatos.filter(
            estado_permiso=permiso,
        )

    if prioridad:

        candidatos = candidatos.filter(
            prioridad=prioridad,
        )

    if busqueda:

        candidatos = candidatos.filter(
            sitio__id_claro__icontains=busqueda,
        )

    comunas = (
        candidatos_base.exclude(
            sitio__comuna__isnull=True,
        )
        .exclude(
            sitio__comuna="",
        )
        .values_list(
            "sitio__comuna",
            flat=True,
        )
        .distinct()
        .order_by(
            "sitio__comuna",
        )
    )

    tipos_zona = (
        candidatos_base.exclude(
            sitio__tipo_zona__isnull=True,
        )
        .exclude(
            sitio__tipo_zona="",
        )
        .values_list(
            "sitio__tipo_zona",
            flat=True,
        )
        .distinct()
        .order_by(
            "sitio__tipo_zona",
        )
    )

    # ========================================================
    # DISPONIBILIDADES
    # ========================================================

    disponibilidades = []

    if batch.configuracion_semana_id:

        disponibilidades = (
            batch.configuracion_semana.disponibilidades_cuadrillas.select_related(
                "cuadrilla_operativa",
            ).order_by(
                "cuadrilla_operativa__orden",
                "cuadrilla_operativa__nombre",
                "cuadrilla",
                "id",
            )
        )

    resumen = obtener_resumen_batch(batch)

    return render(
        request,
        "planificacion/semanal/detalle.html",
        {
            "batch": batch,
            # Compatibilidad con template actual.
            "mensual": mensual,
            # Nueva arquitectura.
            "mensuales_origen": (mensuales_origen),
            "incluidos": incluidos,
            "candidatos": candidatos[:200],
            "comunas": comunas,
            "tipos_zona": tipos_zona,
            "estados_permiso": (SitioPlanificado.ESTADOS_PERMISO),
            "prioridades": (SitioPlanificado.PRIORIDADES),
            "disponibilidades": disponibilidades,
            "resumen": resumen,
            "filtros": {
                "comuna": comuna,
                "tipo_zona": tipo_zona,
                "permiso": permiso,
                "prioridad": prioridad,
                "q": busqueda,
            },
        },
    )


# ============================================================
# AGREGAR SITIOS AL BATCH
# ============================================================


@require_POST
@rol_requerido(*ROLES_PLANIFICACION)
def agregar_sitios_batch(
    request,
    batch_id,
):
    batch = get_object_or_404(
        BatchPlanificacionSemanal,
        pk=batch_id,
    )

    if batch.estado not in [
        "borrador",
        "propuesto",
        "gestion_permisos",
    ]:

        messages.error(
            request,
            (
                "No se pueden agregar sitios a un batch "
                "que ya se encuentra en esta etapa."
            ),
        )

        return redirect(
            "planificacion:detalle_planificacion_semanal",
            batch_id=batch.pk,
        )

    sitio_ids = request.POST.getlist(
        "sitio_ids",
    )

    if not sitio_ids:

        messages.warning(
            request,
            "Debes seleccionar al menos un sitio.",
        )

        return redirect(
            "planificacion:detalle_planificacion_semanal",
            batch_id=batch.pk,
        )

    es_reserva = request.POST.get("es_reserva") == "1"

    cantidad = agregar_sitios_al_batch(
        batch=batch,
        sitio_ids=sitio_ids,
        usuario=request.user,
        es_reserva=es_reserva,
    )

    if cantidad:

        messages.success(
            request,
            (f"{cantidad} sitio(s) " "agregado(s) al batch."),
        )

    else:

        messages.warning(
            request,
            "No se agregó ningún sitio al batch.",
        )

    return redirect(
        "planificacion:detalle_planificacion_semanal",
        batch_id=batch.pk,
    )


# ============================================================
# QUITAR SITIO DEL BATCH
# ============================================================


@require_POST
@rol_requerido(*ROLES_PLANIFICACION)
def quitar_sitio_batch(
    request,
    item_id,
):
    item = get_object_or_404(
        SitioBatchSemanal.objects.select_related(
            "batch",
        ),
        pk=item_id,
    )

    batch = item.batch

    motivo = request.POST.get(
        "motivo",
        "",
    ).strip()

    quitar_sitio_del_batch(
        item_batch=item,
        usuario=request.user,
        motivo=motivo,
    )

    messages.success(
        request,
        "Sitio retirado del batch.",
    )

    return redirect(
        "planificacion:detalle_planificacion_semanal",
        batch_id=batch.pk,
    )


# ============================================================
# CERRAR PROPUESTA
# ============================================================


@require_POST
@rol_requerido(*ROLES_PLANIFICACION)
def cerrar_propuesta_semanal(
    request,
    batch_id,
):
    batch = get_object_or_404(
        BatchPlanificacionSemanal,
        pk=batch_id,
    )

    try:

        cerrar_propuesta_batch(
            batch=batch,
            usuario=request.user,
        )

    except ValueError as exc:

        messages.error(
            request,
            str(exc),
        )

    else:

        messages.success(
            request,
            ("La selección quedó cerrada " "como propuesta semanal."),
        )

    return redirect(
        "planificacion:detalle_planificacion_semanal",
        batch_id=batch.pk,
    )


# ============================================================
# ENVIAR A GESTIÓN DE PERMISOS
# ============================================================


@require_POST
@rol_requerido(*ROLES_PLANIFICACION)
def enviar_gestion_permisos_batch(
    request,
    batch_id,
):
    batch = get_object_or_404(
        BatchPlanificacionSemanal,
        pk=batch_id,
    )

    try:

        marcar_gestion_permisos_enviada(
            batch=batch,
            usuario=request.user,
        )

    except ValueError as exc:

        messages.error(
            request,
            str(exc),
        )

    else:

        messages.success(
            request,
            ("La propuesta quedó marcada como enviada " "para gestión de permisos."),
        )

    return redirect(
        "planificacion:detalle_planificacion_semanal",
        batch_id=batch.pk,
    )


# ============================================================
# ACTUALIZAR PERMISO
# ============================================================


@require_POST
@rol_requerido(*ROLES_PLANIFICACION)
def actualizar_permiso_sitio_batch(
    request,
    item_id,
):
    item = get_object_or_404(
        SitioBatchSemanal.objects.select_related(
            "sitio_planificado",
        ),
        pk=item_id,
    )

    nuevo_permiso = request.POST.get(
        "estado_permiso",
        "",
    ).strip()

    estados_validos = {valor for valor, _ in SitioPlanificado.ESTADOS_PERMISO}

    if nuevo_permiso not in estados_validos:

        return JsonResponse(
            {
                "ok": False,
                "error": "Estado de permiso inválido.",
            },
            status=400,
        )

    try:

        actualizar_permiso_desde_batch(
            item=item,
            nuevo_permiso=nuevo_permiso,
            usuario=request.user,
        )

    except ValueError as exc:

        return JsonResponse(
            {
                "ok": False,
                "error": str(exc),
            },
            status=400,
        )

    item.refresh_from_db()

    item.sitio_planificado.refresh_from_db()

    return JsonResponse(
        {
            "ok": True,
            "estado_permiso": (item.sitio_planificado.estado_permiso),
            "estado_permiso_display": (
                item.sitio_planificado.get_estado_permiso_display()
            ),
            "estado_sitio": (item.sitio_planificado.estado),
            "estado_sitio_display": (item.sitio_planificado.get_estado_display()),
            "estado_batch": item.estado,
            "estado_batch_display": (item.get_estado_display()),
        }
    )


# ============================================================
# CONFIRMAR PARA PLANIFICACIÓN DIARIA
# ============================================================


@require_POST
@rol_requerido(*ROLES_PLANIFICACION)
def confirmar_sitios_batch(
    request,
    batch_id,
):
    batch = get_object_or_404(
        BatchPlanificacionSemanal,
        pk=batch_id,
    )

    item_ids = request.POST.getlist(
        "item_ids",
    )

    if not item_ids:

        messages.warning(
            request,
            ("Debes seleccionar al menos un sitio " "disponible para confirmar."),
        )

        return redirect(
            "planificacion:detalle_planificacion_semanal",
            batch_id=batch.pk,
        )

    try:

        cantidad = confirmar_sitios_para_planificacion(
            batch=batch,
            item_ids=item_ids,
            usuario=request.user,
        )

    except ValueError as exc:

        messages.error(
            request,
            str(exc),
        )

    else:

        messages.success(
            request,
            (f"{cantidad} sitio(s) quedaron confirmados " "para planificación diaria."),
        )

    return redirect(
        "planificacion:detalle_planificacion_semanal",
        batch_id=batch.pk,
    )

# ============================================================
# ACTUALIZAR ESTADO MASIVO DE SITIOS DEL BATCH
# ============================================================


@require_POST
@rol_requerido(*ROLES_PLANIFICACION)
@transaction.atomic
def actualizar_estado_masivo_sitios_batch(
    request,
    batch_id,
):
    """
    Actualiza de forma masiva el estado de varios
    SitioBatchSemanal pertenecientes al mismo batch.

    IMPORTANTE
    ==========================================================

    Esta acción modifica:

        SitioBatchSemanal.estado

    y mantiene sincronizado SitioPlanificado cuando
    corresponde.

    No modifica Operaciones.

    REGLAS PRINCIPALES
    ==========================================================

    disponible:
        requiere permiso aprobado o no_requiere.
        SitioPlanificado -> listo_planificar

    gestion_permiso:
        SitioPlanificado -> gestionando_permiso

    confirmado:
        requiere permiso aprobado o no_requiere.
        SitioPlanificado -> listo_planificar

    rechazado:
        SitioPlanificado -> bloqueado

    seleccionado:
        SitioPlanificado -> pendiente
        salvo que ya tenga permiso aprobado/no_requiere,
        en cuyo caso queda listo_planificar.

    sin_respuesta:
        SitioPlanificado -> gestionando_permiso
    """

    # ========================================================
    # BATCH
    # ========================================================

    batch = get_object_or_404(
        BatchPlanificacionSemanal.objects.select_for_update(),
        pk=batch_id,
    )

    # ========================================================
    # IDS
    # ========================================================

    item_ids = request.POST.getlist(
        "item_ids",
    )

    if not item_ids:

        messages.warning(
            request,
            "Debes seleccionar al menos un sitio.",
        )

        return redirect(
            "planificacion:detalle_planificacion_semanal",
            batch_id=batch.pk,
        )

    # ========================================================
    # NUEVO ESTADO
    # ========================================================

    nuevo_estado = (
        request.POST.get(
            "nuevo_estado_batch",
            "",
        )
        or ""
    ).strip()

    estados_permitidos = {
        "seleccionado",
        "gestion_permiso",
        "disponible",
        "confirmado",
        "rechazado",
        "sin_respuesta",
    }

    if nuevo_estado not in estados_permitidos:

        messages.error(
            request,
            "El estado seleccionado no es válido.",
        )

        return redirect(
            "planificacion:detalle_planificacion_semanal",
            batch_id=batch.pk,
        )

    # ========================================================
    # ITEMS DEL BATCH
    # ========================================================

    items = list(
        SitioBatchSemanal.objects.select_for_update()
        .filter(
            batch=batch,
            id__in=item_ids,
        )
        .select_related(
            "sitio_planificado",
            "sitio_planificado__sitio",
        )
        .order_by(
            "id",
        )
    )

    if not items:

        messages.warning(
            request,
            "No se encontraron sitios válidos para actualizar.",
        )

        return redirect(
            "planificacion:detalle_planificacion_semanal",
            batch_id=batch.pk,
        )

    # ========================================================
    # CONTADORES
    # ========================================================

    actualizados = 0

    omitidos = 0

    errores = []

    # ========================================================
    # PROCESAR
    # ========================================================

    for item in items:

        sitio_planificado = item.sitio_planificado

        identificador = (
            sitio_planificado.sitio.id_claro
            or sitio_planificado.sitio.id_sites
            or f"Sitio {sitio_planificado.sitio_id}"
        )

        # ====================================================
        # NO TOCAR ESTADOS RETIRADOS
        # ====================================================

        if item.estado in {
            "excluido",
            "reemplazado",
        }:

            omitidos += 1

            continue

        permiso_utilizable = sitio_planificado.estado_permiso in {
            "aprobado",
            "no_requiere",
        }

        # ====================================================
        # DISPONIBLE
        # ====================================================

        if nuevo_estado == "disponible":

            if not permiso_utilizable:

                omitidos += 1

                errores.append(
                    (
                        f"{identificador}: no puede pasar a "
                        "Disponible porque el permiso no está "
                        "aprobado ni marcado como No requiere."
                    )
                )

                continue

            item.estado = "disponible"

            if sitio_planificado.estado not in {
                "completado",
                "cancelado",
                "bloqueado",
                "en_ejecucion",
                "en_ruta",
            }:

                sitio_planificado.estado = "listo_planificar"

        # ====================================================
        # CONFIRMADO
        # ====================================================

        elif nuevo_estado == "confirmado":

            if not permiso_utilizable:

                omitidos += 1

                errores.append(
                    (
                        f"{identificador}: no puede pasar a "
                        "Confirmado porque el permiso no está "
                        "aprobado ni marcado como No requiere."
                    )
                )

                continue

            item.estado = "confirmado"

            if sitio_planificado.estado not in {
                "completado",
                "cancelado",
                "bloqueado",
                "en_ejecucion",
                "en_ruta",
            }:

                sitio_planificado.estado = "listo_planificar"

        # ====================================================
        # GESTIÓN DE PERMISO
        # ====================================================

        elif nuevo_estado == "gestion_permiso":

            item.estado = "gestion_permiso"

            if sitio_planificado.estado not in {
                "completado",
                "cancelado",
                "bloqueado",
                "en_ejecucion",
                "en_ruta",
            }:

                sitio_planificado.estado = "gestionando_permiso"

        # ====================================================
        # RECHAZADO
        # ====================================================

        elif nuevo_estado == "rechazado":

            item.estado = "rechazado"

            sitio_planificado.estado_permiso = "rechazado"

            if sitio_planificado.estado not in {
                "completado",
                "cancelado",
                "en_ejecucion",
                "en_ruta",
            }:

                sitio_planificado.estado = "bloqueado"

        # ====================================================
        # SIN RESPUESTA
        # ====================================================

        elif nuevo_estado == "sin_respuesta":

            item.estado = "sin_respuesta"

            if sitio_planificado.estado not in {
                "completado",
                "cancelado",
                "bloqueado",
                "en_ejecucion",
                "en_ruta",
            }:

                sitio_planificado.estado = "gestionando_permiso"

        # ====================================================
        # SELECCIONADO
        # ====================================================

        elif nuevo_estado == "seleccionado":

            item.estado = "seleccionado"

            if sitio_planificado.estado not in {
                "completado",
                "cancelado",
                "bloqueado",
                "en_ejecucion",
                "en_ruta",
            }:

                if permiso_utilizable:

                    sitio_planificado.estado = "listo_planificar"

                else:

                    sitio_planificado.estado = "pendiente"

        # ====================================================
        # GUARDAR SITIO PLANIFICADO
        # ====================================================

        sitio_planificado.actualizado_por = request.user

        sitio_planificado.save(
            update_fields=[
                "estado_permiso",
                "estado",
                "actualizado_por",
                "actualizado_en",
            ]
        )

        # ====================================================
        # GUARDAR ITEM BATCH
        # ====================================================

        item.save(
            update_fields=[
                "estado",
                "actualizado_en",
            ]
        )

        actualizados += 1

    # ========================================================
    # ACTUALIZAR ESTADO GENERAL DEL BATCH SI CORRESPONDE
    # ========================================================

    if nuevo_estado == "confirmado":

        tiene_disponibles_sin_confirmar = batch.sitios.filter(
            estado="disponible",
        ).exists()

        if not tiene_disponibles_sin_confirmar:

            batch.estado = "listo_planificar"

            batch.actualizado_por = request.user

            batch.save(
                update_fields=[
                    "estado",
                    "actualizado_por",
                    "actualizado_en",
                ]
            )

    # ========================================================
    # MENSAJES
    # ========================================================

    if actualizados:

        messages.success(
            request,
            (f"{actualizados} sitio(s) fueron actualizados " "correctamente."),
        )

    if omitidos:

        messages.warning(
            request,
            (f"{omitidos} sitio(s) no pudieron cambiar " "al estado solicitado."),
        )

    for error in errores[:10]:

        messages.warning(
            request,
            error,
        )

    if len(errores) > 10:

        messages.warning(
            request,
            (f"Existen {len(errores) - 10} advertencia(s) " "adicionales."),
        )

    # ========================================================
    # REDIRECT
    # ========================================================

    return redirect(
        "planificacion:detalle_planificacion_semanal",
        batch_id=batch.pk,
    )
